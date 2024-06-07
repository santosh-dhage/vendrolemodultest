from .models import *
from .serializers import *
from rest_framework.response import Response
from rest_framework import status
import logging
from rest_framework import serializers
from rest_framework import viewsets
from .models import MachineMaster

from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from django.db.models import Q
from django.core.exceptions import ValidationError

class MachineMasterViewSet(viewsets.ModelViewSet):
    queryset = MachineMaster.objects.all()
    serializer_class = MachineMasterSerializer
    permission_classes=[IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            queryset = MachineMaster.objects.all().order_by('-id')
            serializer = MachineMasterSerializer(queryset, many=True)
            return Response({'success': 1, 'message': 'Manage Membership List', 'result': serializer.data})
        except Exception as e:
            return Response({'success': 0, 'message': 'Not Found', 'result': str(e)})

    def retrieve(self, request, pk, *args, **kwargs):
        try:
            chp = MachineMaster.objects.get(pk=pk)
            serializer = MachineMasterSerializer(chp)
            return Response({'success': 1, 'message': 'Manage Membership', 'result': serializer.data})
        except MachineMaster.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def create(self, request, *args, **kwargs):
        try:
            serializer = MachineMasterSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            
            return Response({'success': 1, 'message': 'Data Created', 'result': serializer.data})
        except ValidationError as ve:
            return Response({'success': 0, 'message': 'Not Created', 'result': ve.detail})

    def update(self, request, pk, *args, **kwargs):
        try:
            chp = MachineMaster.objects.get(pk=pk)
            if 'rqcode' not in request.data:
                return Response({'success': 0, 'message': 'rqcode is required'}, status=status.HTTP_400_BAD_REQUEST)
            serializer = MachineMasterSerializer(chp, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})
        except MachineMaster.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def destroy(self, request, pk, *args, **kwargs):
        try:
            machine = MachineMaster.objects.get(pk=pk)

            # Ensure the machine belongs to the requesting tenant
            if machine.tenant != request.tenant:
                return Response({'success': 0, 'message': 'You do not have permission to delete this machine'}, 
                                status=status.HTTP_403_FORBIDDEN)

            machine.delete()
            # chp = MachineMaster.objects.get(pk=pk)
            # chp.delete()
            return Response({'success': 1, 'message': 'Data Deleted'})
        except MachineMaster.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})
from django.http import JsonResponse
from django.db.models import Count

def get_created_machine_count(request):
    if request.method=='GET':
        data=MachineMaster.objects.count()
   
        return JsonResponse({'machine_count_all': data})
        
 
    


@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def create_mapping_user(request):
    try:
        if request.method == 'POST':
            # Expect user_email and machine_id in the request data
            user_email = request.data.get('user_email')
            machine_id = request.data.get('machine_id')

            user = User.objects.get(email=user_email)
            machine = MachineMaster.objects.get(machine_id=machine_id)

            # Check if a mapping already exists
            existing_mapping = MachineUserMapping.objects.filter(machine=machine, assigned_user=user)

            if not existing_mapping.exists():
                # Create a new mapping if it doesn't exist
                mapping = MachineUserMapping.objects.create(machine=machine, assigned_user=user)
                serializer = MachineUserMappingSerializer(mapping)
                return Response({'success': 1, 'result': serializer.data}, status=status.HTTP_201_CREATED)
            else:
                return Response({'success': 0, 'result': 'Mapping already exists'}, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'GET':
            users = User.objects.all().order_by('-id')
            machines = MachineMaster.objects.all().order_by('-id')
            mappings_data = []

            for user in users:
                for machine in machines:
                    # Check if a mapping already exists
                    existing_mapping = MachineUserMapping.objects.filter(machine=machine, assigned_user=user)

                    if not existing_mapping.exists():
                        # Create a new mapping if it doesn't exist
                        mapping = MachineUserMapping.objects.create(machine=machine, assigned_user=user)
                        serializer = MachineUserMappingSerializer(mapping)
                        mappings_data.append(serializer.data)
                    else:
                        # Include existing mappings in the response
                        serializer = MachineUserMappingSerializer(existing_mapping.first())
                        mappings_data.append(serializer.data)

            return Response({'success': 1, 'result': mappings_data}, status=status.HTTP_201_CREATED)

    except User.DoesNotExist:
        return Response({'success': 0, 'result': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    except MachineMaster.DoesNotExist:
        return Response({'success': 0, 'result': 'Machine not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'success': 0, 'result': f'An error occurred: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class QrCodeViewSet(viewsets.ModelViewSet):
    queryset = QrCode.objects.all()
    serializer_class = QrCodeSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            queryset = QrCode.objects.all().order_by('-id')
            serializer = QrCodeSerializer(queryset, many=True)
            
            return Response({'success': 1, 'message': 'Machine master List', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error listing Machine master: {str(e)}')
            return Response({'success': 0, 'message': 'Not Found'})

    def retrieve(self, request, pk, *args, **kwargs):
        try:
            user = QrCode.objects.get(pk=pk)
            serializer = QrCodeSerializer(user)
            return Response({'success': 1, 'message': 'Machine master', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error retrieving Machine master: {str(e)}')
            return Response({'success': 0, 'message': 'Not Found'})

    def create(self, request, *args, **kwargs):
        try:
            serializer = QrCodeSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({'success': 1, 'message': 'Machine master created successfully', 'result': serializer.data}, status=status.HTTP_201_CREATED)
        except serializers.ValidationError as ve:
            return Response({'success': 0, 'message': ve.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logging.error(f'Error creating MachineMaster: {str(e)}')
            return Response({'success': 0, 'message': 'An error occurred while creating Machine master'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, pk,*args, **kwargs):
        try:
            instance = QrCode.objects.get(pk=pk)
            serializer = QrCodeSerializer(instance, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({'success': 1, 'message': 'Machine master updated successfully', 'result': serializer.data})
        except serializers.ValidationError as ve:
            return Response({'success': 0, 'message': ve.detail}, status=status.HTTP_400_BAD_REQUEST)
        except MachineMaster.DoesNotExist:
            return Response({'success': 0, 'message': 'Machine master not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logging.error(f'Error updating MachineMaster: {str(e)}')
            return Response({'success': 0, 'message': 'An error occurred while updating Machine master'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, pk,*args, **kwargs):
        try:
            instance = QrCode.objects.get(pk=pk)
            instance.delete()
            return Response({'success': 1, 'message': 'Machine master deleted successfully'})
        except MachineMaster.DoesNotExist:
            return Response({'success': 0, 'message': 'Machine master not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logging.error(f'Error deleting MachineMaster: {str(e)}')
            return Response({'success': 0, 'message': 'An error occurred while deleting Machine master'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MachineUserMappingViewSet(viewsets.ModelViewSet):
    queryset = MachineUserMapping.objects.all()
    serializer_class = MachineUserMappingSerializer
    # permission_classes = [IsAuthenticated]
    def list(self, request, *args, **kwargs):
        try:
            queryset = MachineUserMapping.objects.all().order_by('-id')
            serializer = MachineUserMappingSerializer(queryset, many=True)
            logging.info('List request processed successfully.')
            return Response({'success': 1, 'message': 'Machine User Mapping list', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing list request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, pk,*args, **kwargs):
        try:
            instance = MachineUserMapping.objects.get(pk=pk)
            serializer = MachineUserMappingSerializer(instance)
            logging.info('Retrieve request processed successfully.')
            return Response({'success': 1, 'message': 'Mapping details', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing retrieve request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def create(self, request, *args, **kwargs):
        # try:
            serializer = MachineUserMappingSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save(created_by=request.user.id)
            logging.info('Create request processed successfully. Mapping created.')
            return Response({'success': 1, 'message': 'Mapping created successfully', 'result': serializer.data}, status=status.HTTP_201_CREATED)
        # except Exception as e:
        #     logging.error(f'Error processing create request: {e}')
        #     return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, pk,**kwargs):
        try:
            instance = MachineUserMapping.objects.get(pk=pk)
            serializer = MachineUserMappingSerializer(instance, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            logging.info('Update request processed successfully. Mapping updated.')
            return Response({'success': 1, 'message': 'Mapping updated successfully', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing update request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, pk,*args, **kwargs):
        try:
            instance = MachineUserMapping.objects.get(pk=pk)
            instance.delete()
            logging.info('Delete request processed successfully. Mapping deleted.')
            return Response({'success': 1, 'message': 'Mapping deleted successfully'})
        except Exception as e:
            logging.error(f'Error processing delete request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_user_map_by_user(request):
    if request.method== 'GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
            serializer = MachineUserMapping1Serializer(machine_user_mappings, many=True)
            return Response({'success':1,'message':'Data Found','result':serializer.data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success':0,'message':'Not Found','result':serializer.errors})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_user_map_by_customer(request):
    if request.method=='GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer=request.user.id , assigned_user__isnull=False).order_by('-id')
            for machine_id in machine_user_mappings:
                print(machine_id)
            serializer = MachineUserMapping1Serializer(machine_user_mappings, many=True)

            return Response({'success':1,'message':'Data Found','result':serializer.data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response(f'Error {e}')
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_and_qrcode(request):
    if request.method=='GET':
        try:
            machine_user_mappings = MachineMaster.objects.all().order_by('-id')
            data =[]
            for machine in machine_user_mappings:
                # print(machine)
                if machine.rqcode:
                    qr_code = QrCode.objects.filter(id=machine.rqcode.id)
                    if len(qr_code) !=0 and qr_code[0].qr:
                        option = {
                            "id" : machine.id,
                            "machine_id" :machine.machine_id,
                            'rqcode' :qr_code[0].qr_code_id,
                            'location':machine.installation_location,
                            'product_type':machine.product.product_type if machine.product else None,
                            'model_no':machine.model_number.model_no if machine.model_number else None,
                            'name':machine.model_number.name if machine.model_number else None,
                            'amount':machine.product.amount if machine.product else None,
                            'machinelease':machine.machinelease if machine else None,
                            'qr_code_img':qr_code[0].qr.url,
                            'created_at':machine.created_at,
                            # 'created_by':machine.created_by,
                            'is_active':machine.is_active,
                            'payment_type':machine.payment_type,
                        }
                        data.append(option)
                    else:
                        option = {
                            "id" : machine.id,
                            "machine_id" :machine.machine_id,
                            'rqcode' : '',
                            'qr_code_img':'',
                            'location':machine.installation_location,
                            'product_type':machine.product.product_type if machine.product else None,
                            'model_no':machine.model_number.model_no if machine.model_number else None,
                            'name':machine.model_number.name if machine.model_number else None,
                            'amount':machine.product.amount if machine.product else None,
                            'machinelease':machine.machinelease if machine else None,
                            'created_at':machine.created_at,
                            # 'created_by':machine.created_by,
                            'is_active':machine.is_active,
                            'payment_type':machine.payment_type,
                        }
                       
                        data.append(option)
                else:
                    option = {
                        "id" : machine.id,
                       
                        "machine_id" :machine.machine_id,
                        'rqcode' : '',
                        'qr_code_img':'',
                        'location':machine.installation_location,
                        'product_type':machine.product.product_type if machine.product else None,
                        'model_no':machine.model_number.model_no if machine.model_number else None,
                        'name':machine.model_number.name if machine.model_number else None,
                        'amount':machine.product.amount if machine.product else None,
                        'machinelease':machine.machinelease if machine else None,
                        'created_at':machine.created_at,
                        # 'created_by':machine.created_by,
                        'is_active':machine.is_active,
                        'payment_type':machine.payment_type,
                    }
                    data.append(option)
                # print(machine_id)
            # serializer = MachineUserMappingSerializer(machine_user_mappings, many=True)
            return Response({'success':1,'message':'Data Found','result':data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response(f'Error {e}')
        
# from django.core.exceptions import ObjectDoesNotExist
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_machine_and_qrcode(request):
#     if request.method == 'GET':
#         try:
#             machine_user_mappings = MachineMaster.objects.all().order_by('-id')
#             data = []

#             for machine in machine_user_mappings:
#                 option = {
#                     "id": machine.id,
#                     "machine_id": machine.machine_id,
#                     'rqcode': machine.rqcode.qr_code_id if machine.rqcode else None,
#                     'qr_code_img': '',
#                     'location': machine.installation_location,
#                     'product_type': machine.product.product_type if machine.product else None,
#                     'model_no': machine.model_number.model_no if machine.model_number else None,
#                     'name': machine.model_number.name if machine.model_number else None,
#                     'amount': machine.product.amount if machine.product else None,
#                     'machinelease': machine.machinelease if machine else None,
#                     'created_at': machine.created_at,
#                     'is_active': machine.is_active,
#                     'payment_type': machine.payment_type,
#                 }

#                 # if machine.rqcode:
#                 #     try:
#                 #         qr_code = QrCode.objects.get(id=machine.rqcode.id)
#                 #         if qr_code.qr:
#                 #             option['rqcode'] = qr_code.qr_code_id
#                 #             option['qr_code_img'] = qr_code.qr.url
#                 #     except ObjectDoesNotExist:
#                 #         pass  # Leave the default empty values if the QR code doesn't exist

#                 data.append(option)

#             return Response({'success': 1, 'message': 'Data Found', 'result': data})
#         except Exception as e:
#             logging.error(f'Error processing request: {e}')
#             return Response({"error": f"Error processing request: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
from django.db import connection
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_products(request):
    try:
        
        products = Product.objects.all()
        serialized_products = []
        for product in products:
            serialized_product = {
                'id': product.id,
                'product_type': product.product_type,
                'amount':product.amount,
                'created_at':product.created_at,
                # Add other fields as needed
            }
            serialized_products.append(serialized_product)
        
        return Response({'success': 1, 'message': 'Products Retrieved Successfully', 'result': serialized_products})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_modelcapacity(request):
    if request.method == 'GET':
        try:
            # Custom logic for GET request
            modelcapacity = ModelCapacity.objects.all()
            serialized_data = [{'id': mc.id, 'name': mc.name, 'model_no': mc.model_no} for mc in modelcapacity]
            return Response({'success': 1, 'message': 'Model capacity retrieved successfully', 'result': serialized_data})
        except Exception as e:
            return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    elif request.method == 'POST':
        try:
            # Custom logic for POST request
            data = request.data
            name = data.get('name')
            model_no = data.get('model_no')
            # Add your custom logic here
            
            # For demonstration, let's create a new ModelCapacity object
            new_model_capacity = ModelCapacity.objects.create(name=name, model_no=model_no)
            serialized_data = {'id': new_model_capacity.id, 'name': new_model_capacity.name, 'model_no': new_model_capacity.model_no}
            
            return Response({'success': 1, 'message': 'Model capacity created successfully', 'result': serialized_data}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @permission_classes([IsAuthenticated])
# def get_machine_and_qrcode(request):
#     if request.method == 'GET':
#         try:
#             machines = MachineMaster.objects.all().order_by('-id')

#             data = []
#             for machine in machines:
#                 print(f"Machine ID: {machine.machine_id}, Product: {machine.product}")

#                 qr_code = machine.rqcode
#                 qr_code_img = ''
#                 if qr_code and qr_code.qr:
#                     qr_code_img = qr_code.qr.url

#                 product_type = None
#                 model_no = None
#                 if machine.product:
#                     product_type = machine.product.name
#                     model_no = machine.product.model_no

#                 option = {
#                     "id": machine.id,
#                     "machine_id": machine.machine_id,
#                     'rqcode': qr_code.qr_code_id if qr_code else '',
#                     'qr_code_img': qr_code_img,
#                     'location': machine.installation_location,
#                     'product_type': product_type,
#                     'model_no': model_no,
#                     'created_at': machine.created_at,
#                     'is_active': machine.is_active,
#                 }
#                 data.append(option)

#             return JsonResponse({'success': 1, 'message': 'Data Found', 'result': data})
#         except Exception as e:
#             logging.error(f'Error processing request: {e}')
#             return JsonResponse({'success': 0, 'message': f'Error: {e}'}, status=500)
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_and_qrcode_for_customer(request):
    if request.method=='GET':
        try: 
            map_obj = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')
            data =[]
            for machine1 in map_obj:
                
                machineid = MachineMaster.objects.filter(id=machine1.machine.id)
                for machine in machineid:
                    # print(machine)
                    if machine.rqcode:
                        qr_code = QrCode.objects.filter(id=machine.rqcode.id)
                        if len(qr_code) !=0 and  qr_code[0].qr :
                            option = {
                                "id" : machine.id,
                                "machine_id" :machine.machine_id,
                                'rqcode' :qr_code[0].qr_code_id,
                                'qr_code_img':qr_code[0].qr.url,
                                'created_at':machine.created_at,
                                # 'created_by':machine.created_by,
                                'is_active':machine.is_active,
                                'payment_type':machine.payment_type,
                                'product_type':machine.product.product_type if machine.product else None,
                                'amount':machine.product.amount if machine.product else None,
                                'model_no':machine.model_number.model_no if machine.model_number else None,
                                'name':machine.model_number.name if machine.model_number else None,
                            }
                            data.append(option)
                        else:
                            option = {
                                "id" : machine.id,
                                "machine_id" :machine.machine_id,
                                'rqcode' : '',
                                'qr_code_img':'',
                                'created_at':machine.created_at,
                                # 'created_by':machine.created_by,
                                'is_active':machine.is_active,
                                'payment_type':machine.payment_type,
                                'product_type':machine.product.product_type if machine.product else None,
                                'amount':machine.product.amount if machine.product else None,
                                'model_no':machine.model_number.model_no if machine.model_number else None,
                                'name':machine.model_number.name if machine.model_number else None,
                            }
                            
                            data.append(option)
                    else:
                        option = {
                            "id" : machine.id,
                            "machine_id" :machine.machine_id,
                            'rqcode' : '',
                            'qr_code_img': '',
                            'created_at':machine.created_at,
                            # 'created_by':machine.created_by,
                            'is_active':machine.is_active,
                            'payment_type':machine.payment_type,
                            'product_type':machine.product.product_type if machine.product else None,
                            'amount':machine.product.amount if machine.product else None,
                            'model_no':machine.model_number.model_no if machine.model_number else None,
                            'name':machine.model_number.name if machine.model_number else None,
                        }
                        data.append(option)
                    # print(machine_id)
                # serializer = MachineUserMappingSerializer(machine_user_mappings, many=True)
            return Response({'success':1,'message':'Data Found','result':data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response(f'Error {e}')
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_and_qrcode_for_user(request):
    if request.method=='GET':
        try: 
            
            map_obj = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
           
            data =[]
            for machine1 in map_obj:
                
                machineid = MachineMaster.objects.filter(id=machine1.machine.id)
                for machine in machineid:
                    # print(machine)
                    if machine.rqcode:
                        qr_code = QrCode.objects.filter(id=machine.rqcode.id)
                        if len(qr_code) !=0 and  qr_code[0].qr :
                            option = {
                                "id" : machine.id,
                                "machine_id" :machine.machine_id,
                                'rqcode' :qr_code[0].qr_code_id,
                                'qr_code_img':qr_code[0].qr.url,
                                'created_at':machine.created_at,
                                # 'created_by':machine.created_by,
                                'is_active':machine.is_active,
                                'payment_type':machine.payment_type,
                                'product_type':machine.product.product_type if machine.product else None,
                                'amount':machine.product.amount if machine.product else None,
                                'model_no':machine.model_number.model_no if machine.model_number else None,
                                'name':machine.model_number.name if machine.model_number else None,
                            }
                            data.append(option)
                        else:
                            option = {
                                "id" : machine.id,
                                "machine_id" :machine.machine_id,
                                'rqcode' : '',
                                'qr_code_img':'',
                                'created_at':machine.created_at,
                                # 'created_by':machine.created_by,
                                'is_active':machine.is_active,
                                'payment_type':machine.payment_type,
                                'product_type':machine.product.product_type if machine.product else None,
                                'amount':machine.product.amount if machine.product else None,
                                'model_no':machine.model_number.model_no if machine.model_number else None,
                                'name':machine.model_number.name if machine.model_number else None,
                            }
                            
                            data.append(option)
                    else:
                        option = {
                            "id" : machine.id,
                            "machine_id" :machine.machine_id,
                            'rqcode' : '',
                            'qr_code_img': '',
                            'created_at':machine.created_at,
                            # 'created_by':machine.created_by,
                            'is_active':machine.is_active,
                            'payment_type':machine.payment_type,
                            'product_type':machine.product.product_type if machine.product else None,
                            'amount':machine.product.amount if machine.product else None,
                            'model_no':machine.model_number.model_no if machine.model_number else None,
                            'name':machine.model_number.name if machine.model_number else None,
                        }
                        data.append(option)
                    # print(machine_id)
                # serializer = MachineUserMappingSerializer(machine_user_mappings, many=True)
            return Response({'success':1,'message':'Data Found','result':data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response(f'Error {e}')

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_unused_qrcode(request):
    if request.method=='GET':
        try:
            machine_user_mappings = MachineMaster.objects.filter(rqcode__isnull=False).order_by('-id')
            data =[]
            send_data =[]
            for machine in machine_user_mappings:
                # print(machine)
                if machine.rqcode:
                    qr_code = QrCode.objects.filter(id=machine.rqcode.id)
                    for qr_codes in qr_code:
                        option = {
                            "id" : qr_codes.id,
                            "qr_code_id" : qr_codes.qr_code_id,
                              
                            "is_active" : qr_codes.is_active,  
                        }
                        if option not in data:
                            data.append(option)
            qr_code1 = QrCode.objects.all()
            for qr_codes1 in qr_code1:
                option1 = {
                    "id" : qr_codes1.id,
                    "qr_code_id" : qr_codes1.qr_code_id,
                    "is_active" : qr_codes1.is_active,  
                }
                if option1 not in data:
                    send_data.append(option1)
            # serializer = MachineUserMappingSerializer(machine_user_mappings, many=True)
            return Response({'success':1,'message':'Data Found','result':send_data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response(f'message {e}')
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_unused_machine_map(request):
    if request.method=='GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.all().order_by('-id')
            data =[]
            send_data =[]
            for machine in machine_user_mappings:
                # print(machine)
                if machine.machine:
                    machineid = MachineMaster.objects.filter(id=machine.machine.id)
                    for machineid in machineid:
                        option = {
                            "id" : machineid.id,
                            "machine_id" : machineid.machine_id,  
                            "is_active": machineid.is_active
                        }
                        if option not in data:
                            data.append(option)
            machineid = MachineMaster.objects.all()
            for machineid1 in machineid:
                option1 = {
                    "id" : machineid1.id,
                    "machine_id" : machineid1.machine_id,
                    "is_active": machineid1.is_active  
                }
                if option1 not in data:
                    send_data.append(option1)
            # serializer = MachineUserMappingSerializer(machine_user_mappings, many=True)
            return Response({'success':1,'message':'Data Found','result':send_data})
        except Exception as e:
            return Response(f'message {e}')
        

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_unused_machine_map_for_customer(request):
    if request.method=='GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer=request.user.id,assigned_user__isnull=True).order_by('-id')
            data =[]
            for machine in machine_user_mappings:
                # print(machine)
                if machine.machine:
                    machineid = MachineMaster.objects.filter(id=machine.machine.id)
                    for machineid in machineid:
                        option = {
                            "id" : machine.id,
                            "machine_id" : machineid.machine_id,  
                            "is_active" : machineid.is_active,  
                        }
                        if option not in data:
                            data.append(option)
            # serializer = MachineUserMappingSerializer(machine_user_mappings, many=True)
            return Response({'success':1,'message':'Data Found','result':data})
        except Exception as e:
            return Response(f'message {e}')

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_and_email_map_user(request):
    if request.method == 'GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer__isnull=True).order_by('-id')
            serializer = MachineUserMapping1Serializer(machine_user_mappings, many=True)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success': 0, 'message': f'Error: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# up and down both are same but down is customised        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_and_map_user_all(request):
    if request.method=='GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer__isnull=True).order_by('-id')
            serializer = MachineUserMapping1Serializer(machine_user_mappings, many=True)

            
            return Response({'success':1,'message':'Data Found','result':serializer.data})
        except Exception as e:
            return Response({'success': 0, 'message': f'Data Not Found, Error: {e}', 'result': []})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_and_email_map_customer(request):
    if request.method == 'GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer__isnull=False).order_by('-id')
            serializer = MachineUserMapping1Serializer(machine_user_mappings, many=True)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success': 0, 'message': f'Error: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# uper wala or niche wala same 
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_and_map(request):
    if request.method == 'GET':
        try:
            machine_user_mappings = MachineUserMapping.objects.all().order_by('-id')
            
            # Create a list of dictionaries only for entries where assigned_customer is None
            data = []
            for machine_mapping in machine_user_mappings:
                if machine_mapping.assigned_customer is None:
                    if machine_mapping.machine:
                        machine_id = machine_mapping.machine.machine_id
                        machine_user_name = machine_mapping.assigned_user.name
                        assigned_user = machine_mapping.assigned_user.email
                        assigned_customer = machine_mapping.assigned_customer
                    else:
                        machine_id = None
                        # assigned_user=None

                    option = {
                        "id": machine_mapping.id,
                        "machine_id": machine_id,
                        "assigned_user": assigned_user,
                        "machine_user_name":machine_user_name,
                        "assigned_customer": assigned_customer,
                        "is_active": machine_mapping.is_active,
                    }
                    data.append(option)

            return Response({'success': 1, 'message': 'Data Found', 'result': data})

        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success': 0, 'message': f'Error: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_mapping_by_customer_id(request, machine_id):
    if request.method == 'GET':
        
        try:
            # Using get_object_or_404 to get a single instance or return a 404 response
            mapping = get_object_or_404(MachineUserMapping, id=machine_id,  assigned_customer__role=3)
            
            # Preparing the response data
            if mapping.assigned_customer:  # Ensure there's a machine to avoid NoneType errors
                machine_info = {
                    "machine_id": mapping.machine.id if mapping.machine else None,
                    "machine_name": mapping.machine.machine_id if mapping.machine else None,  # Example field, adjust based on your model
                    "email": mapping.assigned_customer.email if mapping.assigned_customer else None,
                    "mobile_no": mapping.assigned_customer.mobile_no if mapping.assigned_customer else None,
                    "address1": mapping.assigned_customer.address1 if mapping.assigned_customer else None,
                    "address2": mapping.assigned_customer.address2 if mapping.assigned_customer else None,
                    "created_at": mapping.assigned_customer.created_at if mapping.assigned_customer else None,
                    "created_by": mapping.assigned_customer.created_by if mapping.assigned_customer else None,
                    "country": mapping.assigned_customer.country if mapping.assigned_customer else None,
                    "is_active": mapping.assigned_customer.is_active if mapping.assigned_customer else None,
                    "status": mapping.assigned_customer.status if mapping.assigned_customer else None,
                    "state": mapping.assigned_customer.state if mapping.assigned_customer else None,
                    "updated_at": mapping.assigned_customer.updated_at if mapping.assigned_customer else None,
                    "updated_by": mapping.assigned_customer.updated_by if mapping.assigned_customer else None,
                    
                    
                }
                return Response({'success': 1, 'message': 'Mapping found', 'result': machine_info})
            else:
                return Response({'success': 0, 'message': 'No machine found for the given mapping'})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success': 0, 'message': f'Error: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# def get_machines_for_user_map(request, id):
#     if request.method == 'GET':
#         try:
            # Ensure that user_id corresponds to a role=4 user
            # The additional check might include the role field in your UserMaster model
            # Assuming the role field is in the assigned_user model
#             user_mappings = MachineUserMapping.objects.filter(assigned_user=user_id, assigned_user__role='4')

from django.shortcuts import get_object_or_404

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_mapping_by_user_id(request, machine_id):
    if request.method == 'GET':
        
        try:
            # Using get_object_or_404 to get a single instance or return a 404 response
            mapping = get_object_or_404(MachineUserMapping, id=machine_id,  assigned_user__role=4)
            
            # Preparing the response data
            machine_info = {
                "machine_id": mapping.machine.id,
                "machine_name": mapping.machine.machine_id,
                "email": mapping.assigned_user.email if mapping.assigned_user else None,
                "mobile_no": mapping.assigned_user.mobile_no if mapping.assigned_user else None,
                "address1": mapping.assigned_user.address1 if mapping.assigned_user else None,
                "address2": mapping.assigned_user.address2 if mapping.assigned_user else None,
                "created_at": mapping.assigned_user.created_at if mapping.assigned_user else None,
                "created_by": mapping.assigned_user.created_by if mapping.assigned_user else None,
                "country": mapping.assigned_user.country if mapping.assigned_user else None,
                "is_active": mapping.assigned_user.is_active if mapping.assigned_user else None,
                "status": mapping.assigned_user.status if mapping.assigned_user else None,
                "state": mapping.assigned_user.state if mapping.assigned_user else None,
                "updated_at": mapping.assigned_user.updated_at if mapping.assigned_user else None,
                "updated_by": mapping.assigned_user.updated_by if mapping.assigned_user else None,
                # ... (other fields)
            }

            return Response({'success': 1, 'message': 'Mapping found', 'result': machine_info})
            
        except MachineUserMapping.DoesNotExist:
            return Response({'success': 0, 'message': 'Mapping not found'}, status=404)
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)


        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_based_on_role(request):
    if request.method == 'GET':
        try:
            # Assuming you have a way to determine the user's role
            user_role = request.user

            if user_role == '3':
                # If the user is a customer, fetch only the machines assigned to them
                machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer=request.user)
            elif user_role == '4':
                # If the user is a regular user, fetch only the machines assigned to them
                machine_user_mappings = MachineUserMapping.objects.filter(assigned_user=request.user)
            else:
                # For other roles, fetch all machines
                machine_user_mappings = MachineUserMapping.objects.all()

            # Create a list of dictionaries for the response
            data = []
            for machine_mapping in machine_user_mappings:
                machine_id = machine_mapping.machine.machine_id if machine_mapping.machine else None

                option = {
                    "id": machine_mapping.id,
                    "machine_id": machine_id,
                    "assigned_user": machine_mapping.assigned_user.email if machine_mapping.assigned_user else None,
                    "assigned_customer": machine_mapping.assigned_customer.email if machine_mapping.assigned_customer else None,
                    "is_active": machine_mapping.is_active,
                }
                data.append(option)

            return Response({'success': 1, 'message': 'Data Found', 'result': data})

        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success': 0, 'message': f'Error: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_generate_report_for_all_record(request):
    try:
        # Query your data for the report here
        report_data = MachineUserMapping.objects.all()

        # Serialize the data directly in the view
        serialized_data = []
        for item in report_data:
            serialized_item = {
                'machine_id': item.machine.machine_id,
                'email': item.assigned_customer.email if item.assigned_customer else None,
                'name': item.assigned_customer.name if item.assigned_customer else None,
                'created_by_customer': item.assigned_customer.created_by if item.assigned_customer else None,
                'created_at_customer': item.assigned_customer.created_at if item.assigned_customer else None,
                'assigned_customer_date': item.assigned_customer_date,
                'assigned_user_date': item.assigned_user_date,
                'user_email':item.assigned_user.email if item.assigned_user else None,
                'address1': item.assigned_user.address1 if item.assigned_user else None,
                'address2': item.assigned_user.address2 if item.assigned_user else None,
                'mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
                'country': item.assigned_user.country if item.assigned_user else None,
                'pincode': item.assigned_user.pincode if item.assigned_user else None,
                'created_at_user': item.assigned_user.created_at if item.assigned_user else None,
                'created_by_user': item.assigned_user.created_by if item.assigned_user else None,
                'created_at': item.created_at,
                'created_by': item.created_by,
                'is_active': item.is_active,
                # Add more fields as needed
            }
            serialized_data.append(serialized_item)
        return Response({'success': 1, 'message': 'Report generated successfully', 'result': serialized_data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'})



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_generate_report_for_customer_record(request):
    try:
        
        users = User.objects.all().order_by('-id')
        
        serialized_data = []
        
        for user in users:
            # Query data for the report
            report_data = MachineUserMapping.objects.filter(assigned_customer__isnull=False, assigned_customer=user).order_by('-id')
            
            # Dictionary to store assigned machines count for each customer
            assigned_count_dict = {}
            
            for item in report_data:
                # If the customer already exists in the dictionary, increment the count
                if item.assigned_customer.id in assigned_count_dict:
                    assigned_count_dict[item.assigned_customer.id] += 1
                # Otherwise, add the customer to the dictionary with count 1
                else:
                    assigned_count_dict[item.assigned_customer.id] = 1
            
            # Serialize the machine details
            serializer = MachineUserMappingSerializer(report_data, many=True)
            
            # Serialize the data
            for item in report_data:
                # Only include the customer's details once
                if item.assigned_customer.id in assigned_count_dict:
                    assigned_count = assigned_count_dict[item.assigned_customer.id]
                    serialized_item = {
                        'id':item.id,
                        'machine_id': item.machine.machine_id,
                        'email': item.assigned_customer.email,
                        'name': item.assigned_customer.name,
                        "customer_mobile_no": item.assigned_customer.mobile_no,
                        'assigned_customer_date': item.assigned_customer_date,
                        'created_by_customer': item.assigned_customer.created_by,
                        'created_at_customer': item.assigned_customer.created_at,
                        'is_active': item.assigned_customer.is_active,
                        'assigned_count': assigned_count,
                        'machines_assigned': serializer.data,
                    }
                    serialized_data.append(serialized_item)
                    
                    # Remove the customer from the dictionary to avoid duplication
                    del assigned_count_dict[item.assigned_customer.id]
                
        return Response({'success': 1, 'message': 'Report generated successfully', 'result': serialized_data})
    
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'})
    # try:
    #     print(request.user)
    #     # Query your data for the report here
    #     report_data = MachineUserMapping.objects.filter(assigned_customer__isnull=False)
    #     # Serialize the data directly in the view
    #     serialized_data = []
    #     for item in report_data:
    #         serialized_item = {
    #             'machine_id': item.machine.machine_id,
    #             'email': item.assigned_customer.email if item.assigned_customer else None,
    #             'name': item.assigned_customer.name if item.assigned_customer else None,
    #             "customer_moblie_no":item.assigned_customer.mobile_no if item.assigned_customer else None,
    #             'assigned_customer_date': item.assigned_customer_date,
    #             'created_by_customer': item.assigned_customer.created_by if item.assigned_customer else None,
    #             'created_at_customer': item.assigned_customer.created_at if item.assigned_customer else None,
    #             'is_active': item.assigned_customer.is_active,
    #         }
    #         serialized_data.append(serialized_item)
    #     return Response({'success': 1, 'message': 'Report generated successfully', 'result': serialized_data})
    # except Exception as e:
    #     return Response({'success': 0, 'message': f'Error: {str(e)}'})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_generate_report_for_user_record(request):
    try:
        
        # Query your data for the report here
        report_data = MachineUserMapping.objects.filter(assigned_customer__isnull=True).order_by('-id')
        # Serialize the data directly in the view
        serialized_data = []
        for item in report_data:
            serialized_item = {
                'machine_id': item.machine.machine_id if item.assigned_user else None,
                'email': item.assigned_user.email if item.assigned_user else None,
                'name': item.assigned_user.name if item.assigned_user else None,
                "customer_moblie_no":item.assigned_user.mobile_no if item.assigned_user else None,
                'assigned_customer_date': item.assigned_user_date,
                'created_by_customer': item.assigned_user.created_by if item.assigned_user else None,
                'created_at_customer': item.assigned_user.created_at if item.assigned_user else None,
                
                # Add more fields as needed
            }
            serialized_data.append(serialized_item)
        return Response({'success': 1, 'message': 'Report generated successfully', 'result': serialized_data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stats_api(request):
    
    try:
        # Total Users
        total_users = User.objects.filter(role=4).count()

        # Total Customers
        total_customers = User.objects.filter(role=3).count()
        
        # Active Users
        active_users = User.objects.filter(is_active=True, role=4).count()

        # Inactive Users
        inactive_users = User.objects.filter(is_active=False, role=4).count()

        # Active Customers
        active_customers = User.objects.filter(is_active=True, role=3).count()

        # Inactive Customers
        inactive_customers = User.objects.filter(is_active=False, role=3).count()

         # Total Machines
        total_machines = MachineMaster.objects.count()

       # Active Machines
        active_machines_queryset = MachineMaster.objects.filter(is_active=True)
        
        # Get IDs of active machines
        active_machine_ids = list(active_machines_queryset.values_list('machine_id', flat=True))

        # Print active machine IDs
        

        # Count of active machines
        active_machines_count = active_machines_queryset.count()

        # Inactive Machines
        inactive_machines = total_machines - active_machines_count

        # Machines with QR code mapping
        machines_with_qr = MachineMaster.objects.exclude(rqcode__isnull=True).count()

        # Static value for coin status (adjust as needed)
        static_coin_status = True  

        # Machines with the static coin status
        machines_with_static_coin = static_coin_status

        # Calculate percentages
        percentage_with_qr = (machines_with_qr / total_machines) * 100 if total_machines > 0 else 0
        percentage_with_static_coin = (machines_with_static_coin / total_machines) * 100 if total_machines > 0 else 0

        # Organize data in a different way
        data = {
            
            'total_users': total_users,
            'active_users': active_users,
            'inactive_users': inactive_users,
        
        
            'total_customers': total_customers,
            'active_customers': active_customers,
            'inactive_customers': inactive_customers,
        
        
            'total_machines': total_machines,
            'active_machines': active_machines_count,
            'inactive_machines': inactive_machines,

            'percentage_with_qr': percentage_with_qr,
            'percentage_with_coin': percentage_with_static_coin,
            
        }

        return Response({'success': 1, 'message':'Data Found','result': data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)
    

# @api_view(['GET'])
# def machine_mapping_percentage(request):
#     # Get the total number of machines
#     total_machines = MachineMaster.objects.count()

#     # Initialize counters for assigned users and customers
#     total_assigned_users = 0
#     total_assigned_customers = 0

#    # Query once and iterate to reduce database hits
#     machines = MachineMaster.objects.all()
#     print(machines)
#     for machine in machines:

#         mappings = MachineUserMapping.objects.filter(machine=machine)
#         print(mappings)
#         total_assigned_users += mappings.filter(assigned_user__isnull=True).count()
#         total_assigned_customers += mappings.filter(assigned_customer__isnull=False).count()

#     # Calculate the percentages
#     total_assignments = total_assigned_users + total_assigned_customers
#     users_percentage = (total_assigned_users / total_assignments) * 100 if total_assignments > 0 else 0
#     customers_percentage = (total_assigned_customers / total_assignments) * 100 if total_assignments > 0 else 0


#     # Ensure the response handles cases with no users or customers
#     if total_machines == 0 or total_assignments == 0:
#         users_percentage = 0
#         customers_percentage = 0

    
#     data={
#         'total_machines': total_machines,
#         'total_assigned_users': total_assigned_users,
#         'total_assigned_customers': total_assigned_customers,
#         'users_percentage': users_percentage,
#         'customers_percentage': customers_percentage,
#     }
    
#     return Response({'success':1,'message':'Data Found','result':data})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def machine_mapping_percentage(request):
    try:
        # Get the total number of machines
        total_machines = MachineMaster.objects.count()

        # Initialize counters for assigned users and customers
        total_assigned_users = 0
        total_assigned_customers = 0
        total_unassigned = 0  # Counter for unassigned machines

        # Query once and iterate to reduce database hits
        machines = MachineMaster.objects.all()
        for machine in machines:
            mappings = MachineUserMapping.objects.filter(machine=machine)
            total_assigned_users += mappings.filter(assigned_customer__isnull=True).count()
            total_assigned_customers += mappings.filter(assigned_customer__isnull=False).count()
            # Increment unassigned counter if no mappings found for the machine
            if not mappings.exists():
                total_unassigned += 1

        # Calculate the total number of machines with any assignments
        total_assignments = total_assigned_users + total_assigned_customers

        # Calculate the percentages
        users_percentage = (total_assigned_users / total_machines) * 100 if total_assignments > 0 else 0
        customers_percentage = (total_assigned_customers / total_machines) * 100 if total_assignments > 0 else 0
        unassigned_percentage = (total_unassigned / total_machines) * 100 if total_machines > 0 else 0

        # Ensure the response handles cases with no users or customers
        if total_machines == 0 or total_assignments == 0:
            users_percentage = 0
            customers_percentage = 0

        data = {
            'total_machines': total_machines,
            'total_assigned_users': total_assigned_users,
            'total_assigned_customers': total_assigned_customers,
            'total_unassigned': total_unassigned,
            'users_percentage': users_percentage,
            'customers_percentage': customers_percentage,
            'unassigned_percentage': unassigned_percentage
        }

        return Response({'success': 1, 'message': 'Data Found', 'result': data})
    except Exception as e:
        return Response({'success': 0, 'message': str(e)})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def percentage_of_machines_with_qr(request):
    try:
        # Total machines
        total_machines = MachineMaster.objects.count()

        # Machines with QR code mapping
        machines_with_qr = MachineMaster.objects.exclude(rqcode__isnull=True).count()

        # Calculate percentages
        if total_machines > 0:
            percentage_with_qr = (machines_with_qr / total_machines) * 100
            percentage_with_static_coin = 100 - percentage_with_qr  # If no QR, consider it 100% coin
        else:
            percentage_with_qr = 0
            percentage_with_static_coin = 0


        data = {
            'percentage_with_qr': percentage_with_qr,
            'percentage_with_static_coin': percentage_with_static_coin,
        }

        return Response({
            'success': 1,
            'message': 'Data Found',
            'result': data
        })
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)


from django.db.models.functions import TruncMonth

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def month_wise_percentage(request):
    try:
        # Users and Customers count per month
        users_per_month = User.objects.filter(role=4).annotate(month=TruncMonth('created_at')).values('month').annotate(count=Count('id'))
        customers_per_month = User.objects.filter(role=3).annotate(month=TruncMonth('created_at')).values('month').annotate(count=Count('id'))

        # Organize data by month
        data = {
            'users': [
                {'month': entry['month'].strftime('%Y-%m'), 'count': entry['count']} for entry in users_per_month
            ],
            'customers': [
                {'month': entry['month'].strftime('%Y-%m'), 'count': entry['count']} for entry in customers_per_month
            ],
        }

        return Response({'success': 1, 'message': 'Data Found', 'result': data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)


# pass without year and month
from django.utils import timezone

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def users_and_customers_created_in_month(request):
    try:
        current_date = timezone.now()
        # year = current_date.year
        # month = current_date.month
        current_year = current_date.year

        # user_count = User.objects.filter(created_at__year=year, created_at__month=month, role='4').count()
        # customer_count = User.objects.filter(created_at__year=year, created_at__month=month, role='3').count()
        months={
            1:'jan',
            2:'feb',
            3:'mar',
            4:'apr',
            5:'may',
            6:'jun',
            7:'jul',
            8:'aug',
            9:'sept',
            10:'oct',
            11:'nov',
            12:'dec'
        }
        user_month_counts = {}
        customer_month_counts={}
        for month in range(1, 13):  # Loop through months from January to December
            user_count = User.objects.filter(created_at__year=current_year, created_at__month=month, role='4').count()
            customer_count = User.objects.filter(created_at__year=current_year, created_at__month=month, role='3').count()
            # month_counts[months[month]] = {'user_count': user_count, 'customer_count': customer_count}
            user_month_counts[months[month]] =  user_count
            customer_month_counts[months[month]] = customer_count

        return Response({'success': 1,'message': f'Count of users and customers created in the current month',
        # 'month_counts': month_counts,
        'user_month_counts': user_month_counts,
        'customer_month_counts': customer_month_counts,
        })
    except Exception as e:
        return Response({'success': 0, 'message': str(e)})


import os
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware, get_default_timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import MachineUserMapping
import datetime
from django.conf import settings

from django.http import HttpResponse
from dateutil import parser
#new
#1

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def read_machine_user_mapping_report_excel(request):
    today_date = timezone.localdate()
    date_from_raw = request.data.get('from')
    date_to_raw = request.data.get('to')

    # Validate input data
    if not date_from_raw or not date_to_raw:
        return Response({'success':0,'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Filter the queryset based on the provided date range
        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to),assigned_user__isnull=False,assigned_user__role=4,assigned_user__created_by=1)
        # Serialize the data
        serialized_data = [{
            'machine_id': item.machine.machine_id,
            'email': item.assigned_user.email if item.assigned_user else None,
            'name': item.assigned_user.name if item.assigned_user else None,
            'mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
            'user_assigned_date':item.assigned_user_date if item.assigned_user else None,
            # 'is_active': item.assigned_user.is_active,
            # Add the rest of your fields here
        } for item in report_data]

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success':0,"message": "No data found for the provided date range."})

        # Process the serialized_data as needed

        # Return a success response with the processed data
        return Response({'success':1,'message': 'Excel file processed successfully', 'result': serialized_data})

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success':0,'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#2
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def read_machine_customer_mapping_report_excel(request):

    today_date = timezone.localdate()
    # Retrieve 'from' and 'to' date parameters from the POST request data
    date_from_raw = request.data.get('from')
    date_to_raw = request.data.get('to')

    # Validate input data
    if not date_from_raw or not date_to_raw:
        return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        users = User.objects.all()
        
        serialized_data = []
        
        for user in users:
            # Query data for the report
            report_data = MachineUserMapping.objects.filter(assigned_customer_date__range=(date_from, date_to),assigned_customer__isnull=False, assigned_customer=user)
            
            # Dictionary to store assigned machines count for each customer
            assigned_count_dict = {}
            
            for item in report_data:
                # If the customer already exists in the dictionary, increment the count
                if item.assigned_customer.id in assigned_count_dict:
                    assigned_count_dict[item.assigned_customer.id] += 1
                # Otherwise, add the customer to the dictionary with count 1
                else:
                    assigned_count_dict[item.assigned_customer.id] = 1
            
            # Serialize the machine details
            serializer = MachineUserMappingSerializer(report_data, many=True)
            
            # Serialize the data
            for item in report_data:
                # Only include the customer's details once
                if item.assigned_customer.id in assigned_count_dict:
                    assigned_count = assigned_count_dict[item.assigned_customer.id]
                    serialized_item = {
                        'machine_id': item.machine.machine_id,
                        'email': item.assigned_customer.email,
                        'name': item.assigned_customer.name,
                        "customer_mobile_no": item.assigned_customer.mobile_no,
                        'assigned_customer_date': item.assigned_customer_date,
                        'created_by_customer': item.assigned_customer.created_by,
                        'created_at_customer': item.assigned_customer.created_at,
                        'is_active': item.assigned_customer.is_active,
                        'assigned_count': assigned_count,
                        'machines_assigned': serializer.data,
                    }
                    serialized_data.append(serialized_item)
                    
                    # Remove the customer from the dictionary to avoid duplication
                    del assigned_count_dict[item.assigned_customer.id]    

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success': 0, "message": "No data found for the provided date range."})

        # Process the serialized_data as needed

        # Return a success response with the processed data
        return Response({'success': 1, 'message': 'Excel file processed successfully', 'result': serialized_data})

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def machine_customer_mapping_report_excel(request):
    # Retrieve 'from' and 'to' date parameters from the POST request data
    date_from_raw = request.data.get('from')
    date_to_raw = request.data.get('to')

    # Validate input data
    if not date_from_raw or not date_to_raw:
        return Response({'success':0,'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw)
        date_to = parser.parse(date_to_raw)

        # Filter the queryset based on the provided date range
        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to))
        # Serialize the data
        serialized_data = [{
            'machine_id': item.machine.machine_id,
            'email': item.assigned_customer.email if item.assigned_customer else None,
            'name': item.assigned_customer.name if item.assigned_customer else None,
            'address1': item.assigned_customer.address1 if item.assigned_customer else None,
            'address2': item.assigned_customer.address2 if item.assigned_customer else None,
            'mobile_no': item.assigned_customer.mobile_no if item.assigned_customer else None,
            'country': item.assigned_customer.country if item.assigned_customer else None,
            'pincode': item.assigned_customer.pincode if item.assigned_customer else None,
            # 'is_active': item.assigned_customer.is_active,
            # Add the rest of your fields here
        } for item in report_data]

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success':0,"message": "No data found for the provided date range."})

        # Process the serialized_data as needed

        # Return a success response with the processed data
        return Response({'success':1,'message': 'Excel file processed successfully', 'result': serialized_data})

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success':0,'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def machine_user_mapping_report_excel(request):
    # Retrieve 'from' and 'to' date parameters from the POST request data
    date_from_raw = request.data.get('from')
    date_to_raw = request.data.get('to')

    # Validate input data
    if not date_from_raw or not date_to_raw:
        return Response({'success':0,'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw)
        date_to = parser.parse(date_to_raw)

        # Filter the queryset based on the provided date range
        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to))
        # Serialize the data
        serialized_data = [{
            'machine_id': item.machine.machine_id,
            'email': item.assigned_customer.email if item.assigned_customer else None,
            'name': item.assigned_customer.name if item.assigned_customer else None,
            'address1': item.assigned_user.address1 if item.assigned_user else None,
            'address2': item.assigned_user.address2 if item.assigned_user else None,
            'mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
            'country': item.assigned_user.country if item.assigned_user else None,
            'pincode': item.assigned_user.pincode if item.assigned_user else None,
            'is_active': item.is_active,
        } for item in report_data]

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success':0,"message": "No data found for the provided date range."})

        # Process the serialized_data as needed

        # Return a success response with the processed data
        return Response({'success':1,'message': 'Excel file processed successfully', 'result': serialized_data})

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success':0,'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



   
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def read_machine_user_mapping_report_pdf(request):
    # Retrieve 'from' and 'to' date parameters from the POST request data
    date_from = request.data.get('from')
    date_to = request.data.get('to')

    # Parse the 'from' and 'to' dates
    date_from_parsed = parse_date(date_from) if date_from else None
    date_to_parsed = parse_date(date_to) if date_to else None

    # Make the datetime objects timezone-aware
    if date_from_parsed:
        date_from_parsed = make_aware(datetime.datetime.combine(date_from_parsed, datetime.time.min), get_default_timezone())
    if date_to_parsed:
        date_to_parsed = make_aware(datetime.datetime.combine(date_to_parsed, datetime.time.max), get_default_timezone())

    # Ensure both dates are provided and valid
    if not date_from_parsed or not date_to_parsed:
        return Response({"message": "Please provide both 'from' and 'to' date parameters in YYYY-MM-DD format."}, status=400)

    # Filter the queryset based on the provided date range
    report_data = MachineUserMapping.objects.filter(created_at__range=[date_from_parsed, date_to_parsed])

    # Serialize the data
    serialized_data = [{
        'machine_id': item.machine.machine_id,
        'assigned_customer_email': item.assigned_customer.email if item.assigned_customer else None,
        'assigned_customer_name': item.assigned_customer.name if item.assigned_customer else None,
        'assigned_user_address1': item.assigned_user.address1 if item.assigned_user else None,
        'assigned_user_address2': item.assigned_user.address2 if item.assigned_user else None,
        'assigned_user_mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
        'assigned_user_country': item.assigned_user.country if item.assigned_user else None,
        'assigned_user_pincode': item.assigned_user.pincode if item.assigned_user else None,
        'is_active': item.is_active,
    } for item in report_data]

    # Check if serialized_data is empty
    if not serialized_data:
        return Response({"message": "No data found for the provided date range."}, status=404)

    # Generate timestamps for the filename
    timestamp_from = date_from_parsed.strftime("%Y%m%d%H%M%S") if date_from_parsed else 'from'
    timestamp_to = date_to_parsed.strftime("%Y%m%d%H%M%S") if date_to_parsed else 'to'

    # Create a new folder inside MEDIA_ROOT for PDF reports
    folder_name = 'pdf_reports'
    folder_path = os.path.join(settings.MEDIA_ROOT, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    # Generate PDF file with timestamps in the filename using os.path.join
    filename = f'MachineUserMappingReport_{timestamp_from}_{timestamp_to}.pdf'
    pdf_file_path = os.path.join(folder_path, filename)

    # Set up the response to return the PDF file and open it in the browser
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Generate the PDF content
    with open(pdf_file_path, 'wb') as pdf_file:
        generate_pdf(serialized_data, pdf_file)

    with open(pdf_file_path, 'rb') as pdf_file:
        response.write(pdf_file.read())


    return Response({'success':1,'message':'Data Found','result':serialized_data})


def generate_pdf(serialized_data, pdf_file):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors

    doc = SimpleDocTemplate(pdf_file, pagesize=letter)
    elements = []

    # Define the table structure
    table_data = [["Machine ID", "Email", "Name", "Address1", "Address2", "Mobile No", "Country", "Pincode", "Is Active"]]
    
    for item in serialized_data:
        table_data.append([
            item['machine_id'],
            item['assigned_customer_email'],
            item['assigned_customer_name'],
            item['assigned_user_address1'],
            item['assigned_user_address2'],
            item['assigned_user_mobile_no'],
            item['assigned_user_country'],
            item['assigned_user_pincode'],
            item['is_active'],
        ])

    # Create the table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
        ('WORDWRAP', (0, 0), (-1, -1), 1),
    ]))

    # Add the table to the elements list
    elements.append(table)

    # Build the PDF document
    doc.build(elements)

#new
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_assigned_machine_by_user(request):
    if request.method == 'GET':
        try:
            # Retrieve the logged-in user
            user = request.user

            # Retrieve the machine assigned to the logged-in user
            mapping = MachineUserMapping.objects.get(assigned_user=user)
            machine_assigned = mapping.machine

            # Get the URL to the QR code file
            qr_code_url = machine_assigned.rqcode.qr.url if machine_assigned.rqcode.qr else None

            # Combine both sets of data
            response_data = {
                'machine_id': machine_assigned.machine_id,
                'qr_code_id': machine_assigned.rqcode.qr_code_id,
                'qr_code_url': qr_code_url,
                'created_at': machine_assigned.created_at
            }

            return Response({'success': 1, 'message': 'Data Found', 'result': response_data})
        except MachineUserMapping.DoesNotExist:
            return Response({'message': 'No machine assigned to this user'}, status=404)





from datetime import datetime, timedelta
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.request import Request as RestRequest

#############################################################################################


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_yesterday_report(request):
    try:
        request = request._request if isinstance(request, RestRequest) else request
        end_date = datetime.now().date() - timedelta(days=1)
        start_date = end_date - timedelta(days=1)
        return generate_report(request, start_date, end_date)
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def generate_today_report(request):
    try:
        request = request._request if isinstance(request, RestRequest) else request
        end_date = datetime.now().date()
        start_date = end_date
        return generate_report(request, start_date, end_date)
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def generate_weekly_report(request):
    try:
        request = request._request if isinstance(request, RestRequest) else request
        end_date = datetime.now().date()
        start_date = end_date - timedelta(weeks=1)
        return generate_report(request, start_date, end_date)
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def generate_monthly_report(request):
    try:
        request = request._request if isinstance(request, RestRequest) else request
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)  # Assuming a month has 30 days
        return generate_report(request, start_date, end_date)
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def generate_three_months_report(request):
    try:
        request = request._request if isinstance(request, RestRequest) else request
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=90)  # Three months has around 90 days
        return generate_report(request, start_date, end_date)
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


@api_view(['GET'])

def generate_report(request, start_date, end_date):
    try:
        report_data = MachineUserMapping.objects.filter(created_at__range=(start_date, end_date),assigned_user__isnull=True)
        serialized_data = [{
            'machine_id': item.machine.machine_id,
            'email': item.assigned_customer.email if item.assigned_customer else None,
            'name': item.assigned_customer.name if item.assigned_customer else None,
            'address1': item.assigned_customer.address1 if item.assigned_customer else None,
            'address2': item.assigned_customer.address2 if item.assigned_customer else None,
            'mobile_no': item.assigned_customer.mobile_no if item.assigned_customer else None,
            'country': item.assigned_customer.country if item.assigned_customer else None,
            'pincode': item.assigned_customer.pincode if item.assigned_customer else None,
            'is_active': item.is_active,
        } for item in report_data]
        return JsonResponse({'success':1,'message':'Data Found','result': serialized_data})
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['GET'])
# def generate_report(request, start_date, end_date):
#     try:
#         include_customer_data = request.GET.get('include_customer_data')
#         include_user_data = request.GET.get('include_user_data')

#         # If neither include_customer_data nor include_user_data is provided, default to both
#         if not include_customer_data and not include_user_data:
#             include_customer_data = True
#             include_user_data = True

#         serialized_data = []
#         if include_customer_data:
#             # Filter for items where assigned_user is null
#             report_data_customer = MachineUserMapping.objects.filter(created_at__range=(start_date, end_date), assigned_customer__isnull=False)
#             # Serialize data for items where assigned_user is null
#             serialized_data_customer = [{
#                 'machine_id': item.machine.machine_id,
#                 'email': item.assigned_customer.email if item.assigned_customer else None,
#                 'name': item.assigned_customer.name if item.assigned_customer else None,
#                 'address1': item.assigned_customer.address1 if item.assigned_customer else None,
#                 'address2': item.assigned_customer.address2 if item.assigned_customer else None,
#                 'mobile_no': item.assigned_customer.mobile_no if item.assigned_customer else None,
#                 'country': item.assigned_customer.country if item.assigned_customer else None,
#                 'pincode': item.assigned_customer.pincode if item.assigned_customer else None,
#                 'is_active': item.is_active,
#             } for item in report_data_customer]
#             serialized_data.extend(serialized_data_customer)

#         if include_user_data:
#             # Filter for items where assigned_customer is null
#             report_data_user = MachineUserMapping.objects.filter(created_at__range=(start_date, end_date), assigned_user__isnull=False)
#             # Serialize data for items where assigned_customer is null
#             serialized_data_user = [{
#                 'machine_id': item.machine.machine_id,
#                 'email': item.assigned_user.email if item.assigned_user else None,
#                 'name': item.assigned_user.name if item.assigned_user else None,
#                 'address1': item.assigned_user.address1 if item.assigned_user else None,
#                 'address2': item.assigned_user.address2 if item.assigned_user else None,
#                 'mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
#                 'country': item.assigned_user.country if item.assigned_user else None,
#                 'pincode': item.assigned_user.pincode if item.assigned_user else None,
#                 'is_active': item.is_active,
#             } for item in report_data_user]
#             serialized_data.extend(serialized_data_user)


#         return JsonResponse({'report_data': serialized_data})
#     except Exception as e:
#         return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
from userapp.models import MStatus,MQTTData
import json
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mstatus_for_assigned_user(request):
    if request.method == 'GET':
        try:
            # Retrieve machine-user mappings for the logged-in user
            map_objs = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')

             # Get all machine IDs from the retrieved mappings
            machine_ids = [map_obj.machine.machine_id for map_obj in map_objs]

            # Define time threshold (two minutes ago)
            threshold_time = timezone.now() - timedelta(minutes=2)

            # Fetch machine data along with their creation times using cursor
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT payload, created_at
                    FROM userapp_mqttdata
                    WHERE created_at >= %s
                    """,
                    [threshold_time]
                )
                rows = cursor.fetchall()

            # Prepare a dictionary to store online/offline status for each machine
            machine_status = defaultdict(lambda: {'status': "OFF", 'last_online_time': None})
            offline_durations = defaultdict(lambda: timedelta())
            online_durations = defaultdict(lambda: timedelta())

            # Iterate through the fetched MQTT data to determine machine statuses and calculate durations
            for row in rows:
                payload = json.loads(row[0])
                machine_id = payload.get('M_Id')
                created_at = row[1]
                if machine_id in machine_ids:
                    if machine_status[machine_id]['status'] == "OFF":
                        machine_status[machine_id]['status'] = "ON"
                        machine_status[machine_id]['last_online_time'] = created_at
                    online_time = machine_status[machine_id]['last_online_time']
                    offline_durations[machine_id] += created_at - online_time if online_time else timedelta()
                    machine_status[machine_id]['last_online_time'] = created_at

            # Calculate online durations for each machine
            for machine_id in machine_ids:
                if machine_status[machine_id]['status'] == "ON":
                    online_time = machine_status[machine_id]['last_online_time']
                    online_durations[machine_id] += timezone.now() - online_time if online_time else timedelta()

            # Fetch latest stock data for each machine
            stocks = []
            for machine_id in machine_ids:
                stock = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()
                if stock:
                    # Determine the color based on stock level
                    color = 'green' if stock.stock > 25 else 'yellow' if 16 <= stock.stock <= 25 else 'red' if 1 <= stock.stock <= 5 else 'Refill'
                    # Append machine status data to the list
                    stocks.append({
                        'M_Id': stock.m_id,
                        'Capacity': stock.capacity,
                        'Stock': stock.stock,
                        'Status': color,
                        'Mode': stock.mode,
                        'Color': color,
                        'created_at': stock.created_at,
                        'created_by': stock.created_by,
                        'offline_duration': offline_durations[machine_id],
                        'online_duration': online_durations[machine_id],
                    })

            # # Prepare the response
            # response_data = {
            #     'success': 1,
            #     'message': 'Data Found',
            #     'result': stocks
            # }

            # return Response(response_data)

            return Response({'success': 1, 'message': 'Data Found', 'result': stocks})
        except MachineUserMapping.DoesNotExist:
            return Response({'message': 'No machines assigned to this user'}, status=404)
        
from collections import defaultdict
from datetime import datetime, timedelta
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mstatus_for_assigned_customer(request):
    if request.method == 'GET':
        try:
            # Retrieve machine-user mappings for the logged-in user's assigned customer
            map_objs = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')

            # Get all machine IDs from the retrieved mappings
            machine_ids = [map_obj.machine.machine_id for map_obj in map_objs]

            # Define time threshold (two minutes ago)
            threshold_time = timezone.now() - timedelta(minutes=2)

            # Fetch machine data along with their creation times using cursor
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT payload, created_at
                    FROM userapp_mqttdata
                    WHERE created_at >= %s
                    """,
                    [threshold_time]
                )
                rows = cursor.fetchall()

            # Prepare a dictionary to store online/offline status for each machine
            machine_status = defaultdict(lambda: {'status': "OFF", 'last_online_time': None})
            offline_durations = defaultdict(lambda: timedelta())
            online_durations = defaultdict(lambda: timedelta())

            # Iterate through the fetched MQTT data to determine machine statuses and calculate durations
            for row in rows:
                payload = json.loads(row[0])
                machine_id = payload.get('M_Id')
                created_at = row[1]
                if machine_id in machine_ids:
                    if machine_status[machine_id]['status'] == "OFF":
                        machine_status[machine_id]['status'] = "ON"
                        machine_status[machine_id]['last_online_time'] = created_at
                    online_time = machine_status[machine_id]['last_online_time']
                    offline_durations[machine_id] += created_at - online_time if online_time else timedelta()
                    machine_status[machine_id]['last_online_time'] = created_at

            # Calculate online durations for each machine
            for machine_id in machine_ids:
                if machine_status[machine_id]['status'] == "ON":
                    online_time = machine_status[machine_id]['last_online_time']
                    online_durations[machine_id] += timezone.now() - online_time if online_time else timedelta()

            # Fetch latest stock data for each machine
            stocks = []
            for machine_id in machine_ids:
                stock = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()
                if stock:
                    # Determine the color based on stock level
                    color = 'green' if stock.stock > 25 else 'yellow' if 16 <= stock.stock <= 25 else 'red' if 1 <= stock.stock <= 5 else 'Refill'
                    # Append machine status data to the list
                    stocks.append({
                        'M_Id': stock.m_id,
                        'Capacity': stock.capacity,
                        'Stock': stock.stock,
                        'Status': color,
                        'Mode': stock.mode,
                        'Color': color,
                        'created_at': stock.created_at,
                        'created_by': stock.created_by,
                        'offline_duration': offline_durations[machine_id],
                        'online_duration': online_durations[machine_id],
                    })

            # Return the response with success message and machine status data
            return Response({'success': 1, 'message': 'Data Found', 'result': stocks})
        
        except MachineUserMapping.DoesNotExist:
            # If no machines are assigned to the user, return a 404 response
            return Response({'success': 0, 'message': 'No machines assigned to this user'}, status=404)        
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_mstatus_for_assigned_customer(request):
#     if request.method == 'GET':
#         try:
#             # Retrieve machine-user mappings for the logged-in user
#             map_objs = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')

#             # Get all machine IDs from MachineMaster
#             machine_ids = [map_obj.machine.machine_id for map_obj in map_objs]

#             # Define time threshold (two minutes ago)
#             threshold_time = timezone.now() - timedelta(minutes=2)

#             # Fetch machine data along with their creation times using cursor
#             with connection.cursor() as cursor:
#                 cursor.execute(
#                     """
#                     SELECT payload, created_at
#                     FROM userapp_mqttdata
#                     WHERE created_at >= %s
#                     """,
#                     [threshold_time]
#                 )
#                 rows = cursor.fetchall()

#             # Prepare dictionaries to store machine statuses
#             machine_status = {}
#             machine_off = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"

#             # Iterate through the fetched data to determine machine statuses
#             for data in rows:
#                 M_Id = json.loads(data['payload']).get('M_Id')
#                 if M_Id in machine_ids and data['created_at'] >= threshold_time:
#                     machine_status[M_Id] = "ON"
#                     machine_off.pop(M_Id, None)  # Remove machine from the OFF list if it's online

#             # Fetch latest stock data for each machine
#             stocks = []
#             for machine_id in machine_ids:
#                 stock = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()
#                 if stock:
#                     color = 'green' if 16 <= stock.stock < 200 else 'yellow' if 6 <= stock.stock <= 15 else 'red' if 1 <= stock.stock <= 5 else 'Refill'
#                     machine_on_off = "ON" if machine_status.get(machine_id) else "OFF"
#                     stocks.append({
#                         'M_Id': stock.m_id,
#                         'Capacity': stock.capacity,
#                         'Stock': stock.stock,
#                         'Status': color,
#                         'Mode': stock.mode,
#                         'Color': color,
#                         'created_at': stock.created_at,
#                         'created_by': stock.created_by,
#                         'machine_status': machine_on_off,
#                     })


#             return Response({'success': 1, 'message': 'Data Found', 'result': stocks})
#         except MachineUserMapping.DoesNotExist:
#             return Response({'success': 0,'message': 'No machines assigned to this user'}, status=404)
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_created_user(request):
    if request.method=='GET':
        user_id = request.get('user_id')
        try:
            user = User.objects.get(id=user_id)
            # Serialize the user data if needed
            user_data = {
                'id': user.id,
                'username': user.username,
                'email': user.email,
            }
            return Response(user_data)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes=[IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            queryset = Product.objects.all().order_by('-id')
            serializer = ProductSerializer(queryset, many=True)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            return Response({'success': 0, 'message': 'Not Found', 'result': str(e)})

    def retrieve(self, request, pk, *args, **kwargs):
        try:
            chp = Product.objects.get(pk=pk)
            serializer = ProductSerializer(chp)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def create(self, request, *args, **kwargs):
        # try:
            # tenant = request.tenant.schema_name  # Assuming request.tenant.sce gives the tenant instance
            serializer = ProductSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            # Set the tenant field before saving the serializer
            # serializer.validated_data['tenant'] = tenant
            serializer.save()

            return Response({'success': 1, 'message': 'Data Created', 'result': serializer.data})
        # except ValidationError as ve:
        #     return Response({'success': 0, 'message': 'Not Created', 'result': ve.detail})

    def update(self, request, pk, *args, **kwargs):
        try:
            chp = Product.objects.get(pk=pk)
            serializer = ProductSerializer(chp, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def destroy(self, request, pk, *args, **kwargs):
        try:
            chp = Product.objects.get(pk=pk)
            chp.delete()
            return Response({'success': 1, 'message': 'Data Deleted'})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})


        

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_customer_dashboard(request):
    try:
        
        if request.method == 'GET':
            # Filter MachineUserMapping based on the logged-in admin user's ID
            assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)
            # print('assigned_customers',assigned_customers)

            # Extracting the customer IDs from assigned_customers
            assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)
            

            # Filter all users assigned to the customers (including users created by customers)
            count_users= User.objects.filter(role = 4,created_by = request.user.id).count()
            

            # print(count_users)

            # Total Users (including all assigned users)
            total_users = count_users
            
            # Active Users
            active_users = User.objects.filter(role = 4,created_by = request.user.id,is_active=True).count()

            # # Inactive Users
            inactive_users = User.objects.filter(role = 4,created_by = request.user.id,is_active=False).count()

            # Total Machines
            total_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids).count()

            # Active Machines
            active_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids, is_active=True).count()

            # Inactive Machines
            inactive_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids, is_active=False).count()

            # Organize data in a different way
            data = {
                'total_users': total_users,
                'active_users': active_users,
                'inactive_users': inactive_users,
                'total_machines': total_machines,
                'active_machines': active_machines,
                'inactive_machines': inactive_machines,
            }

            return Response({'success': 1, 'message': 'Data Found', 'result': data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)
    

# @api_view(['GET'])
# def get_customer_user_machine_mapping_percentage(request):
#     try:
#         print('hello',request.user)
#         # Filter MachineUserMapping based on the logged-in customer user's ID
#         assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)

#         # Extract the IDs of the assigned customers
#         assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)

#         # Get the total number of machines associated with the assigned customers
#         total_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids).count()

#         # Initialize counters for assigned users and customers
#         total_assigned_users = 0
#         total_assigned_customers = 0

#         # Query once and iterate to reduce database hits
#         machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids)
#         for machine in machines:
#             mappings = MachineUserMapping.objects.filter(machine=machine)
#             total_assigned_users += mappings.filter(assigned_user__isnull=False).count()
#             total_assigned_customers += mappings.filter(assigned_customer__isnull=True).count()

#         # Calculate the percentages
#         total_assignments = total_assigned_users + total_assigned_customers
#         users_percentage = (total_assigned_users / total_assignments) * 100 if total_assignments > 0 else 0
#         customers_percentage = (total_assigned_customers / total_assignments) * 100 if total_assignments > 0 else 0

#         # Ensure the response handles cases with no users or customers
#         if total_machines == 0 or total_assignments == 0:
#             users_percentage = 0
#             customers_percentage = 0

#         data = {
#             'total_machines': total_machines,
#             'total_assigned_users': total_assigned_users,
#             'total_assigned_customers': total_assigned_customers,
#             'users_percentage': users_percentage,
#             'customers_percentage': customers_percentage,
#         }

#         return Response({'success': 1, 'message': 'Data Found', 'result': data})
#     except Exception as e:
#         return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_user_machine_mapping_percentage(request):
    try:
        
        # Filter MachineUserMapping based on the logged-in customer user's ID
        assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)

        # Extract the IDs of the assigned customers
        assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)

        # Get the total number of machines associated with the assigned customers
        total_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids).count()

        # Initialize counters for assigned users and customers
        total_assigned_users = 0
        total_assigned_customers = 0

        # Query once and iterate to reduce database hits
        machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids)
        for machine in machines:
            mappings = MachineUserMapping.objects.filter(machine=machine)
            total_assigned_users += mappings.filter(assigned_user__isnull=False).count()
            total_assigned_customers += mappings.filter(assigned_customer__isnull=True).count()

        # Calculate the total number of unassigned machines
        total_unassigned_machines = total_machines - (total_assigned_users + total_assigned_customers)

        # Calculate the percentages
        total_assignments = total_assigned_users + total_assigned_customers + total_unassigned_machines
        users_percentage = (total_assigned_users / total_assignments) * 100 if total_assignments > 0 else 0
        customers_percentage = (total_assigned_customers / total_assignments) * 100 if total_assignments > 0 else 0
        unassigned_machines_percentage = (total_unassigned_machines / total_assignments) * 100 if total_assignments > 0 else 0

        # Ensure the response handles cases with no users or customers
        if total_machines == 0 or total_assignments == 0:
            users_percentage = 0
            customers_percentage = 0
            unassigned_machines_percentage = 0

        data = {
            'total_machines': total_machines,
            'total_assigned_users': total_assigned_users,
            'total_assigned_customers': total_assigned_customers,
            'total_unassigned_machines': total_unassigned_machines,
            'users_percentage': users_percentage,
            'customers_percentage': customers_percentage,
            'unassigned_machines_percentage': unassigned_machines_percentage,
        }

        # Check if 'total_unassigned_machines' is not already in 'data' before appending
        if 'total_unassigned_machines' not in data:
            data['total_unassigned_machines'] = total_unassigned_machines

        return Response({'success': 1, 'message': 'Data Found', 'result': data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_user_percentage_of_machines_with_qr(request):
    try:
        assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)
        assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)

        # Print machine_id values
        for machine_id in assigned_customer_ids:
            print(machine_id)

        total_machines = MachineMaster.objects.filter(id__in=assigned_customer_ids).count()
    

        if total_machines > 0:
            machines_with_qr = MachineMaster.objects.filter(rqcode__isnull=False, id__in=assigned_customer_ids).count()
            # print(machines_with_qr)
            percentage_with_qr = (machines_with_qr / total_machines) * 100
            percentage_with_static_coin = 100 - percentage_with_qr
        else:
            percentage_with_qr = 0
            percentage_with_static_coin = 0

        data = {
            'percentage_with_qr': percentage_with_qr,
            'percentage_with_static_coin': percentage_with_static_coin,
        }

        return Response({'success': 1, 'message': 'Data Found', 'result': data}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_user_percentage_of_machines_with_qr(request):
    try:
        assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)
        assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)

        # Print machine_id values
        for machine_id in assigned_customer_ids:
            print(machine_id)

        total_machines = MachineMaster.objects.filter(id__in=assigned_customer_ids).count()
    

        if total_machines > 0:
            machines_with_qr = MachineMaster.objects.filter(rqcode__isnull=False, id__in=assigned_customer_ids).count()
            # print(machines_with_qr)
            percentage_with_qr = (machines_with_qr / total_machines) * 100
            percentage_with_static_coin = 100 - percentage_with_qr
        else:
            percentage_with_qr = 0
            percentage_with_static_coin = 0

        data = {
            'percentage_with_qr': percentage_with_qr,
            'percentage_with_static_coin': percentage_with_static_coin,
        }

        return Response({'success': 1, 'message': 'Data Found', 'result': data}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_role_dashboard_qr_coin(request):
    if request.method == 'GET':
        try:
            
            # Filter MachineUserMapping objects based on the logged-in user
            assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)
            
            assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)
            
            machine_names = MachineMaster.objects.filter(id__in=assigned_customer_ids).values_list('machine_id', flat=True)

            
            # Retrieve MStatus objects associated with assigned customer IDs
            mstatus_objects = MStatus.objects.filter(m_id__in=machine_names)
            

            # Retrieve machine names associated with the assigned customer IDs
            
            

            # Initialize counts of 'COIN' and 'QR' modes
            coin_count = 0
            qr_count = 0

            # Count occurrences of each mode type
            for obj in mstatus_objects:
                if obj.mode == 'COIN':
                    coin_count += 1
                elif obj.mode == 'QR':
                    qr_count += 1

            # Calculate percentages
            total_count = coin_count + qr_count
            coin_percentage = (coin_count / total_count) * 100 if total_count != 0 else 0
            qr_percentage = (qr_count / total_count) * 100 if total_count != 0 else 0

            # Prepare response data
            data = {
                'coin_count': coin_count,
                'qr_count': qr_count,
                'coin_percentage': coin_percentage,
                'qr_percentage': qr_percentage
            }

            response_data = {
                'success': 1,
                'message': 'Data Found',
                'result': data
            }

            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_users_created_in_month(request):
    try:
        if request.method=='GET':
            
            current_date = timezone.now()
            current_year = current_date.year

            months = {
                1: 'jan',
                2: 'feb',
                3: 'mar',
                4: 'apr',
                5: 'may',
                6: 'jun',
                7: 'jul',
                8: 'aug',
                9: 'sept',
                10: 'oct',
                11: 'nov',
                12: 'dec'
            }
            user_month_counts = {}
            # customer_month_counts = {}

            for month in range(1, 13):  # Loop through months from January to December
                user_count = User.objects.filter(role='4', created_at__year=current_year, created_at__month=month, created_by=request.user.id).count()
                # customer_count = User.objects.filter(created_at__year=current_year, created_at__month=month, role='3', id__in=assigned_customer_ids).count()
               
                user_month_counts[months[month]] = user_count
                # customer_month_counts[months[month]] = customer_count

            return Response({
                'success': 1,
                'message': 'Count of users and customers created in each month',
                'user_month_counts': user_month_counts,
                # 'customer_month_counts': customer_month_counts,
            })
    except Exception as e:
        return Response({'success': 0, 'message': str(e)})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_user_dashboard(request):
    try:
        
        if request.method == 'GET':
            # Filter MachineUserMapping based on the logged-in admin user's ID
            assigned_customers = MachineUserMapping.objects.filter(assigned_user=request.user.id)

           

            # Extracting the customer IDs from assigned_customers
            assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)
           

            # Filter all users assigned to the customers (including users created by customers)
            # count_users= User.objects.filter(role = 4,created_by = request.user.id).count()
            # print('count user',count_users)

            # # print(count_users)

            # # Total Users (including all assigned users)
            # total_users = count_users
            
            # Active Users
            # active_users = User.objects.filter(role = 4,created_by = request.user.id,is_active=True).count()

            # # # Inactive Users
            # inactive_users = User.objects.filter(role = 4,created_by = request.user.id,is_active=False).count()

            # Total Machines
            total_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids).count()

            # Active Machines
            active_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids, is_active=True).count()

            # Inactive Machines
            inactive_machines = MachineMaster.objects.filter(machineusermapping__machine_id__in=assigned_customer_ids, is_active=False).count()

            # Organize data in a different way
            data = {
                # 'total_users': total_users,
                # 'active_users': active_users,
                # 'inactive_users': inactive_users,
                'total_machines': total_machines,
                'active_machines': active_machines,
                'inactive_machines': inactive_machines,
            }

            return Response({'success': 1, 'message': 'Data Found', 'result': data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_cutomer_user_dashboard_machine_mapping_percentage(request):
    try:
        # Get the logged-in admin user
        admin_user = request.user

        

        # Get machines assigned to the admin user
        assigned_machines = MachineMaster.objects.filter(machineusermapping__assigned_user=admin_user).distinct()

        # Count the total number of assigned machines
        total_machines = assigned_machines.count()

        # Initialize counters for assigned users and customers
        total_assigned_users = 0
        total_assigned_customers = 0

        # Query all assigned machines and iterate to reduce database hits
        for machine in assigned_machines:
            mappings = MachineUserMapping.objects.filter(machine=machine)
            total_assigned_users += mappings.filter(assigned_user__isnull=False).count()
            total_assigned_customers += mappings.filter(assigned_customer__isnull=True).count()

        # Calculate the percentages
        total_assignments = total_assigned_users + total_assigned_customers
        users_percentage = (total_assigned_users / total_assignments) * 100 if total_assignments > 0 else 0
        customers_percentage = (total_assigned_customers / total_assignments) * 100 if total_assignments > 0 else 0

        # Ensure the response handles cases with no users or customers
        if total_machines == 0 or total_assignments == 0:
            users_percentage = 0
            customers_percentage = 0

        data = {
            'total_machines': total_machines,
            'total_assigned_users': total_assigned_users,
            'total_assigned_customers': total_assigned_customers,
            'users_percentage': users_percentage,
            'customers_percentage': customers_percentage,
        }
        return Response({'success': 1, 'message': 'Data Found', 'result': data})
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=500)
    

# @api_view(['GET', 'DELETE'])
# def site_settings_api_for_logo(request):
#     print(request.tenant)
#     if hasattr(request, 'tenant') and request.tenant:
#         # Attempt to retrieve site settings for the current tenant
#         settings_queryset = SiteSettings.objects.filter(tenant=request.tenant)

#         # Check if settings exist for the current tenant
#         if settings_queryset.exists():
#             # Assuming there's only one SiteSettings object per tenant, you can retrieve the first one
#             settings = settings_queryset.first()

#             # Prepare response data with logo and favicon URLs
#             data = {
#                 'logo': settings.logo.url if settings.logo else '',
#                 'favicon': settings.favicon.url if settings.favicon else ''
#             }
#             return Response(data)
#         # return Response(data)
#     elif request.method == 'DELETE':
#         try:
#             settings = SiteSettings.objects.get(tenant=request.tenant)
#             settings.delete()
#             return Response({'detail': 'Site settings deleted successfully'})
#         except SiteSettings.DoesNotExist:
#             return Response({'detail': 'Site settings not found'}, status=status.HTTP_404_NOT_FOUND)
# @api_view(['GET'])
# def site_settings_api_for_logo(request):
#     # Ensure that the tenant attribute exists in the request object
#     # It's assumed that you're using Django Tenant Schemas or similar multi-tenancy setup
#     if hasattr(request, 'tenant') and request.tenant:
#         # Attempt to retrieve site settings for the current tenant
#         settings_queryset = SiteSettings.objects.filter(tenant=request.tenant)
#         print(settings_queryset,'set')

#         if settings_queryset.exists():
#             settings = settings_queryset.first()

#             data = {
#                 'logo': settings.logo.url if settings.logo else None,
#                 'favicon': settings.favicon.url if settings.favicon else None
#             }
#             return Response({'success':1,'message':'Data Found','result':data})
#         else:
#             return Response({'success':0,'message': 'Site settings not found for the current tenant'}, status=status.HTTP_404_NOT_FOUND)
#     else:
#         return Response({'success':0,'message': 'Tenant not found in the request'}, status=status.HTTP_400_BAD_REQUEST)
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def site_settings_api_for_logo(request):
    # Ensure that the tenant attribute exists in the request object
    # It's assumed that you're using Django Tenant Schemas or similar multi-tenancy setup
    if hasattr(request, 'tenant') and request.tenant:
        # Attempt to retrieve site settings for the current tenant
        settings_queryset = SiteSettings.objects.filter(tenant=request.tenant)
        
        if settings_queryset.exists():
            settings = settings_queryset.first()

            data = {
                'logo': settings.logo.url if settings.logo else get_default_logo_url(),
                'favicon': settings.favicon.url if settings.favicon else get_default_favicon_url()
            }
            return Response({'success': 1, 'message': 'Data Found', 'result': data})
        else:
            # If site settings are not found, return default logo URL
            default_logo_url = get_default_logo_url()
            default_favicon_url = get_default_favicon_url()
            data = {
                'logo': default_logo_url,
                'favicon': default_favicon_url
            }
            return Response({'success': 1, 'message': 'Default Data', 'result': data})
    else:
        return Response({'success': 0, 'message': 'Tenant not found in the request'}, status=status.HTTP_400_BAD_REQUEST)


def get_default_logo_url():
    return f"{settings.MEDIA_URL}default_logo.png" if hasattr(settings, 'MEDIA_URL') else None
# D:\ivendsoft_tenant\public\media\logo_ivendsoft_logo_2__1_-removebg-preview (1).png

def get_default_favicon_url():
    return f"{settings.MEDIA_URL}favicon.ico" if hasattr(settings, 'MEDIA_URL') else None

@api_view(['POST', 'PUT'])
# @permission_classes([IsAuthenticated])
def tenant_upload_logo(request):
    try:
        tenant = request.tenant
        settings, created = SiteSettings.objects.get_or_create(tenant=tenant)

        
        logo_file = request.FILES.get('logo')
        favicon_file = request.FILES.get('favicon')

        if logo_file:
            # Delete old logo file
            if settings.logo:
                settings.logo.delete()
            # Save the new logo file
            settings.logo = logo_file
            settings.save()
            logo_path = settings.logo.url  # Retrieve the URL of the saved logo
        else:
            # Use default logo image path
            logo_path = None
        
        if favicon_file:
            # Save or update the favicon file
            settings.favicon = favicon_file
            settings.save()

        data_img = {
            'logo_path': settings.logo.url if settings.logo else None,
            'favicon_path': settings.favicon.url if settings.favicon else None
        }

        if created:
            message = 'Logo uploaded successfully'
            status_code = status.HTTP_201_CREATED
        else:
            message = 'Logo updated successfully'
            status_code = status.HTTP_200_OK


        return Response({'success': 1, 'message': 'Logo updated successfully', 'result': data_img}, status=status.HTTP_200_OK)
        # else:
        #     return Response({'success': 0, 'message': 'Only POST or PUT method allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def post_customer_role_user_for_mapping_report_excel(request):
    try:
        date_from_raw = request.GET.get('from_date')
       
        date_to_raw = request.GET.get('to_date')
        

        # # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success':0,'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()
        
        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Filter report data based on the logged-in customer user's ID
        report_data = MachineUserMapping.objects.filter(assigned_user_date__range=(date_from, date_to),assigned_customer=request.user.id)
        
        serialized_data = []
        for item in report_data:
            # Only include mapped machines where assigned_user is not None
            if item.assigned_user:
                serialized_item = {
                    'machine_id': item.machine.machine_id if item.machine else None,
                    'user_id': item.assigned_user.id,
                    'name': item.assigned_user.name,
                    'user_mobile_no': item.assigned_user.mobile_no,
                    'email': item.assigned_user.email,
                    'assigned_machine_date':item.assigned_user_date
                }
                serialized_data.append(serialized_item)

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success':0,"message": "No data found for the provided date range."})

        # Return a success response with the processed data
        return Response({'success':1,'message': 'Excel file processed successfully', 'result': serialized_data})

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success':0,'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def post_customer_role_user_payment_report_excel(request):
    try:
        # Retrieve 'from' and 'to' date parameters from the POST request data
       
        date_from_raw = request.data.get('from')
        date_to_raw = request.data.get('to')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Filter MachineUserMapping queryset based on the provided date range and user
        machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer=request.user.id)

        # Initialize a dictionary to hold mode counts per machine
        mode_counts_per_machine = defaultdict(lambda: defaultdict(int))

        # Iterate over each MachineUserMapping object
        for mapping in machine_user_mappings:
            # Retrieve machine_id from MachineUserMapping
            machine_id = mapping.machine.machine_id

            # Filter MStatus records for the current machine and date range
            mstatus_records = MStatus.objects.filter(m_id=machine_id, created_at__range=(date_from, date_to))

            # Count mode occurrences for the current machine
            for record in mstatus_records:
                if record.mode:
                    mode_counts_per_machine[machine_id][record.mode] += 1

        # Initialize an empty list to hold the result
        data_for_response = []

        # Iterate over mode counts per machine to prepare data for response
        for machine_id, modes in mode_counts_per_machine.items():
            total_modes = sum(modes.values())
            total_coin = modes.get('COIN', 0) * 5
            total_qr = modes.get('QR', 0) * 5
            total_amount = total_coin + total_qr
            
            data_for_response.append({
                'machine_id': machine_id,
                'coin_mode_counts': modes.get('COIN', 0),
                'qr_mode_counts': modes.get('QR', 0),
                'total_modes': total_modes,
                'total_amount': total_amount
            })

        return Response({
            'success': 1,
            'message': 'Data found',
            'result': data_for_response
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)



# @api_view(['GET'])
# def get_product_list(request):
#     if request.method == 'GET':
#         data = []
#         products = Product.objects.all().order_by('-id')
#         for product in products:
#             user_data = {
#                 'id':product.id,
#                 # 'machine_id': product.machine_id,
#                 # 'name': product.name if product.name else None,
#                 'product_type': product.product_type if product.product_type else None,
#                 # 'model_no': product.model_no if product.model_no else None,
#                 'amount': product.amount if product.amount else None,
#                 'created_at': product.created_at if product.created_at else None,
#             }
#             data.append(user_data)
#         return Response({'success': 1, 'message': 'Data Found', 'result':data})
    
# @api_view(['GET'])
# def get_product_model(request):
#     if request.method == 'GET':
#         data = []
#         products = Product.objects.all().order_by('-id')
#         for product in products:
#             user_data = {
#                 'id':product.id,
#                 # 'machine_id': product.machine_id,
#                 'name': product.name if product.name else None,
#                 # 'product_type': product.product_type if product.product_type else None,
#                 'model_no': product.model_no if product.model_no else None,
#                 # 'amount': product.amount if product.amount else None,
#                 'created_at': product.created_at if product.created_at else None,
#             }
#             data.append(user_data)
#         return Response({'success': 1, 'message': 'Data Found', 'result':data})

# @api_view(['GET'])
# def get_product_list(request):
#     if request.method == 'GET':
#         data = []
#         products = Product.objects.all().order_by('-id')
#         for product in products:
#             user_data = {
#                 'id': product.id,
#                 'product_type': product.product_type,
#                 'amount': product.amount,
#                 'created_at': product.created_at,
#             }
#             # Remove empty fields from the response
#             user_data = {key: value for key, value in user_data.items() if value is not None}
#             data.append(user_data)
#         return Response({'success': 1, 'message': 'Data Found', 'result': data})
    
# @api_view(['GET'])
# def get_product_model(request):
#     if request.method == 'GET':
#         data = []
#         products = Product.objects.all().order_by('-id')
#         for product in products:
#             user_data = {
#                 'id': product.id,
#                 'name': product.name,
#                 'model_no': product.model_no,
#                 'created_at': product.created_at,
#             }
#             # Remove empty fields from the response
#             user_data = {key: value for key, value in user_data.items() if value is not None}
#             data.append(user_data)
#         return Response({'success': 1, 'message': 'Data Found', 'result': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_product_list(request):
    if request.method == 'GET':
        data = []
        products = Product.objects.all().order_by('-id')
        for product in products:
            user_data = {
                'id': product.id,
                'product_type': product.product_type,
                'amount': product.amount,
                'created_at': product.created_at if product.created_at else None,
            }
            # Remove empty fields from the response
            user_data = {key: value for key, value in user_data.items() if value is not None}
            data.append(user_data)
        return Response({'success': 1, 'message': 'Data Found', 'result': data})

class ModelCapacityViewSet(viewsets.ModelViewSet):
    queryset = ModelCapacity.objects.all()
    serializer_class = ModelCapacitySerializer
    permission_classes=[IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            queryset = ModelCapacity.objects.all().order_by('-id')
            serializer = ModelCapacitySerializer(queryset, many=True)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            return Response({'success': 0, 'message': 'Not Found', 'result': str(e)})

    def retrieve(self, request, pk, *args, **kwargs):
        try:
            chp = ModelCapacity.objects.get(pk=pk)
            serializer = ModelCapacitySerializer(chp)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def create(self, request, *args, **kwargs):
        # try:
            
            serializer = ModelCapacitySerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
    
            serializer.save()

            return Response({'success': 1, 'message': 'Data Created', 'result': serializer.data})
        # except ValidationError as ve:
        #     return Response({'success': 0, 'message': 'Not Created', 'result': ve.detail})

    def update(self, request, pk, *args, **kwargs):
        try:
            chp = ModelCapacity.objects.get(pk=pk)
            serializer = ModelCapacitySerializer(chp, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def destroy(self, request, pk, *args, **kwargs):
        try:
            chp = ModelCapacity.objects.get(pk=pk)
            chp.delete()
            return Response({'success': 1, 'message': 'Data Deleted'})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_product_model(request):
    if request.method == 'GET':
        data = []
        modelcapacitys = ModelCapacity.objects.all().order_by('-id')
        for product in modelcapacitys:
            user_data = {
                'id': product.id,
                'name': product.name,
                'model_no': product.model_no,
                'created_at': product.created_at if product.created_at else None,
            }
            # Remove empty fields from the response
            user_data = {key: value for key, value in user_data.items() if value is not None}
            data.append(user_data)
        return Response({'success': 1, 'message': 'Data Found', 'result': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_details_view(request, id):
    try:
        machine = MachineMaster.objects.get(id=id)
        data = {
            "machine_id": machine.machine_id,
            "amount": machine.product.amount if machine.product else None,
            "product_type": machine.product.product_type if machine.product else None,
            # "rqcode": machine.rqcode.id if machine.rqcode else None,
            "model_number": machine.model_number.model_no if machine.model_number else None,
            "name": machine.model_number.name if machine.model_number else None,
            "payment_type": machine.payment_type,
            "installation_location": machine.installation_location,
            "is_active": machine.is_active,
            "created_at": machine.created_at
        }
        return Response({'success':1,'message':'Data Found','result':data})
    except MachineMaster.DoesNotExist:
        return Response({'success':0,'message': 'Machine with this ID does not exist.'})
    
# @api_view(['POST'])
# def create_color(request):
#     serializer = ColorStoreSerializer(data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response({'success':1,'message':'Data Found','result':serializer.data}, status=status.HTTP_201_CREATED)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# @api_view(['PUT'])
# def update_color(request, color_id):
#     try:
#         color_store = ColorStore.objects.get(pk=color_id)
#     except ColorStore.DoesNotExist:
#         return Response({'success':0,'mesage': 'Color not found'}, status=status.HTTP_404_NOT_FOUND)

#     serializer = ColorStoreSerializer(color_store, data=request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response({'success':1,'message':'Data Found','result':serializer.data}, status=status.HTTP_200_OK)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
class ColorStoreViewSet(viewsets.ModelViewSet):
    queryset = ColorStore.objects.all()
    serializer_class = ColorStoreSerializer
    # permission_classes=[IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            queryset = ColorStore.objects.all().order_by('-id')
            serializer = ColorStoreSerializer(queryset, many=True)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            return Response({'success': 0, 'message': 'Not Found', 'result': str(e)})

    def retrieve(self, request, pk, *args, **kwargs):
        try:
            chp = ColorStore.objects.get(pk=pk)
            serializer = ColorStoreSerializer(chp)
            return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def create(self, request, *args, **kwargs):
        try:
            
            serializer = ColorStoreSerializer(data=request.data)
            
            serializer.is_valid(raise_exception=True)
    
            serializer.save()

            return Response({'success': 1, 'message': 'Data Created', 'result': serializer.data})
        except ValidationError as ve:
            return Response({'success': 0, 'message': 'Not Created', 'result': ve.detail})

    def update(self, request, pk, *args, **kwargs):
        try:
            chp = ColorStore.objects.get(pk=pk)
            serializer = ColorStoreSerializer(chp, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def destroy(self, request, pk, *args, **kwargs):
        try:
            chp = ColorStore.objects.get(pk=pk)
            chp.delete()
            return Response({'success': 1, 'message': 'Data Deleted'})
        except Product.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

# from rest_framework.permissions import IsAuthenticated, IsAdminUser
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def tenant_list(request):
    if request.method == 'GET':
        tenants = Tenant.objects.all()
        tenant_list = [{'id': tenant.id, 'name': tenant.schema_name} for tenant in tenants]
        return Response({'success':1,'message':'Data Found','result':tenant_list})
    
    # elif request.method == 'POST':
    #     if not request.user.is_superuser:
    #         return Response({"detail": "Only superadmins can create tenants."}, status=status.HTTP_403_FORBIDDEN)
    #     data = request.data
    #     tenant = Tenant.objects.create(name=data['name'])
    #     return Response({'id': tenant.id, 'name': tenant.name}, status=status.HTTP_201_CREATED)

# @api_view(['GET', 'PUT', 'DELETE'])
# @permission_classes([IsAuthenticated, IsAdminUser])
# def tenant_detail(request, pk):
#     try:
#         tenant = Tenant.objects.get(pk=pk)
#     except Tenant.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)
    
#     if request.method == 'GET':
#         return Response({'id': tenant.id, 'name': tenant.schema_name})
    
#     elif request.method == 'PUT':
#         return Response({"detail": "Updates to tenants are not allowed."}, status=status.HTTP_403_FORBIDDEN)
    
#     elif request.method == 'DELETE':
#         return Response({"detail": "Deletion of tenants is not allowed."}, status=status.HTTP_403_FORBIDDEN)

# @api_view(['GET', 'POST'])
# @permission_classes([IsAuthenticated])
# def tenant_data_list(request):
#     if request.method == 'GET':
#         if hasattr(request.user, 'customuser'):
#             tenant_data = Tenant.objects.filter(tenant=request.user.customuser.tenant)
#             data_list = [{'id': data.id, 'data': data.data} for data in tenant_data]
#             return Response(data_list)
#         else:
#             return Response({"detail": "User is not associated with any tenant."}, status=status.HTTP_403_FORBIDDEN)
    
#     elif request.method == 'POST':
#         if hasattr(request.user, 'customuser'):
#             data = request.data
#             data['tenant'] = request.user.customuser.tenant.id
#             tenant_data = Tenant.objects.create(**data)
#             return Response({'id': tenant_data.id, 'data': tenant_data.data}, status=status.HTTP_201_CREATED)
#         else:
#             return Response({"detail": "User is not associated with any tenant."}, status=status.HTTP_403_FORBIDDEN)

# @api_view(['GET', 'PUT', 'DELETE'])
# @permission_classes([IsAuthenticated])
# def tenant_data_detail(request, pk):
#     try:
#         tenant_data = TenantData.objects.get(pk=pk)
#     except TenantData.DoesNotExist:
#         return Response(status=status.HTTP_404_NOT_FOUND)
    
#     if tenant_data.tenant != request.user.customuser.tenant:
#         return Response({"detail": "You don't have permission to access this data."}, status=status.HTTP_403_FORBIDDEN)
    
#     if request.method == 'GET':
#         return Response({'id': tenant_data.id, 'data': tenant_data.data})
    
#     elif request.method == 'PUT':
#         data = request.data
#         for key in data.keys():
#             setattr(tenant_data, key, data[key])
#         tenant_data.save()
#         return Response({'id': tenant_data.id, 'data': tenant_data.data})
    
#     elif request.method == 'DELETE':
#         tenant_data.delete()
#         return Response(status=status.HTTP_204_NO_CONTENT)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def module_permissions(request):
#     if request.user.is_authenticated:
#         tenant = request.tenant
#         print(tenant)
#         # permissions = ModulePermission.objects.get(tenant=tenant)
#         # data = {
#         #     'can_view_module1': permissions.can_view_module1,
#         #     'can_view_module2': permissions.can_view_module2,
#         #     # Add more modules as needed
#         # }
#         return Response()
#     else:
#         return Response({"detail": "User is not authenticated."}, status=status.HTTP_401_UNAUTHORIZED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_count_for_customer(request):
    try:
        # Get all machine user mappings
        all_mappings = MachineUserMapping.objects.all()

        # Initialize a dictionary to store counts for each customer
        customer_counts = {}

        # Count machines for each customer
        for mapping in all_mappings:
            if mapping.assigned_customer:
                customer_name = mapping.assigned_customer.name
                if customer_name not in customer_counts:
                    customer_counts[customer_name] = {
                        'machine_count': 0,
                        'machine_ids': set()
                    }
                customer_counts[customer_name]['machine_count'] += 1
                customer_counts[customer_name]['machine_ids'].add(mapping.machine.machine_id)

        # Construct response data
        response_data = []
        for customer_name, count_info in customer_counts.items():
            response_data.append({
                'customer_name': customer_name,
                'machine_count': count_info['machine_count'],
                'machine_ids': list(count_info['machine_ids'])
            })

        return Response({'success':1,'message':'Data Found','result':response_data})

    except Exception as e:
        # Handle exceptions appropriately, perhaps return an error response
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
# @api_view(['GET'])
# def get_machine_count_for_customer(request):
#     try:
#         # Aggregate counts for all customers and retrieve customer name along with machine IDs
#         machine_counts = MachineUserMapping.objects.values('assigned_customer__name').annotate(
#             machine_count=Count('machine_id'),
#             machine_ids=Count('machine__machine_id', distinct=True)
#         ).filter(assigned_customer__isnull=False)
        
#         # Construct response data
#         response_data = [
#             {
#                 'customer_name': item['assigned_customer__name'],
#                 'machine_count': item['machine_count'],
#                 'machine_ids': list(MachineUserMapping.objects.filter(assigned_customer__name=item['assigned_customer__name']).values_list('machine__machine_id', flat=True))
#             }
#             for item in machine_counts
#         ]

#         return Response({'success':1,'message':'Data Found','result':response_data})
#     except Exception as e:
#         # Handle exceptions appropriately, perhaps return an error response
#         return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_count_for_user(request):
    try:
        # Aggregate counts for all customers and retrieve customer name along with machine IDs
        machine_counts = MachineUserMapping.objects.values('assigned_user__name').annotate(
            machine_count=Count('machine_id'),
            machine_ids=Count('machine__machine_id', distinct=True)
        ).filter(assigned_user__isnull=False)
        
        # Construct response data
        response_data = [
            {
                'customer_name': item['assigned_user__name'],
                'machine_count': item['machine_count'],
                'machine_ids': list(MachineUserMapping.objects.filter(assigned_user__name=item['assigned_user__name']).values_list('machine__machine_id', flat=True))
            }
            for item in machine_counts
        ]

        return Response({'success':1,'message':'Data Found','result':response_data})
    except Exception as e:
        # Handle exceptions appropriately, perhaps return an error response
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_user_dashboard_percentage_of_machines_with_qr(request):
    try:
        assigned_customers = MachineUserMapping.objects.filter(assigned_user=request.user.id)
        assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)

        # Print machine_id values
        for machine_id in assigned_customer_ids:
            print(machine_id)

        total_machines = MachineMaster.objects.filter(id__in=assigned_customer_ids).count()
    

        if total_machines > 0:
            machines_with_qr = MachineMaster.objects.filter(rqcode__isnull=False, id__in=assigned_customer_ids).count()
            
            percentage_with_qr = (machines_with_qr / total_machines) * 100
            percentage_with_static_coin = 100 - percentage_with_qr
        else:
            percentage_with_qr = 0
            percentage_with_static_coin = 0

        data = {
            'percentage_with_qr': percentage_with_qr,
            'percentage_with_static_coin': percentage_with_static_coin,
        }

        return Response({'success': 1, 'message': 'Data Found', 'result': data}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_user_dashboard_role_dashboard_qr_coin(request):
    if request.method == 'GET':
        try:
            
            # Filter MachineUserMapping objects based on the logged-in user
            assigned_customers = MachineUserMapping.objects.filter(assigned_user=request.user.id)
            
            assigned_customer_ids = assigned_customers.values_list('machine_id', flat=True)
            
            machine_names = MachineMaster.objects.filter(id__in=assigned_customer_ids).values_list('machine_id', flat=True)

            
            # Retrieve MStatus objects associated with assigned customer IDs
            mstatus_objects = MStatus.objects.filter(m_id__in=machine_names)
 
            # Initialize counts of 'COIN' and 'QR' modes
            coin_count = 0
            qr_count = 0

            # Count occurrences of each mode type
            for obj in mstatus_objects:
                if obj.mode == 'COIN':
                    coin_count += 1
                elif obj.mode == 'QR':
                    qr_count += 1

            # Calculate percentages
            total_count = coin_count + qr_count
            coin_percentage = (coin_count / total_count) * 100 if total_count != 0 else 0
            qr_percentage = (qr_count / total_count) * 100 if total_count != 0 else 0

            # Prepare response data
            data = {
                'coin_count': coin_count,
                'qr_count': qr_count,
                'coin_percentage': coin_percentage,
                'qr_percentage': qr_percentage
            }

            response_data = {
                'success': 1,
                'message': 'Data Found',
                'result': data
            }

            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'success': 0, 'message': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def assigned_customer_machine_count(request):
    if request.method == 'GET':
       # Retrieve all users
        users = User.objects.all()
        # Create a list to store data for each user
        assigned_data = []
        
        # Iterate over each user
        for user in users:
            # Retrieve machines assigned to the user
            machines_assigned = MachineUserMapping.objects.filter(assigned_customer=user)
            
            # Count the number of assigned machines
            assigned_count = machines_assigned.count()
            
            # Serialize the machine details
            serializer = MachineUserMappingSerializer(machines_assigned, many=True)
            
            # Create dictionary for user's data
            user_data = {
                'assigned_count': assigned_count,
                'machines_assigned': serializer.data,
                'email': user.email
            }
            
            # Append user's data to the list
            assigned_data.append(user_data)
        
        return Response({'success':1,'message':'Data Found','result':assigned_data})

from openpyxl.utils import get_column_letter

import logging as logger
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
import os
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from dateutil import parser


from openpyxl import Workbook, styles
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_download_excel_user(request):
    tenant = request.tenant.user_id
    today_date = timezone.localdate()

    date_from_raw = request.GET.get('from_date')
    date_to_raw = request.GET.get('to_date')

    if not date_from_raw or not date_to_raw:
        return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    date_from = parser.parse(date_from_raw).date()
    date_to = parser.parse(date_to_raw).date() + timedelta(days=1) - timedelta(seconds=1)

    machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)
    threshold_time = timezone.now() - timedelta(minutes=2)

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT organization_name
            FROM saasapp_companyregistration
            WHERE user_id = %s
        """, [tenant])
        row = cursor.fetchone()
        organization_name = row[0] if row else None

    report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to), assigned_customer__isnull=True)

    serialized_data = [{
        'sr_no': index + 1,
        'name': item.assigned_user.name if item.assigned_user else None,
        'user_mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
        'email': item.assigned_user.email if item.assigned_user else None,
        'machine_id': item.machine.machine_id if item.machine else None,
        'machine_assigned_date': item.created_at.strftime('%Y-%m-%d'),
    } for index, item in enumerate(report_data)]

    workbook = Workbook()
    worksheet = workbook.active

    if organization_name:
        # Limit the merge to fewer columns to avoid affecting the sr_no column
        worksheet.merge_cells('A4:F4')
        cell = worksheet['A4']
        cell.value = f"Organization Name: {organization_name}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(wrap_text=True)

    worksheet.merge_cells('A1:F2')
    cell = worksheet['A1']
    cell.value = "User Machine Report"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet.merge_cells('A3:F3')
    date_cell = worksheet['A3']
    date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
    date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ["Sr No", "Name", "Mobile", "Email", "Machine ID", "Machine\nAssigned Date"]
    for col_num, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=5, column=col_num, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'),
                                    left=Side(style='medium'), right=Side(style='medium'))

    for row_num, row_data in enumerate(serialized_data, start=6):
        for col_num, value in enumerate(row_data.values(), start=1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                 left=Side(style='thin'), right=Side(style='thin'))

    # Adjust the column width excluding the organization name row
    # Adjust the column width
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        column_letter = chr(64 + column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Set a fixed width for the "Sr No" column
    worksheet.column_dimensions['A'].width = 10

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

    workbook.save(response)

    return response
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def admin_download_excel_user(request):
#     # Retrieve tenant information based on the logged-in user
#     tenant = request.tenant.user_id
    
#     today_date = timezone.localdate()
    

#     date_from_raw = request.GET.get('from_date')
#     date_to_raw = request.GET.get('to_date')

#     # Validate input data
#     if not date_from_raw or not date_to_raw:
#         return Response({'success':0,'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

#     # Parse date strings into datetime objects using dateutil.parser
#     date_from = parser.parse(date_from_raw).date()
    
#     date_to = parser.parse(date_to_raw).date()
    

#     date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)
#     # Get all machine IDs from MachineMaster
#     machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

#     # Define time threshold (two minutes ago)
#     threshold_time = timezone.now() - timedelta(minutes=2)
    

#     # Fetch organization name from database
#     with connection.cursor() as cursor:
#         cursor.execute("""
#             SELECT organization_name
#             FROM saasapp_companyregistration
#             WHERE user_id = %s
#         """, [tenant])
#         row = cursor.fetchone()
#         organization_name = row[0] if row else None

#     # Fetch machine data along with their creation times using cursor
#     report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to), assigned_customer__isnull=True)
    

#     serialized_data = [{
#         'sr_no': index + 1,  # Serial number
#         'name': item.assigned_user.name if item.assigned_user else None,
#         'user_mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
#         'email': item.assigned_user.email if item.assigned_user else None,
#         'machine_id': item.machine.machine_id if item.machine else None,
#         'machine_assigned_date': item.created_at.strftime('%Y-%m-%d'),
#     } for index, item in enumerate(report_data)]

    
#     # Create a new Excel workbook
#     workbook = Workbook()
#     worksheet = workbook.active

#     # Add the organization name at the top of the worksheet
#     if organization_name:
#         worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
#         cell = worksheet['A4']  # Get the top-left cell of the merged range
#         cell.value = f"Organization Name: {organization_name}"
#         cell.font = Font(bold=True,size=12)
#         cell.alignment = Alignment(wrap_text=True)  # Align text to the center

#     # Add the "Customer Machine Report" label in the middle of the third row
#     worksheet.merge_cells('A1:F2')  # Merge cells for the label
#     cell = worksheet['A1']  # Get the top-left cell of the merged range
#     cell.value = "User Machine Report"
#     cell.font = Font(bold=True,size=16)  # Make the label bold
#     cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#     worksheet.merge_cells('A3:F3')  # Merge cells for the date range
#     date_cell = worksheet['A3']  # Get the top-left cell of the merged range
#     date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
#     date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    
#     # Add headers for customer section
#     headers = ["Sr No", "Name", "Mobile", "Email", "Machine ID", "Machine\nAssigned Date"]
#     for col_num, header in enumerate(headers, start=1):
#         header_cell = worksheet.cell(row=5, column=col_num, value=header)
#         header_cell.font = Font(bold=True)
#         header_cell.alignment = Alignment(horizontal='center', vertical='center')
#         header_cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), 
#                                         left=Side(style='medium'), right=Side(style='medium'))

#         # If data exists, populate the worksheet starting from row 6
#         for row_num, row_data in enumerate(serialized_data, start=6):
#             for col_num, value in enumerate(row_data.values(), start=1):
#                 cell = worksheet.cell(row=row_num, column=col_num, value=value)
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
#                                      left=Side(style='thin'), right=Side(style='thin'))
    
#     # Auto-adjust column widths based on content size
#     for col in worksheet.columns:
#             max_length = 0
#             column = col[0].column  # Get the column index
#             column_letter = chr(65 + column - 1)  # Convert column index to letter
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2
#             worksheet.column_dimensions[column_letter].width = adjusted_width

#     # Set the response headers for file download
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

#     # Write the Excel workbook to the response
#     workbook.save(response)

#     return response



# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def admin_download_excel_customer(request):
#     try:
#         # Retrieve tenant information based on the logged-in user
#         tenant = request.tenant.user_id
#         today_date = timezone.localdate()

#         date_from_raw = request.GET.get('from_date')
#         date_to_raw = request.GET.get('to_date')

#         print(request.user)

#         # Validate input data
#         if not date_from_raw or not date_to_raw:
#             return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

#         # Parse date strings into datetime objects using dateutil.parser
#         date_from = parser.parse(date_from_raw).date()
#         date_to = parser.parse(date_to_raw).date()

#         date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)
        
#         # Fetch organization name from database
#         with connection.cursor() as cursor:
#             cursor.execute("""
#                 SELECT organization_name
#                 FROM saasapp_companyregistration
#                 WHERE user_id = %s
#             """, [tenant])
#             row = cursor.fetchone()
#             organization_name = row[0] if row else None

#         # Fetch machine data along with their creation times using cursor
#         report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to), assigned_customer__isnull=False)
#         serialized_data = [{
#             'sr_no': index + 1,  # Serial number
#             'name': item.assigned_customer.name if item.assigned_customer else None,
#             'mobile': item.assigned_customer.mobile_no if item.assigned_customer else None,
#             'email': item.assigned_customer.email if item.assigned_customer else None,
#             'machine_id': item.machine.machine_id if item.machine else None,
#             'machine_assigned_date': item.created_at.strftime('%Y-%m-%d'),
#         } for index, item in enumerate(report_data)]

#         # Create a new Excel workbook
#         workbook = Workbook()
#         worksheet = workbook.active

#         # Add the organization name at the top of the worksheet
#         if organization_name:
#             worksheet.merge_cells('C1:D2')  # Merge cells for the organization name
#             cell = worksheet['C1']  # Get the top-left cell of the merged range
#             cell.value = organization_name
#             cell.font = Font(bold=True)
#             cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#         # Add the "Customer Machine Report" label in the middle of the third row
#         worksheet.merge_cells('C3:D3')  # Merge cells for the label
#         cell = worksheet['C3']  # Get the top-left cell of the merged range
#         cell.value = "Customer Machine Report"
#         cell.font = Font(bold=True)  # Make the label bold
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#         # Merge cells for "FROM" and "TO" dates in cell E1
#         worksheet.merge_cells('E1:E3')  # Merge cells for the date range
#         date_range_cell = worksheet['E1']  # Get the top-left cell of the merged range
#         date_range_cell.value = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
#         date_range_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#         border_bold = styles.Border(left=styles.Side(style='medium'), 
#                                      right=styles.Side(style='medium'), 
#                                      top=styles.Side(style='medium'), 
#                                      bottom=styles.Side(style='medium'))

#         # Add borders to the cells in the header row
#         # Define the headers list
#         headers = ["Sr No", "Name", "Mobile", "Email", "Machine ID", "Machine Assigned Date"]
#         for col_num, header in enumerate(headers, start=1):
#             worksheet.cell(row=4, column=col_num, value=header).font = Font(bold=True)  # Make headers bold
#             worksheet.cell(row=4, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
        
#         # If data exists, populate the worksheet starting from row 5
#         for row_num, row_data in enumerate(serialized_data, start=5):
#             for col_num, value in enumerate(row_data.values(), start=1):
#                 worksheet.cell(row=row_num, column=col_num,value = value).alignment=Alignment(horizontal='center', vertical='center')

#         for col in worksheet.columns:
#             max_length = 0
#             column = col[0].column  # Get the column index
#             column_letter = chr(65 + column - 1)  # Convert column index to letter
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2
#             worksheet.column_dimensions[column_letter].width = adjusted_width

#         # Set the response headers for file download
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

#         # Write the Excel workbook to the response
#         workbook.save(response)

#         return response

#     except Exception as e:
#         return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
from openpyxl.styles import Font, Alignment, Border, Side

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_download_excel_customer(request):
    try:
        # Retrieve tenant information based on the logged-in user
        tenant = request.tenant.user_id
        today_date = timezone.localdate()

        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)
        
        # Fetch organization name from database
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT organization_name
                FROM saasapp_companyregistration
                WHERE user_id = %s
            """, [tenant])
            row = cursor.fetchone()
            organization_name = row[0] if row else None

        # Fetch machine data along with their creation times using cursor
        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to), assigned_customer__isnull=False)
        serialized_data = [{
            'sr_no': index + 1,  # Serial number
            'name': item.assigned_customer.name if item.assigned_customer else None,
            'mobile': item.assigned_customer.mobile_no if item.assigned_customer else None,
            'email': item.assigned_customer.email if item.assigned_customer else None,
            'machine_id': item.machine.machine_id if item.machine else None,
            'machine_assigned_date': item.created_at.strftime('%Y-%m-%d'),
        } for index, item in enumerate(report_data)]

        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active

        # Add the organization name at the top of the worksheet
        if organization_name:
            worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
            cell = worksheet['A4']  # Get the top-left cell of the merged range
            cell.value = f"Organization Name: {organization_name}"
            cell.font = Font(bold=True,size=12)
            cell.alignment = Alignment(wrap_text=True)  # Align text to the center

        # Add the "Customer Machine Report" label in the middle of the third row
        worksheet.merge_cells('A1:F2')  # Merge cells for the label
        cell = worksheet['A1']  # Get the top-left cell of the merged range
        cell.value = "Customer Machine Report"
        cell.font = Font(bold=True,size=16)  # Make the label bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        worksheet.merge_cells('A3:F3')  # Merge cells for the date range
        date_cell = worksheet['A3']  # Get the top-left cell of the merged range
        date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
       
        # Add borders to the cells in the header row
        # Define the headers list
        # Add borders to the cells in the header row
        headers = ["Sr No", "Name", "Mobile", "Email", "Machine ID", "Machine Assigned Date"]
        for col_num, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=5, column=col_num, value=header)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), 
                                        left=Side(style='medium'), right=Side(style='medium'))

        # If data exists, populate the worksheet starting from row 6
        for row_num, row_data in enumerate(serialized_data, start=6):
            for col_num, value in enumerate(row_data.values(), start=1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), 
                                     left=Side(style='thin'), right=Side(style='thin'))

        # for col in worksheet.columns:
        #     max_length = 0
        #     column = col[0].column  # Get the column index
        #     column_letter = chr(65 + column - 1)  # Convert column index to letter
        #     for cell in col:
        #         try:
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     worksheet.column_dimensions[column_letter].width = adjusted_width
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            column_letter = chr(64 + column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set a fixed width for the "Sr No" column
        worksheet.column_dimensions['A'].width = 10

        # Set the response headers for file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # Write the Excel workbook to the response
        workbook.save(response)

        return response

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def download_stock_data(request):
    # Get the date range from request parameters or set default dates
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if not from_date or not to_date:
        return Response({'success': False, 'message': 'Please provide both from_date and to_date parameters'})

    # Fetch all entries within the specified date range where stock equals capacity
    queryset = MStatus.objects.filter(
        created_at__date__gte=from_date,
        created_at__date__lte=to_date,
        stock=F('capacity')  # Filter where stock equals capacity
    ).values('m_id', 'created_at__date', 'stock').annotate(
        count_capacity_equals_stock=Count('id')
    )
from openpyxl.styles import Font, Alignment
from django.db.models import  F
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def admin_download_excel_refill(request):
#     tenant = request.tenant.user_id
#     date_from_raw = request.GET.get('from_date')
#     date_to_raw = request.GET.get('to_date')

#     if not date_from_raw or not date_to_raw:
#         return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

#     date_from = parser.parse(date_from_raw).date()
#     date_to = parser.parse(date_to_raw).date()
#     date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

#     machine_data = []
#     serial_number = 1  # Initialize serial number counter

#     all_machine_ids = MStatus.objects.filter(created_at__date__range=[date_from, date_to]).values_list('m_id', flat=True).distinct()
#     for machine_id in all_machine_ids:
#         last_entry = MStatus.objects.filter(m_id=machine_id, created_at__date__range=[date_from, date_to]).order_by('-id').first()

#         if last_entry:
#             stock_before_refill = last_entry.stock
#             # Count instances where stock is equal to capacity for the current machine per day
#             stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

#             stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
#             if stock_after_refill_qs.exists():
#                 stock_after_refill = stock_after_refill_qs.first().stock
#             else:
#                 stock_after_refill = None

#             last_second_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id')[1:2].first()
#             if last_second_entry:
#                 last_second_stock = last_second_entry.stock
#             else:
#                 last_second_stock = None

#             if stock_after_refill is not None:
#                 refill_quantity = stock_after_refill - last_second_stock
#             else:
#                 refill_quantity = None

#             # Fetch refill date from MStatus model
#             refill_date = last_entry.created_at.date() if last_entry.created_at else None

#             # Fetch location from MachineMaster model
#             machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
#             location = machine_master_data.installation_location if machine_master_data else None

#             # Add data for the current machine to the list
#             machine_data.append({
#                 'sr_no': serial_number,
#                 "machine_id": machine_id,
#                 "stock_before_refill": last_second_stock,
#                 "refill_quantity": refill_quantity,
#                 "stock_after_refill": stock_after_refill,
#                 "stock_capacity_equal_count": stock_capacity_equal_count,
#                 'location': location,
#                 "refill_date": refill_date,
#             })
#             if stock_after_refill == last_entry.capacity:
#                 machine_data.append({
#                     "sr_no": serial_number + 1,
#                     "machine_id": machine_id,
#                     "stock_before_refill": last_entry.capacity,
#                     "refill_quantity": 0,
#                     "stock_after_refill": 0,
#                     "stock_capacity_equal_count": stock_capacity_equal_count,
#                     'location': location,
#                     "refill_date": refill_date,
#                 })
#             serial_number += 2  # Increment serial number by 2 if both entries are added
#         else:
#             # If no entry exists for the machine, set both stock values to None
#             machine_data.append({
#                 'sr_no': serial_number,
#                 "machine_id": machine_id,
#                 "stock_before_refill": None,
#                 "stock_after_refill": None,
#                 "stock_capacity_equal_count": None,  # No entries, so count is 0
#                 "refill_quantity": None,
#                 'location': location,
#                 "refill_date": None,
#             })
#             serial_number += 1  # Increment serial number even if no data is added

#     # Calculate grand totals
    # grand_total = {
    #     "stock_before_refill": sum(data.get("stock_before_refill", 0) or 0 for data in machine_data),
    #     "refill_quantity": sum(data.get("refill_quantity", 0) or 0 for data in machine_data),
    #     "stock_after_refill": sum(data.get("stock_after_refill", 0) or 0 for data in machine_data),
    #     "stock_capacity_equal_count": sum(data.get("stock_capacity_equal_count", 0) or 0 for data in machine_data),
    # }

    # organization_name = None
    # with connection.cursor() as cursor:
    #     cursor.execute("""
    #         SELECT organization_name
    #         FROM saasapp_companyregistration
    #         WHERE user_id = %s
    #     """, [tenant])
    #     row = cursor.fetchone()
    #     organization_name = row[0] if row else None

    # workbook = Workbook()
    # worksheet = workbook.active

    # if organization_name:
    #     worksheet.merge_cells('C1:D2')
    #     cell = worksheet['C1']
    #     cell.value = organization_name
    #     cell.font = Font(bold=True)
    #     cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # worksheet.merge_cells('C3:D3')
    # cell = worksheet['C3']
    # cell.value = "Refill Machine Report"
    # cell.font = Font(bold=True)
    # cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # worksheet.merge_cells('E1:E3')
    # date_range_cell = worksheet['E1']
    # date_range_cell.value = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
    # date_range_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     headers = ["Sr No", "Machine ID", "Stock Before Refill", "Refill Quantity", "Stock After Refill", "Refill Count", 'Location',
#                "Refill Date"]
#     for col, header in enumerate(headers, start=1):
#         worksheet.cell(row=4, column=col, value=header).font = Font(bold=True)
#         worksheet.cell(row=4, column=col).alignment = Alignment(horizontal='center', vertical='center')

#     for row_num, data in enumerate(machine_data, start=5):
#         for col_num, value in enumerate(data.values(), start=1):
#             worksheet.cell(row=row_num, column=col_num, value=value).alignment = Alignment(horizontal='center', vertical='center')

    # # Add grand total row
    # grand_total_row = len(machine_data) + 6  # Add 6 for headers and blank row
    # for col_num, value in enumerate(grand_total.values(), start=3):
    #     worksheet.cell(row=grand_total_row, column=col_num, value=value).alignment = Alignment(horizontal='center', vertical='center')

    # # Add label for grand total row
    # worksheet.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
    # label_cell = worksheet.cell(row=grand_total_row, column=1, value="Grand Total")
    # label_cell.font = Font(bold=True)
    # label_cell.alignment = Alignment(horizontal='right', vertical='center')
    
#     for col in worksheet.columns:
#         max_length = 0
#         column = col[0].column
#         column_letter = chr(65 + column - 1)
#         for cell in col:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(cell.value)
#             except:
#                 pass
#         adjusted_width = (max_length + 2) * 1.2
#         worksheet.column_dimensions[column_letter].width = adjusted_width

#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

#     workbook.save(response)

#     return response

    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_pdf_customer(request):
    try:
        # Retrieve tenant information based on the logged-in user
        tenant = request.tenant
        today_date = datetime.now().date()
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return HttpResponse({'success': 0, 'message': 'Both from and to dates are required.'}, status=400)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        # Get all machine IDs from MachineMaster
        machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

        # Fetch organization name from database
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT organization_name
                FROM saasapp_companyregistration
                WHERE user_id = %s
            """, [tenant.id])
            row = cursor.fetchone()
            organization_name = row[0] if row else None

        # Fetch machine data along with their creation times
        report_data = MachineUserMapping.objects.filter(
            created_at__range=(date_from, date_to),
            assigned_customer__isnull=False
        )

        # Create PDF using ReportLab
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Add logo, organization name, and date range in a single row
        header_data = []
       

        if organization_name:
            org_name_para = Paragraph(organization_name, styles['Heading1'])
            header_data.append(org_name_para)

        date_range = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
        date_range_para = Paragraph(date_range, styles['Normal'])
        header_data.append(date_range_para)

        header_table = Table([header_data], colWidths=[1 * inch, None, 2 * inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(header_table)
        elements.append(Spacer(1, 12))  # Add vertical spacing

        # Add table for report data
        data = [['Name', 'Mobile', 'Email', 'Machine ID', 'Machine Assigned Date']]
        for item in report_data:
            data.append([
                item.assigned_customer.name if item.assigned_customer else '',
                item.assigned_customer.mobile_no if item.assigned_customer else '',
                item.assigned_customer.email if item.assigned_customer else '',
                item.machine.machine_id if item.machine else '',
                item.created_at.strftime('%Y-%m-%d') if item.created_at else ''
            ])

        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), (0.7, 0.7, 0.7)),
            ('TEXTCOLOR', (0, 0), (-1, 0), (1, 1, 1)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), (0.9, 0.9, 0.9)),
            ('GRID', (0, 0), (-1, -1), 1, (0.7, 0.7, 0.7))
        ]))
        elements.append(t)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        # Set the response headers for file download
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report.pdf"'

        # Write the PDF content to the response
        response.write(buffer.getvalue())
        return response

    except Exception as e:
        logger.error("An error occurred while generating PDF: %s", str(e))
        return HttpResponse({'success': 0, 'message': 'An error occurred while generating PDF.'}, status=500)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_pdf_user(request):
    try:
        # Retrieve tenant information based on the logged-in user
        tenant = request.tenant
        today_date = datetime.now().date()
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return HttpResponse({'success': 0, 'message': 'Both from and to dates are required.'}, status=400)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        # Get all machine IDs from MachineMaster
        machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

        # Fetch organization name from database
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT organization_name
                FROM saasapp_companyregistration
                WHERE user_id = %s
            """, [tenant.id])
            row = cursor.fetchone()
            organization_name = row[0] if row else None

        # Fetch machine data along with their creation times
        report_data = MachineUserMapping.objects.filter(
            created_at__range=(date_from, date_to),
            assigned_customer__isnull=True
        )

        # Create PDF using ReportLab
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Add logo, organization name, and date range in a single row
        header_data = []
        

        if organization_name:
            org_name_para = Paragraph(organization_name, styles['Heading1'])
            header_data.append(org_name_para)

        date_range = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
        date_range_para = Paragraph(date_range, styles['Normal'])
        header_data.append(date_range_para)

        header_table = Table([header_data], colWidths=[1 * inch, None, 2 * inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(header_table)
        elements.append(Spacer(1, 12))  # Add vertical spacing

        # Add table for report data
        data = [['Name', 'Mobile', 'Email', 'Machine ID', 'Machine Assigned Date']]
        for item in report_data:
            data.append([
                item.assigned_user.name if item.assigned_user else '',
                item.assigned_user.mobile_no if item.assigned_user else '',
                item.assigned_user.email if item.assigned_user else '',
                item.machine.machine_id if item.machine else '',
                item.created_at.strftime('%Y-%m-%d') if item.created_at else ''
            ])

        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), (0.7, 0.7, 0.7)),
            ('TEXTCOLOR', (0, 0), (-1, 0), (1, 1, 1)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), (0.9, 0.9, 0.9)),
            ('GRID', (0, 0), (-1, -1), 1, (0.7, 0.7, 0.7))
        ]))
        elements.append(t)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        # Set the response headers for file download
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report.pdf"'

        # Write the PDF content to the response
        response.write(buffer.getvalue())
        return response

    except Exception as e:
        logger.error("An error occurred while generating PDF: %s", str(e))
        return HttpResponse({'success': 0, 'message': 'An error occurred while generating PDF.'}, status=500)
from reportlab.lib import colors


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_download_pdf_refill(request):
    tenant = request.tenant
    # Get the from_date and to_date from the request data
    date_from_raw = request.GET.get('from_date')
    date_to_raw = request.GET.get('to_date')

    # Validate input data
    if not date_from_raw or not date_to_raw:
        return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Parse date strings into datetime objects using dateutil.parser
    date_from = parser.parse(date_from_raw).date()
    date_to = parser.parse(date_to_raw).date()

    machine_data = []

    # Fetch all distinct machine IDs within the date range
    all_machine_ids = MStatus.objects.filter(created_at__date__range=[date_from, date_to]).values_list('m_id', flat=True).distinct()

    for machine_id in all_machine_ids:
        # Fetch the last entry for the current machine sorted by id
        last_entry = MStatus.objects.filter(m_id=machine_id, created_at__date__range=[date_from, date_to]).order_by('-id').first()

        if last_entry:
            # Fetch the second-last entry for the current machine sorted by id
            second_last_entry = MStatus.objects.filter(m_id=machine_id, id__lt=last_entry.id).order_by('-id').first()

            # If both last and second-last entries exist
            if second_last_entry:
                stock_before_refill = second_last_entry.stock
            else:
                stock_before_refill = None

            stock_after_refill = last_entry.stock

            # Count instances where stock is equal to capacity for the current machine
            stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

            # Calculate refill stock if stock and capacity are equal
            if stock_after_refill == last_entry.capacity:
                refill_stock = stock_after_refill - stock_before_refill
            else:
                refill_stock = None

            # Extract month and day from the date
            month = last_entry.created_at.month
            day = last_entry.created_at.day

            # Calculate per day and per month counts within the date range
            per_day_count = MStatus.objects.filter(m_id=machine_id, created_at__date__range=[date_from, date_to], created_at__day=day).count()
            per_month_count = MStatus.objects.filter(m_id=machine_id, created_at__date__range=[date_from, date_to], created_at__month=month).count()

            machine_data.append({
                "machine_id": machine_id,
                "date": last_entry.created_at.date().isoformat(),
                "month": month,
                "day": day,
                "stock_before_refill": stock_before_refill,
                "stock_after_refill": stock_after_refill,
                "stock_capacity_equal_count": stock_capacity_equal_count,
                "refill_stock": refill_stock,
                "per_day_count": per_day_count,
                "per_month_count": per_month_count,
            })
        else:
            # If no entry exists for the machine within the date range, set default values
            machine_data.append({
                "machine_id": machine_id,
                "date": None,
                "month": None,
                "day": None,
                "stock_before_refill": None,
                "stock_after_refill": None,
                "stock_capacity_equal_count": 0,  # No entries, so count is 0
                "refill_stock": None,
                "per_day_count": 0,
                "per_month_count": 0,
            })

    # Fetch organization name from the database
    organization_name = None
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT organization_name
            FROM saasapp_companyregistration
            WHERE user_id = %s
        """, [tenant.id])
        row = cursor.fetchone()
        organization_name = row[0] if row else None

    # Create response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Add organization name
    if organization_name:
        elements.append(Paragraph(organization_name, styles['Title']))

    # Add date range
    elements.append(Paragraph(f"Date Range: {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}", styles['Title']))

    # Add table headers
    headers = ["Machine ID", "Date", "Month", "Day", "Stock Before Refill", "Stock After Refill",
               "Stock Capacity Equal Count", "Refill Stock", "Per Day Count", "Per Month Count"]
    
    # Transpose machine data
    transposed_data = [[header] for header in headers]
    for data in machine_data:
        for i, value in enumerate(data.values()):
            transposed_data[i].append(str(value))

    table = Table(transposed_data, colWidths=[100] * len(transposed_data), repeatRows=1)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(table)

    # Build PDF document
    doc.build(elements)

    return response

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_customer_machine_report_detail_by_id(request,customer_name):
# # def get_customer_machine_report_detail_by_id(request,customer_id):
#     try:
#         # Query the user (customer) with the specified ID
#         customer = User.objects.get(name=customer_name)
#         # customer = User.objects.get(id=customer_id)

#         print(customer,'id')
        
#         machine_data = []
        
#         # Query machine-customer mappings for the specified customer ID
#         machines = MachineUserMapping.objects.filter(assigned_customer=customer)
        
#         # Iterate through the machine-customer mappings and create a list of dictionaries
#         for machine in machines:
#             data = {
#                 'id':machine.id,
#                 'machine_id': machine.machine.machine_id,
#                 'customer_id': machine.assigned_customer.id,
#                 'customer_assigned_date': machine.assigned_customer_date,
#                 'customer_name': machine.assigned_customer.name,
#                 # Add other fields as needed
#             }
#             machine_data.append(data)


#         return Response({'success':1,'message':'Data Found','result':machine_data})
    
#     except MachineUserMapping.DoesNotExist:
#         # Handle the case where the machine ID does not exist
#         return Response({'success': 0, 'message': 'Machine ID does not exist'}, status=status.HTTP_404_NOT_FOUND)
    
#     except Exception as e:
#         # Handle other exceptions, log the error, and return an appropriate error response
#         return Response({'success': 0, 'message': f'Error retrieving machine details: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_machine_report_detail_by_id(request, customer_name):
    try:
        # Query users (customers) with the specified name
        customers = User.objects.filter(name=customer_name)
        if customers.exists():
            # Depending on your logic, you might want to handle multiple customers with the same name
            customer = customers.first()  # Selecting the first user with the given name

            machine_data = []
            
            # Query machine-customer mappings for the specified customer
            machines = MachineUserMapping.objects.filter(assigned_customer=customer)
            
            # Iterate through the machine-customer mappings and create a list of dictionaries
            for machine in machines:
                data = {
                    'id': machine.id,
                    'machine_id': machine.machine.machine_id,
                    'customer_id': machine.assigned_customer.id,
                    'customer_assigned_date': machine.assigned_customer_date,
                    'customer_name': machine.assigned_customer.name,
                    # Add other fields as needed
                }
                machine_data.append(data)

            return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})
        else:
            return Response({'success': 0, 'message': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)
    
    except MachineUserMapping.DoesNotExist:
        # Handle the case where the machine ID does not exist
        return Response({'success': 0, 'message': 'Machine ID does not exist'}, status=status.HTTP_404_NOT_FOUND)
    
    except Exception as e:
        # Handle other exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error retrieving machine details: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
import xlsxwriter


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_machine_report_detail_by_name(request, customer_name):
    # try:
        # Retrieve 'from' and 'to' date parameters from the request query parameters
        date_from_str = request.GET.get('from_date')
        date_to_str = request.GET.get('to_date')

        # Parse date strings into date objects
        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None
        
        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)
        # Query the user (customer) with the specified name
        customer = User.objects.get(name=customer_name)
        
        machine_data = []
        
        # Query machine-customer mappings for the specified customer
        machines = MachineUserMapping.objects.filter(assigned_customer=customer)

        # Apply date range filter if provided
        if date_from:
            machines = machines.filter(assigned_customer_date__gte=date_from)
        if date_to:
            machines = machines.filter(assigned_customer_date__lte=date_to)
        
        # Iterate through the machine-customer mappings and create a list of dictionaries
        for machine in machines:
            data = {
                'customer_name': machine.assigned_customer.name,
                'machine_id': machine.machine.machine_id,
                'customer_assigned_date': machine.assigned_customer_date.strftime('%Y-%m-%d'),  # Remove timezone information
                
                # Add other fields as needed
            }
            machine_data.append(data)

        # Create an Excel workbook
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customer_machine_report.xlsx"'
        workbook = xlsxwriter.Workbook(response)
        worksheet = workbook.add_worksheet()

       # Define bold format for headers
       
        bold_format = workbook.add_format({'bold': True})

        # Set alignment for all cells to center
        cell_format_center = workbook.add_format({'align': 'center', 'valign': 'vcenter'})

        # Merge cells for the report title and write it
        worksheet.merge_range('A1:E1', 'Customer Machine Report', bold_format)
        worksheet.write('A1', 'Customer Machine Report', bold_format)

        # Set alignment for the merged cell
        worksheet.set_row(0, None, cell_format_center)

        # Write customer name and date range
        worksheet.write('A2', 'Customer Name:', bold_format)
        worksheet.write('B2', customer_name)
        worksheet.write('D2', 'Report Date From:', bold_format)
        worksheet.write('E2', date_from.strftime('%Y-%m-%d') if date_from else '')

        # Set alignment for static headers
        worksheet.set_row(1, None, cell_format_center)  # Set alignment for the second row

        # Write headers
        headers = ['Sr No','Customer Name' ,'Machine ID', 'Customer Assigned Date']
        for col, header in enumerate(headers):
            worksheet.write(3, col, header, bold_format)
            worksheet.set_column(col, col, 20, cell_format_center)  # Set column width and alignment for headers

        # Write data
        sr_no = 1  # Initialize serial number counter
        for row, data in enumerate(machine_data, start=4):  # Start from row 3
            worksheet.write(row, 0, sr_no, cell_format_center)  # Write serial number with center alignment
            sr_no += 1  # Increment serial number for the next row
            for col, value in enumerate(data.values(), start=1):  # Start from column 1
                worksheet.write(row, col, value, cell_format_center)  # Write data with center alignment

        # Write Report Date To on the right side
        if date_to:
            worksheet.write('D3', 'Report Date To:', bold_format)
            worksheet.write('E3', date_to.strftime('%Y-%m-%d'), cell_format_center)

        # Close the workbook
        workbook.close()

        return response

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_stock_by_machine_user_role_refill_excel_download(request):
#     try:
        
#         date_from_raw = request.GET.get('from_date')
#         date_to_raw = request.GET.get('to_date')

#         # Validate input data
#         if not date_from_raw or not date_to_raw:
#             return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

#         # Parse date strings into datetime objects using dateutil.parser
#         date_from = parser.parse(date_from_raw).date()
#         date_to = parser.parse(date_to_raw).date()

#         date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

#         # Fetch all distinct machine IDs associated with the logged-in user
#         assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
        
#         serialized_data = [{'org_name': item.assigned_customer.organization,'location':item.machine.installation_location} for item in assigned_machines if item.assigned_customer and item.assigned_customer.organization]
#         print(serialized_data,'serialized_data')
#         machine_data = []

#         for machine in assigned_machines:
#             org=machine.assigned_user.organization
#             print(org)
#             queryset = MStatus.objects.filter(
#                 m_id=machine.machine.machine_id,
#                 created_at__date__range=[date_from, date_to],
#                 stock=F('capacity')
#             ).values('m_id', 'created_at__date', 'stock').annotate(
#                 count_capacity_equals_stock=Count('id')
#             )

#             for entry in queryset:
#                 stock_after_refill = entry['stock']
#                 stock_before_date = MStatus.objects.filter(
#                     m_id=machine.machine.machine_id,
#                     created_at__date__lt=entry['created_at__date']
#                 ).order_by('-created_at').values_list('stock', flat=True).first()

#                 refill_quantity = None
#                 if stock_after_refill is not None and stock_before_date is not None:
#                     refill_quantity = stock_after_refill - stock_before_date

#                 machine_data.append({
#                     'm_id': entry['m_id'],
#                     'date': entry['created_at__date'],
#                     'stock_before_refill': stock_before_date,
#                     'refill_quantity': refill_quantity,
#                     'stock_after_refill': entry['stock'],
#                     'stock_capacity_equal_count': entry['count_capacity_equals_stock'],
#                     'location':machine.machine.installation_location
#                 })

#         grand_total = {
#             "stock_before_refill": sum(data.get("stock_before_refill", 0) or 0 for data in machine_data),
#             "refill_quantity": sum(data.get("refill_quantity", 0) or 0 for data in machine_data),
#             "stock_after_refill": sum(data.get("stock_after_refill", 0) or 0 for data in machine_data),
#             "stock_capacity_equal_count": sum(data.get("stock_capacity_equal_count", 0) or 0 for data in machine_data),
#         }
        
#         # Create a new Excel workbook
#         workbook = Workbook()
#         worksheet = workbook.active

#         organization_name = org
        
#         if organization_name:
#             worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
#             cell = worksheet['A4']  # Get the top-left cell of the merged range
#             cell.value = f"Organization Name: {organization_name}"
#             cell.font = Font(bold=True,size=12)
#             cell.alignment = Alignment(wrap_text=True)  # Align text to the center

#         # Add the "Customer Machine Report" label in the middle of the third row
#         worksheet.merge_cells('A1:F2')  # Merge cells for the label
#         cell = worksheet['A1']  # Get the top-left cell of the merged range
#         cell.value = "Refill Machine Report"
#         cell.font = Font(bold=True,size=16)  # Make the label bold
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#         worksheet.merge_cells('A3:F3')  # Merge cells for the date range
#         date_cell = worksheet['A3']  # Get the top-left cell of the merged range
#         date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
#         date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
       
#         # Add headers for machine data
#         # Set up border styles
#         border_style = Side(border_style="thin")
#         border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

#         # Add headers for machine data
#         headers = ["Sr No", "Machine ID", "Before Refill Stock", "Refill Quantity", "After Refill Stock", "Refill Count", 'Location', "Refill Date"]
#         for col, header in enumerate(headers, start=1):
#             header_cell = worksheet.cell(row=5, column=col, value=header)
#             header_cell.font = Font(bold=True)  # Make headers bold
#             header_cell.alignment = Alignment(horizontal='center', vertical='center')
#             # Adding border to header cells
#             header_cell.border = border

#         # Write machine data to the worksheet
#         for row_num, data in enumerate(machine_data, start=6):
#             for col_num, key in enumerate(["m_id", "stock_before_refill", "refill_quantity", "stock_after_refill", "stock_capacity_equal_count", "location", "date"], start=1):
#                 cell = worksheet.cell(row=row_num, column=col_num, value=data.get(key, ""))
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 # Adding border to data cells
#                 cell.border = border

#         # Add grand total row
#         grand_total_row = len(machine_data) + 6  # Add 6 for headers and blank row
#         for col_num, value in enumerate(grand_total.values(), start=3):
#             cell = worksheet.cell(row=grand_total_row, column=col_num, value=value)
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#             # Adding border to grand total row cells
#             cell.border = border

#         # Add label for grand total row
#         worksheet.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
#         label_cell = worksheet.cell(row=grand_total_row, column=1, value="Grand Total")
#         label_cell.font = Font(bold=True)
#         label_cell.alignment = Alignment(horizontal='right', vertical='center')
        
#         # Adjust column widths based on content length
#         for col in worksheet.columns:
#             max_length = 0
#             column = col[0].column  # Get the column index
#             column_letter = chr(65 + column - 1)  # Convert column index to letter
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2
#             worksheet.column_dimensions[column_letter].width = adjusted_width

#         # Set the response headers for file download
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

#         # Write the Excel workbook to the response
#         workbook.save(response)

#         return response
#     except Exception as e:
#         # Handle exceptions, log the error, and return an appropriate error response
#         return Response({'success': 0, 'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_stock_by_machine_user_role_refill_excel_download(request):
#     try:
        
#         date_from_raw = request.GET.get('from_date')
#         date_to_raw = request.GET.get('to_date')

#         # Validate input data
#         if not date_from_raw or not date_to_raw:
#             return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

#         # Parse date strings into datetime objects using dateutil.parser
#         date_from = parser.parse(date_from_raw).date()
#         date_to = parser.parse(date_to_raw).date()

#         date_to = date_to + timedelta(days=1) - timedelta(seconds=1)

#         # Fetch all distinct machine IDs associated with the logged-in user
#         assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
        
#         serialized_data = [{'org_name': item.assigned_customer.organization,'location':item.machine.installation_location} for item in assigned_machines if item.assigned_customer and item.assigned_customer.organization]
#         print(serialized_data,'serialized_data')
#         machine_data = []

#         for machine in assigned_machines:
#             org = machine.assigned_user.organization
#             print(org)
#             queryset = MStatus.objects.filter(
#                 m_id=machine.machine.machine_id,
#                 created_at__date__range=[date_from, date_to],
#                 stock=F('capacity')
#             ).values('m_id', 'created_at__date', 'stock').annotate(
#                 count_capacity_equals_stock=Count('id')
#             )

#             for entry in queryset:
#                 stock_after_refill = entry['stock']
#                 stock_before_date = MStatus.objects.filter(
#                     m_id=machine.machine.machine_id,
#                     created_at__date__lt=entry['created_at__date']
#                 ).order_by('-created_at').values_list('stock', flat=True).first()

#                 refill_quantity = None
#                 if stock_after_refill is not None and stock_before_date is not None:
#                     refill_quantity = stock_after_refill - stock_before_date

#                 machine_data.append({
#                     'm_id': entry['m_id'],
#                     'date': entry['created_at__date'],
#                     'stock_before_refill': stock_before_date,
#                     'refill_quantity': refill_quantity,
#                     'stock_after_refill': entry['stock'],
#                     'stock_capacity_equal_count': entry['count_capacity_equals_stock'],
#                     'location': machine.machine.installation_location
#                 })

#         grand_total = {
#             "stock_before_refill": sum(data.get("stock_before_refill", 0) or 0 for data in machine_data),
#             "refill_quantity": sum(data.get("refill_quantity", 0) or 0 for data in machine_data),
#             "stock_after_refill": sum(data.get("stock_after_refill", 0) or 0 for data in machine_data),
#             "stock_capacity_equal_count": sum(data.get("stock_capacity_equal_count", 0) or 0 for data in machine_data),
#         }
        
#         # Create a new Excel workbook
#         workbook = Workbook()
#         worksheet = workbook.active

#         organization_name =  serialized_data[0].get('org_name') if serialized_data else "N/A"
        
#         if organization_name:
#             worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
#             cell = worksheet['A4']  # Get the top-left cell of the merged range
#             cell.value = f"Organization Name: {organization_name}"
#             cell.font = Font(bold=True, size=12)
#             cell.alignment = Alignment(wrap_text=True)  # Align text to the center

#         # Add the "Customer Machine Report" label in the middle of the third row
#         worksheet.merge_cells('A1:F2')  # Merge cells for the label
#         cell = worksheet['A1']  # Get the top-left cell of the merged range
#         cell.value = "Refill Machine Report"
#         cell.font = Font(bold=True, size=16)  # Make the label bold
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#         worksheet.merge_cells('A3:F3')  # Merge cells for the date range
#         date_cell = worksheet['A3']  # Get the top-left cell of the merged range
#         date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
#         date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
       
#         # Add headers for machine data
#         # Set up border styles
#         border_style = Side(border_style="thin")
#         border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

#         # Add headers for machine data
#         headers = ["Sr No", "Machine ID", "Before Refill Stock", "Refill Quantity", "After Refill Stock", "Refill Count", "Location", "Refill Date"]
#         for col, header in enumerate(headers, start=1):
#             header_cell = worksheet.cell(row=5, column=col, value=header)
#             header_cell.font = Font(bold=True)  # Make headers bold
#             header_cell.alignment = Alignment(horizontal='center', vertical='center')
#             # Adding border to header cells
#             header_cell.border = border

#         # Write machine data to the worksheet
#         if machine_data:
#             for row_num, data in enumerate(machine_data, start=6):
#                 for col_num, key in enumerate(["m_id", "stock_before_refill", "refill_quantity", "stock_after_refill", "stock_capacity_equal_count", "location", "date"], start=1):
#                     cell = worksheet.cell(row=row_num, column=col_num, value=data.get(key, ""))
#                     cell.alignment = Alignment(horizontal='center', vertical='center')
#                     # Adding border to data cells
#                     cell.border = border

#             # Add grand total row
#             grand_total_row = len(machine_data) + 6  # Add 6 for headers and blank row
#             for col_num, value in enumerate(grand_total.values(), start=3):
#                 cell = worksheet.cell(row=grand_total_row, column=col_num, value=value)
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 # Adding border to grand total row cells
#                 cell.border = border

#             # Add label for grand total row
#             worksheet.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
#             label_cell = worksheet.cell(row=grand_total_row, column=1, value="Grand Total")
#             label_cell.font = Font(bold=True)
#             label_cell.alignment = Alignment(horizontal='right', vertical='center')
#         # else:
#         #     # Ensure the worksheet has at least the headers even if there's no data
#         #     worksheet.append(['Sr No', 'Machine ID', 'Before Refill Stock', 'Refill Quantity', 'After Refill Stock', 'Refill Count', 'Location', 'Refill Date'])

#         # Adjust column widths based on content length
#         for col in worksheet.columns:
#             max_length = 0
#             column = col[0].column  # Get the column index
#             column_letter = chr(65 + column - 1)  # Convert column index to letter
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2
#             worksheet.column_dimensions[column_letter].width = adjusted_width

#         # Set the response headers for file download
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

#         # Write the Excel workbook to the response
#         workbook.save(response)

#         return response
#     except Exception as e:
#         # Handle exceptions, log the error, and return an appropriate error response
#         return Response({'success': 0, 'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stock_by_machine_user_role_refill_excel_download(request):
    try:
        
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timedelta(days=1) - timedelta(seconds=1)

        # Fetch all distinct machine IDs associated with the logged-in user
        assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
        
        serialized_data = [{'org_name': item.assigned_user.organization,'location':item.machine.installation_location} for item in assigned_machines if item.assigned_user and item.assigned_user.organization]
        
        machine_data = []

        for machine in assigned_machines:
            org = machine.assigned_user.organization
            
            queryset = MStatus.objects.filter(
                m_id=machine.machine.machine_id,
                created_at__date__range=[date_from, date_to],
                stock=F('capacity')
            ).values('m_id', 'created_at__date', 'stock').annotate(
                count_capacity_equals_stock=Count('id')
            )

            for entry in queryset:
                stock_after_refill = entry['stock']
                stock_before_date = MStatus.objects.filter(
                    m_id=machine.machine.machine_id,
                    created_at__date__lt=entry['created_at__date']
                ).order_by('-created_at').values_list('stock', flat=True).first()

                refill_quantity = None
                if stock_after_refill is not None and stock_before_date is not None:
                    refill_quantity = stock_after_refill - stock_before_date

                machine_data.append({
                    'm_id': entry['m_id'],
                    'date': entry['created_at__date'],
                    'stock_before_refill': stock_before_date,
                    'refill_quantity': refill_quantity,
                    'stock_after_refill': entry['stock'],
                    'stock_capacity_equal_count': entry['count_capacity_equals_stock'],
                    'location': machine.machine.installation_location
                })

        grand_total = {
            "stock_before_refill": sum(data.get("stock_before_refill", 0) or 0 for data in machine_data),
            "refill_quantity": sum(data.get("refill_quantity", 0) or 0 for data in machine_data),
            "stock_after_refill": sum(data.get("stock_after_refill", 0) or 0 for data in machine_data),
            "stock_capacity_equal_count": sum(data.get("stock_capacity_equal_count", 0) or 0 for data in machine_data),
        }
        
        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active

        organization_name = serialized_data[0].get('org_name') if serialized_data else "N/A"
        
        if organization_name:
            worksheet.merge_cells('A4:H4')  # Merge cells for the organization name
            cell = worksheet['A4']  # Get the top-left cell of the merged range
            cell.value = f"Organization Name: {organization_name}"
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)  # Align text to the center

        # Add the "Customer Machine Report" label in the middle of the third row
        worksheet.merge_cells('A1:H2')  # Merge cells for the label
        cell = worksheet['A1']  # Get the top-left cell of the merged range
        cell.value = "Refill Machine Report"
        cell.font = Font(bold=True, size=16)  # Make the label bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        worksheet.merge_cells('A3:H3')  # Merge cells for the date range
        date_cell = worksheet['A3']  # Get the top-left cell of the merged range
        date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
       
        # Add headers for machine data
        # Set up border styles
        border_style = Side(border_style="thin")
        border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

        # Add headers for machine data
        headers = ["Sr No", "Machine ID", "Before Refill Stock", "Refill Quantity", "After Refill Stock", "Refill Count", "Location", "Refill Date"]
        for col, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=5, column=col, value=header)
            header_cell.font = Font(bold=True)  # Make headers bold
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Adding border to header cells
            header_cell.border = border

        # Write machine data to the worksheet
        if machine_data:
            for row_num, data in enumerate(machine_data, start=6):
                worksheet.cell(row=row_num, column=1, value=row_num-5)  # Adding serial number
                for col_num, key in enumerate(["m_id", "stock_before_refill", "refill_quantity", "stock_after_refill", "stock_capacity_equal_count", "location", "date"], start=2):
                    cell = worksheet.cell(row=row_num, column=col_num, value=data.get(key, ""))
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Adding border to data cells
                    cell.border = border

            # Add grand total row
            grand_total_row = len(machine_data) + 6  # Add 6 for headers and blank row
            for col_num, value in enumerate(grand_total.values(), start=3):
                cell = worksheet.cell(row=grand_total_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Adding border to grand total row cells
                cell.border = border

            # Add label for grand total row
            worksheet.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
            label_cell = worksheet.cell(row=grand_total_row, column=1, value="Grand Total")
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='right', vertical='center')

        # # Ensure the worksheet has at least the headers even if there's no data
        # else:
        #     worksheet.append(['Sr No', 'Machine ID', 'Before Refill Stock', 'Refill Quantity', 'After Refill Stock', 'Refill Count', 'Location', 'Refill Date'])

        # Adjust column widths based on content length
        # for col in worksheet.columns:
        #     max_length = 0
        #     column = col[0].column  # Get the column index
        #     column_letter = chr(65 + column - 1)  # Convert column index to letter
        #     for cell in col:
        #         try:
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     worksheet.column_dimensions[column_letter].width = adjusted_width
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column index
            column_letter = chr(65 + column - 1)  # Convert column index to letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set the width for the "Sr No" column
        worksheet.column_dimensions['A'].width = 10
        # Set the response headers for file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # Write the Excel workbook to the response
        workbook.save(response)

        return response
    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_machine_mapping_report_refill(request):
    try:
        tenant = request.tenant.user_id
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Fetch all distinct machine IDs associated with the logged-in user
        assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
        all_machine_ids = [mapping.machine.machine_id for mapping in assigned_machines]

        machine_data = []

        for machine_id in all_machine_ids:
            # Fetch the last entry for the current machine sorted by id
            last_entry = MStatus.objects.filter(m_id=machine_id, created_at__range=[date_from, date_to]).order_by('-id').first()

            if last_entry:
                    stock_before_refill = last_entry.stock
                    # Count instances where stock is equal to capacity for the current machine per day
                    # stock_capacity_equal_count_per_day = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).annotate(
                    #     day=Trunc('created_at', 'day', output_field=DateField())
                    # ).values('day').annotate(count=Count('id')).order_by('day')
                    stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

                    stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
                    if stock_after_refill_qs.exists():
                        stock_after_refill = stock_after_refill_qs.first().stock
                    else:
                        stock_after_refill = None
                    

                    if stock_after_refill is not None:
                        refill_quantity = stock_after_refill - stock_before_refill
                    else:
                        refill_quantity = None

                    # Fetch refill date from MStatus model
                    refill_date = last_entry.created_at.date() if last_entry.created_at else None

                    # Fetch location from MachineMaster model
                    machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
                    location = machine_master_data.installation_location if machine_master_data else None

                    # Add data for the current machine to the list
                    machine_data.append({
                        "machine_id": machine_id,
                        "stock_before_refill": stock_before_refill,
                        "stock_after_refill": stock_after_refill,
                        "refill_quantity": refill_quantity,
                        "stock_capacity_equal_count": stock_capacity_equal_count,
                        'location':location,
                        "refill_date": refill_date,
                    })
            # else:
            #     # If no entry exists for the machine, set both stock values to None
            #     machine_data.append({
            #         "machine_id": None,
            #         "stock_before_refill": 0,
            #         "stock_after_refill": 0,
            #         "stock_capacity_equal_count": 0,  # No entries, so count is 0
            #     })

        # Check if machine_data is empty
        if not machine_data:
            return Response({'success': 0, "message": "No data found for the provided date range."})
        
        # Return the processed data
        return Response({'success': 1, 'message': 'Data processed successfully', 'result': machine_data})

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error processing data: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_payment_report(request):
    try:
        date_from_raw = request.GET.get('from_date')
        
        date_to_raw = request.GET.get('to_date')

        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()
        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Filter MachineUserMapping queryset based on the provided date range and user
        machine_user_mappings = MachineUserMapping.objects.filter(assigned_user=request.user.id)

        # Initialize a dictionary to hold mode counts per machine
        mode_counts_per_machine = defaultdict(lambda: defaultdict(int))

        # Iterate over each MachineUserMapping object
        for mapping in machine_user_mappings:
            # Retrieve machine_id from MachineUserMapping
            machine_id = mapping.machine.machine_id

            # Filter MStatus records for the current machine and date range
            mstatus_records = MStatus.objects.filter(m_id=machine_id, created_at__range=(date_from, date_to))

            # Count mode occurrences for the current machine
            for record in mstatus_records:
                if record.mode:
                    mode_counts_per_machine[machine_id][record.mode] += 1

        # Initialize an empty list to hold the result
        data_for_response = []

        # Get all machine IDs associated with the user
        user_machine_ids = set(mapping.machine.machine_id for mapping in machine_user_mappings)

        # Iterate over all machine IDs to prepare data for response
        for machine_id in user_machine_ids:
            modes = mode_counts_per_machine.get(machine_id, {})
            total_modes = sum(modes.values())
            total_coin = modes.get('COIN', 0) * 5
            total_qr = modes.get('QR', 0) * 5
            total_amount = total_coin + total_qr
            
            data_for_response.append({
                'machine_id': machine_id,
                'coin_mode_counts': modes.get('COIN', 0),
                'qr_mode_counts': modes.get('QR', 0),
                'total_modes': total_modes,
                'total_amount': total_amount
            })

        return Response({
            'success': 1,
            'message': 'Data found',
            'result': data_for_response
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def user_payment_report_excel_download(request):
#     try:
#         # Retrieve 'from' and 'to' date parameters from the GET request data
#         date_from_raw = request.GET.get('from_date')
#         date_to_raw = request.GET.get('to_date')

#         # Validate input data
#         if not date_from_raw or not date_to_raw:
#             return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

#         # Parse date strings into datetime objects using dateutil.parser
#         date_from = parser.parse(date_from_raw).date()
#         date_to = parser.parse(date_to_raw).date()

#         date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

#         # Filter MachineUserMapping queryset based on the provided date range and user
#         machine_user_mappings = MachineUserMapping.objects.filter(assigned_user=request.user)

#         # Initialize a dictionary to hold mode counts per machine
#         mode_counts_per_machine = defaultdict(lambda: defaultdict(int))

#         # Iterate over each MachineUserMapping object
#         for mapping in machine_user_mappings:
#             # Retrieve machine_id from MachineUserMapping
#             machine_id = mapping.machine.machine_id

#             # Retrieve the organization associated with the assigned user
#             org = mapping.assigned_user.organization if mapping.assigned_user else None

#             # Filter MStatus records for the current machine and date range
#             mstatus_records = MStatus.objects.filter(m_id=machine_id, created_at__range=(date_from, date_to))

#             # Count mode occurrences for the current machine
#             for record in mstatus_records:
#                 if record.mode:
#                     mode_counts_per_machine[machine_id][record.mode] += 1

#         # Initialize an empty list to hold the result
#         data_for_response = []

#         # Get all machine IDs associated with the user
#         user_machine_ids = set(mapping.machine.machine_id for mapping in machine_user_mappings)

#         # Iterate over all machine IDs to prepare data for response
#         for machine_id in user_machine_ids:
#             modes = mode_counts_per_machine.get(machine_id, {})
#             total_modes = sum(modes.values())
#             total_coin = modes.get('COIN', 0) * 10
#             total_qr = modes.get('QR', 0) * 10
#             total_amount = total_coin + total_qr
            
#             data_for_response.append({
#                 'machine_id': machine_id,
#                 'coin_mode_counts': modes.get('COIN', 0),
#                 'qr_mode_counts': modes.get('QR', 0),
#                 'total_modes': total_modes,
#                 'total_amount': total_amount,
#                 'organization': org  # Include organization information in the response
#             })

#         # Check if there is no data available
#         if not data_for_response:
#             # If no data available, return a response indicating the same
#             return Response({'success': 0, 'message': 'No data available for the specified date range.'}, status=status.HTTP_200_OK)

#         # Create a new Excel workbook
#         workbook = Workbook()
#         worksheet = workbook.active

#         # Add organization name
#         organization_name = data_for_response[0].get('organization')
#         # Add the organization name at the top of the worksheet
#         if organization_name:
#             worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
#             cell = worksheet['A4']  # Get the top-left cell of the merged range
#             cell.value = f"Organization Name: {organization_name}"
#             cell.font = Font(bold=True, size=12)
#             cell.alignment = Alignment(wrap_text=True)  # Align text to the center

#         # Add the "Customer Machine Report" label in the middle of the third row
#         worksheet.merge_cells('A1:F2')  # Merge cells for the label
#         cell = worksheet['A1']  # Get the top-left cell of the merged range
#         cell.value = "Payment Machine Report"
#         cell.font = Font(bold=True, size=16)  # Make the label bold
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#         worksheet.merge_cells('A3:F3')  # Merge cells for the date range
#         date_cell = worksheet['A3']  # Get the top-left cell of the merged range
#         date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
#         date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

#         # Add headers for machine data
#         headers = ["Sr No", "Machine ID", "Coin", "QR", "Total", "Total Amount"]
#         for col, header in enumerate(headers, start=1):
#             header_cell = worksheet.cell(row=5, column=col, value=header)
#             header_cell.font = Font(bold=True)  # Make headers bold
#             header_cell.alignment = Alignment(horizontal='center', vertical='center')
#             # Adding border to header cells
#             # border = Border(bottom=Side(border_style="thin"))
#             border_style = Side(border_style="thin")
#             border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
#             header_cell.border = border

#         # Write machine data to the worksheet
#         for row_num, data in enumerate(data_for_response, start=6):
#             worksheet.cell(row=row_num, column=1, value=row_num - 5).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=2, value=data['machine_id']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=3, value=data['coin_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=4, value=data['qr_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=5, value=data['total_modes']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=6, value=data['total_amount']).alignment = Alignment(horizontal='center', vertical='center')
#             # Adding border to data cells
#             for col in range(1, 7):
#                 cell = worksheet.cell(row=row_num, column=col)
#                 border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
#                 cell.border = border

#         # Adjust column widths based on content length
#         for col in worksheet.columns:
#             max_length = 0
#             column = col[0].column  # Get the column index
#             column_letter = chr(65 + column - 1)  # Convert column index to letter
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2
#             worksheet.column_dimensions[column_letter].width = adjusted_width

#         # Set the response headers for file download
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

#         # Write the Excel workbook to the response
#         workbook.save(response)

#         return response

#     except Exception as e:
#         # Handle exceptions, log the error, and return an appropriate error response
#         return Response({'success': 0, 'message': f'Error processing data: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_payment_report_excel_download(request):
    try:
        # Retrieve 'from' and 'to' date parameters from the GET request data
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()
        date_to = date_to + timedelta(days=1) - timedelta(seconds=1)

        # Filter MachineUserMapping queryset based on the provided date range and user
        machine_user_mappings = MachineUserMapping.objects.filter(assigned_user=request.user)

        # Initialize a dictionary to hold mode counts per machine
        mode_counts_per_machine = defaultdict(lambda: defaultdict(int))

        # Iterate over each MachineUserMapping object
        for mapping in machine_user_mappings:
            # Retrieve machine_id from MachineUserMapping
            machine_id = mapping.machine.machine_id

            # Retrieve the organization associated with the assigned user
            org = mapping.assigned_user.organization if mapping.assigned_user else None

            # Filter MStatus records for the current machine and date range
            mstatus_records = MStatus.objects.filter(m_id=machine_id, created_at__range=(date_from, date_to))

            # Count mode occurrences for the current machine
            for record in mstatus_records:
                if record.mode:
                    mode_counts_per_machine[machine_id][record.mode] += 1

        # Initialize an empty list to hold the result
        data_for_response = []

        # Get all machine IDs associated with the user
        user_machine_ids = set(mapping.machine.machine_id for mapping in machine_user_mappings)

        # Iterate over all machine IDs to prepare data for response
        for machine_id in user_machine_ids:
            modes = mode_counts_per_machine.get(machine_id, {})
            total_modes = sum(modes.values())
            total_coin = modes.get('COIN', 0) * 10
            total_qr = modes.get('QR', 0) * 10
            total_amount = total_coin + total_qr

            data_for_response.append({
                'machine_id': machine_id,
                'coin_mode_counts': modes.get('COIN', 0),
                'qr_mode_counts': modes.get('QR', 0),
                'total_modes': total_modes,
                'total_amount': total_amount,
                'organization': org  # Include organization information in the response
            })

        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active

        # Add organization name
        organization_name = data_for_response[0].get('organization') if data_for_response else "N/A"
        # Add the organization name at the top of the worksheet
        worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
        cell = worksheet['A4']  # Get the top-left cell of the merged range
        cell.value = f"Organization Name: {organization_name}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(wrap_text=True)  # Align text to the center

        # Add the "Customer Machine Report" label in the middle of the third row
        worksheet.merge_cells('A1:F2')  # Merge cells for the label
        cell = worksheet['A1']  # Get the top-left cell of the merged range
        cell.value = "Payment Machine Report"
        cell.font = Font(bold=True, size=16)  # Make the label bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        worksheet.merge_cells('A3:F3')  # Merge cells for the date range
        date_cell = worksheet['A3']  # Get the top-left cell of the merged range
        date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        # Add headers for machine data
        headers = ["Sr No", "Machine ID", "Coin", "QR", "Total", "Total Amount"]
        for col, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=5, column=col, value=header)
            header_cell.font = Font(bold=True)  # Make headers bold
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Adding border to header cells
            border_style = Side(border_style="thin")
            border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            header_cell.border = border

        # Write machine data to the worksheet
        if data_for_response:
            for row_num, data in enumerate(data_for_response, start=6):
                worksheet.cell(row=row_num, column=1, value=row_num - 5).alignment = Alignment(horizontal='center', vertical='center')
                worksheet.cell(row=row_num, column=2, value=data['machine_id']).alignment = Alignment(horizontal='center', vertical='center')
                worksheet.cell(row=row_num, column=3, value=data['coin_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
                worksheet.cell(row=row_num, column=4, value=data['qr_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
                worksheet.cell(row=row_num, column=5, value=data['total_modes']).alignment = Alignment(horizontal='center', vertical='center')
                worksheet.cell(row=row_num, column=6, value=data['total_amount']).alignment = Alignment(horizontal='center', vertical='center')
                # Adding border to data cells
                for col in range(1, 7):
                    cell = worksheet.cell(row=row_num, column=col)
                    border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
                    cell.border = border

        # Adjust column widths based on content length
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            column_letter = chr(64 + column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set a fixed width for the "Sr No" column
        worksheet.column_dimensions['A'].width = 10

        # Set the response headers for file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # Write the Excel workbook to the response
        workbook.save(response)

        return response

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error processing data: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customer_role_user_report_excel_download(request):
    try:
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()
        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        report_data = MachineUserMapping.objects.filter(assigned_user_date__range=(date_from, date_to), assigned_customer=request.user.id)

        serialized_data = []
        for item in report_data:
            if item.assigned_customer:
                serialized_item = {
                    'machine_id': item.machine.machine_id if item.machine else None,
                    'name': item.assigned_user.name,
                    'user_mobile_no': item.assigned_user.mobile_no,
                    'email': item.assigned_user.email,
                    'organization': item.assigned_customer.organization,
                    'assigned_machine_date': item.assigned_user_date.strftime('%Y-%m-%d')
                }
                serialized_data.append(serialized_item)
               

        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active
        
        
        # if serialized_data and serialized_data[0].get('organization'):
        #     worksheet.merge_cells('C1:D2')
        #     cell = worksheet['C1']
        #     cell.value = serialized_data[0]['organization']
        #     cell.font = Font(bold=True)
        #     cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # worksheet.merge_cells('C3:D3')
        # cell = worksheet['C3']
        # cell.value = "User Machine Report"
        # cell.font = Font(bold=True)
        # cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # worksheet.merge_cells('E1:E3')
        # date_range_cell = worksheet['E1']
        # date_range_cell.value = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
        # date_range_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if serialized_data and serialized_data[0].get('organization'):
            
            worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
            cell = worksheet['A4']  # Get the top-left cell of the merged range
            cell.value = f"Organization Name: {serialized_data[0]['organization']}"
            cell.font = Font(bold=True,size=12)
            cell.alignment = Alignment(wrap_text=True)  # Align text to the center

        # Add the "Customer Machine Report" label in the middle of the third row
        worksheet.merge_cells('A1:F2')  # Merge cells for the label
        cell = worksheet['A1']  # Get the top-left cell of the merged range
        cell.value = "User Machine Report"
        cell.font = Font(bold=True,size=16)  # Make the label bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        worksheet.merge_cells('A3:F3')  # Merge cells for the date range
        date_cell = worksheet['A3']  # Get the top-left cell of the merged range
        date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
       
        # Set up border styles
        border_style = Side(border_style="thin")
        border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

        # Add headers for machine data
        headers = ["Sr No", "Machine ID",  "Name", "Mobile No", "Email", "Assigned Date"]
        for col, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=5, column=col, value=header)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Adding border to header cells
            header_cell.border = border

        # Write machine data to the worksheet if available
        if serialized_data:
            for row_num, data in enumerate(serialized_data, start=6):
                # Adding Sr No
                worksheet.cell(row=row_num, column=1, value=row_num - 5).alignment = Alignment(horizontal='center', vertical='center')
                # Adding border to Sr No cell
                worksheet.cell(row=row_num, column=1).border = border
                
                for col_num, key in enumerate(["machine_id", "name", "user_mobile_no", "email", "assigned_machine_date"], start=2):
                    cell = worksheet.cell(row=row_num, column=col_num, value=data.get(key, ""))
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Adding border to data cells
                    cell.border = border

        # Adjust column widths based on content length
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            column_letter = chr(64 + column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set a fixed width for the "Sr No" column
        worksheet.column_dimensions['A'].width = 10

        # Set the response headers for file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # Write the Excel workbook to the response
        workbook.save(response)

        return response

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customer_role_payment_report_excel_download(request):
    try:
        tenant=request.tenant.user_id
        # Retrieve 'from' and 'to' date parameters from the POST request data
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()
        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Filter MachineUserMapping queryset based on the provided date range and user
        machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer=request.user.id)
        
        
        
        # Initialize a dictionary to hold mode counts per machine
        mode_counts_per_machine = defaultdict(lambda: defaultdict(int))

        # Iterate over each MachineUserMapping object
        for mapping in machine_user_mappings:
            # Retrieve machine_id from MachineUserMapping
            machine_id = mapping.machine.machine_id
            organization_name=mapping.assigned_customer.organization

            # Filter MStatus records for the current machine and date range
            mstatus_records = MStatus.objects.filter(m_id=machine_id, created_at__range=(date_from, date_to))

            # Count mode occurrences for the current machine
            for record in mstatus_records:
                if record.mode:
                    mode_counts_per_machine[machine_id][record.mode] += 1

        # Initialize an empty list to hold the result
        data_for_response = []

        # Iterate over mode counts per machine to prepare data for response
        for machine_id, modes in mode_counts_per_machine.items():
            total_modes = sum(modes.values())
            total_coin = modes.get('COIN', 0) * 10
            total_qr = modes.get('QR', 0) * 10
            total_amount = total_coin + total_qr
            
            data_for_response.append({
                'machine_id': machine_id,
                'coin_mode_counts': modes.get('COIN', 0),
                'qr_mode_counts': modes.get('QR', 0),
                'total_modes': total_modes,
                'total_amount': total_amount
            })

        

        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active

        # Add organization name
        if organization_name:
            worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
            cell = worksheet['A4']  # Get the top-left cell of the merged range
            cell.value = f"Organization Name: {organization_name}"
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)  # Align text to the center

        # Add the "Customer Machine Report" label in the middle of the third row
        worksheet.merge_cells('A1:F2')  # Merge cells for the label
        cell = worksheet['A1']  # Get the top-left cell of the merged range
        cell.value = "Payment Machine Report"
        cell.font = Font(bold=True, size=16)  # Make the label bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        worksheet.merge_cells('A3:F3')  # Merge cells for the date range
        date_cell = worksheet['A3']  # Get the top-left cell of the merged range
        date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        # Add headers for machine data
        headers = ["Sr No","Machine ID", "Coin", "QR", "Total", "Total Amount"]
        for col, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=5, column=col, value=header)
            header_cell.font = Font(bold=True)  # Make headers bold
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Adding border to header cells
            # border = Border(bottom=Side(border_style="thin"))
            border_style = Side(border_style="thin")
            border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            header_cell.border = border
        # Write machine data to the worksheet
        for row_num, data in enumerate(data_for_response, start=6):
            worksheet.cell(row=row_num, column=1, value=row_num - 4).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=2, value=data['machine_id']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=3, value=data['coin_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=4, value=data['qr_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=5, value=data['total_modes']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=6, value=data['total_amount']).alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(1, 7):
                cell = worksheet.cell(row=row_num, column=col)
                border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
                cell.border = border
        # Adjust column widths based on content length
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            column_letter = chr(64 + column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set a fixed width for the "Sr No" column
        worksheet.column_dimensions['A'].width = 10

        # Set the response headers for file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # Write the Excel workbook to the response
        workbook.save(response)

        return response
    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customer_refill_machine(request):
    try:
        
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Fetch all distinct machine IDs associated with the logged-in user
        assigned_machines = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')
        
        all_machine_ids = [mapping.machine.machine_id for mapping in assigned_machines]

        machine_data = []

        for machine_id in all_machine_ids:
            
            # Fetch the last entry for the current machine sorted by id
            last_entry = MStatus.objects.filter(m_id=machine_id, created_at__range=[date_from, date_to]).order_by('-id').first()
            
            if last_entry:
                stock_before_refill = last_entry.stock

                stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

                stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
                if stock_after_refill_qs.exists():
                    stock_after_refill = stock_after_refill_qs.first().stock
                else:
                    stock_after_refill = None
                

                if stock_after_refill is not None:
                    refill_quantity = stock_after_refill - stock_before_refill
                else:
                    refill_quantity = None

                # Fetch refill date from MStatus model
                refill_date = last_entry.created_at.date() if last_entry.created_at else None

                # Fetch location from MachineMaster model
                machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
                location = machine_master_data.installation_location if machine_master_data else None


                # Add data for the current machine to the list
                machine_data.append({
                    "machine_id": machine_id,
                    "stock_before_refill": stock_before_refill,
                    "refill_quantity": refill_quantity,
                    "stock_after_refill": stock_after_refill,
                    "stock_capacity_equal_count": stock_capacity_equal_count,
                    'location':location,
                    "refill_date": refill_date,
                })
            # else:
            #     # If no entry exists for the machine, set both stock values to None
            #     machine_data.append({
            #         "machine_id": None,
            #         "stock_before_refill": None,
            #         "refill_quantity": None,
            #         "stock_after_refill": None,
            #         "stock_capacity_equal_count": None,  # No entries, so count is 0
            #     })

        # Check if machine_data is empty
        if not machine_data:
            return Response({'success': 0, "message": "No data found for the provided date range."})
        

        return Response({'success': 1, 'message': 'Data Found','result':machine_data})

    
    except Exception as e:
        import inspect,traceback
        # Get the name of the current function
        current_function = inspect.currentframe().f_code.co_name
        # Log the function name along with the error message
        logger.error(f"Error in function {current_function}: {str(e)}")
        logger.error(traceback.format_exc())  # Log the traceback
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error processing data: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def customer_refill_machine_report_download(request):
#     try:
        
#         date_from_raw = request.GET.get('from_date')
#         date_to_raw = request.GET.get('to_date')

#         # Validate input data
#         if not date_from_raw or not date_to_raw:
#             return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

#         # Parse date strings into datetime objects using dateutil.parser
#         date_from = parser.parse(date_from_raw).date()
#         date_to = parser.parse(date_to_raw).date()

#         date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

#         # Fetch all distinct machine IDs associated with the logged-in user
#         assigned_machines = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')

#         serialized_data = []
#         for item in assigned_machines:
#             if item.assigned_user:
#                 organization_name = item.assigned_user.organization
#                 if organization_name:
#                     serialized_data.append({'org_name': organization_name})

#         all_machine_ids = [mapping.machine.machine_id for mapping in assigned_machines]

#         machine_data = []

#         for machine_id in all_machine_ids:
#             # Fetch the last entry for the current machine sorted by id
#             last_entry = MStatus.objects.filter(m_id=machine_id, created_at__range=[date_from, date_to]).order_by('-id').first()

#             if last_entry:
#                 stock_before_refill = last_entry.stock
#                 # Count instances where stock is equal to capacity for the current machine per day
#                 stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'),created_at__range=[date_from, date_to]).count()

#                 stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'),created_at__range=[date_from, date_to])
#                 if stock_after_refill_qs.exists():
#                     stock_after_refill = stock_after_refill_qs.first().stock
#                 else:
#                     stock_after_refill = None

#                 if stock_after_refill is not None:
#                     refill_quantity = stock_after_refill - stock_before_refill
#                 else:
#                     refill_quantity = None

#                 # Fetch refill date from MStatus model
#                 refill_date = last_entry.created_at.date() if last_entry.created_at else None

#                 # Fetch location from MachineMaster model
#                 machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
#                 location = machine_master_data.installation_location if machine_master_data else None


#                 # Add data for the current machine to the list
#                 machine_data.append({
#                     "machine_id": machine_id,
#                     "stock_before_refill": stock_before_refill,
#                     "refill_quantity": refill_quantity,
#                     "stock_after_refill": stock_after_refill,
#                     "stock_capacity_equal_count": stock_capacity_equal_count,
#                     'location':location,
#                     "refill_date": refill_date,
#                 })
            
#         # Calculate grand totals
#         grand_total = {
#             "stock_before_refill": sum(data.get("stock_before_refill", 0) or 0 for data in machine_data),
#             "refill_quantity": sum(data.get("refill_quantity", 0) or 0 for data in machine_data),
#             "stock_after_refill": sum(data.get("stock_after_refill", 0) or 0 for data in machine_data),
#             "stock_capacity_equal_count": sum(data.get("stock_capacity_equal_count", 0) or 0 for data in machine_data),
#             # "per_day_count": sum(data.get("per_day_count", 0) or 0 for data in machine_data),
#             # "per_month_count": sum(data.get("per_month_count", 0) or 0 for data in machine_data),
#         }
#         # Check if machine_data is empty
#         if not machine_data:
#             return Response({'success': 0, "message": "No data found for the provided date range."})

#         # Create a new Excel workbook
#         workbook = Workbook()
#         worksheet = workbook.active

#         # Fetch organization name from the serialized data
#         organization_name = serialized_data[0].get('org_name') if serialized_data else None
#         if organization_name:
#             worksheet.merge_cells('C1:D2')
#             cell = worksheet['C1']
#             cell.value = organization_name
#             cell.font = Font(bold=True)
#             cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#         # Add the "Refill Machine Report" label
#         worksheet.merge_cells('C3:D3')
#         cell = worksheet['C3']
#         cell.value = "Refill Machine Report"
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#         # Merge cells for "FROM" and "TO" dates
#         worksheet.merge_cells('E1:E3')
#         date_range_cell = worksheet['E1']
#         date_range_cell.value = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
#         date_range_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#         # Add headers for machine data
#         headers = ["Sr No", "Machine ID", "Before Refill Stock", "Refill Quantity", "After Refill Stock", "Refill Count",'location',"refill_date"]
#         for col, header in enumerate(headers, start=1):
#             worksheet.cell(row=4, column=col, value=header).font = Font(bold=True)
#             worksheet.cell(row=4, column=col).alignment = Alignment(horizontal='center', vertical='center')

#         # Write machine data to the worksheet
#         for row_num, data in enumerate(machine_data, start=5):
#             worksheet.cell(row=row_num, column=1, value=row_num - 4).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=2, value=data['machine_id']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=3, value=data['stock_before_refill']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=4, value=data['refill_quantity']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=5, value=data['stock_after_refill']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=6, value=data['stock_capacity_equal_count']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=7, value=data['location']).alignment = Alignment(horizontal='center', vertical='center')
#             worksheet.cell(row=row_num, column=8, value=data['refill_date']).alignment = Alignment(horizontal='center', vertical='center')

#          # Add grand total row
#         grand_total_row = len(machine_data) + 6  # Add 6 for headers and blank row
#         for col_num, value in enumerate(grand_total.values(), start=3):
#             worksheet.cell(row=grand_total_row, column=col_num, value=value).alignment = Alignment(horizontal='center', vertical='center')

#         # Add label for grand total row
#         worksheet.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
#         label_cell = worksheet.cell(row=grand_total_row, column=1, value="Grand Total")
#         label_cell.font = Font(bold=True)
#         label_cell.alignment = Alignment(horizontal='right', vertical='center')

#         # Adjust column widths based on content length
#         for col in worksheet.columns:
#             max_length = 0
#             column = col[0].column
#             column_letter = chr(65 + column - 1)
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2
#             worksheet.column_dimensions[column_letter].width = adjusted_width

#         # Set the response headers for file download
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

#         # Write the Excel workbook to the response
#         workbook.save(response)

#         return response

#     except Exception as e:
#         # Handle exceptions, log the error, and return an appropriate error response
#         return Response({'success': 0, 'message': f'Error processing data: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_download_excel_refill(request,machine_id):
    tenant = request.tenant.user_id
    date_from_raw = request.GET.get('from_date')
    date_to_raw = request.GET.get('to_date')

    if not date_from_raw or not date_to_raw:
        return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    date_from = parser.parse(date_from_raw).date()
    date_to = parser.parse(date_to_raw).date()
    date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)


    all_machine_ids=MStatus.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        m_id=machine_id,
        stock=F('capacity')  # Filter where stock capacity equals stock
    ).values('m_id', 'created_at__date', 'stock').annotate(
        count_capacity_equals_stock=Count('id')
    )

    machine_data = []
    serial_number = 1
    for entry in all_machine_ids:
        stock_after_refill = entry['stock']

        stock_before_date = MStatus.objects.filter(
            m_id=machine_id,
            created_at__date__lt=entry['created_at__date']
        ).order_by('-created_at').values_list('stock', flat=True).first()
        
        refill_quantity = None
        if stock_after_refill is not None and stock_before_date is not None:
            # Calculate refill quantity if stock_after_refill and stock_before_date are not None
            refill_quantity = stock_after_refill - stock_before_date

        # Fetch location from MachineMaster model
        machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
        location = machine_master_data.installation_location if machine_master_data else None

        # Append data to the result
        machine_data.append({
            "sr_no": serial_number,
            'date': entry['created_at__date'],
            'm_id': entry['m_id'],
            'stock_before_date': stock_before_date,
            'refill_quantity': refill_quantity,
            'stock': entry['stock'],
            'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
            'location': location,
        })

        serial_number += 1
    grand_total = {
        "stock_before_date": sum(data.get("stock_before_date", 0) or 0 for data in machine_data),
        "refill_quantity": sum(data.get("refill_quantity", 0) or 0 for data in machine_data),
        "stock": sum(data.get("stock", 0) or 0 for data in machine_data),
        "count_capacity_equals_stock": sum(data.get("count_capacity_equals_stock", 0) or 0 for data in machine_data),
    }

    organization_name = None
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT organization_name
            FROM saasapp_companyregistration
            WHERE user_id = %s
        """, [tenant])
        row = cursor.fetchone()
        organization_name = row[0] if row else None

    workbook = Workbook()
    worksheet = workbook.active

    if organization_name:
        worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
        cell = worksheet['A4']  # Get the top-left cell of the merged range
        cell.value = f"Organization Name: {organization_name}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(wrap_text=True)  # Align text to the center
        # cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    # Add the "Refill Machine Report" label in the middle of the third row
    worksheet.merge_cells('A1:F2')  # Merge cells for the label
    cell = worksheet['A1']  # Get the top-left cell of the merged range
    cell.value = "Refill Machine Report"
    cell.font = Font(bold=True, size=16)  # Make the label bold
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
    # cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    worksheet.merge_cells('A3:F3')  # Merge cells for the date range
    date_cell = worksheet['A3']  # Get the top-left cell of the merged range
    date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
    date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
    # date_cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), bottom=Side(style='medium'))

    headers = ["Sr No","Refill Date","Machine ID", "Stock Before Refill", "Refill Quantity", "Stock After Refill", "Refill Count",   'Location']
    for col, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=5, column=col, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    for row_num, data in enumerate(machine_data, start=6):
        for col_num, value in enumerate(data.values(), start=1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), bottom=Side(style='thin'))

    # Add grand total row
    grand_total_row = len(machine_data) + 6  # Add 6 for headers and blank row
    for col_num, value in enumerate(grand_total.values(), start=4):
        cell = worksheet.cell(row=grand_total_row, column=col_num, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), bottom=Side(style='medium'))

    # Add label for grand total row
    worksheet.merge_cells(start_row=grand_total_row, start_column=3, end_row=grand_total_row, end_column=3)
    label_cell = worksheet.cell(row=grand_total_row, column=3, value="Grand Total")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal='right', vertical='center')
    label_cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), bottom=Side(style='medium'))

    # for col in worksheet.columns:
    #     max_length = 0
    #     column = col[0].column
    #     column_letter = chr(65 + column - 1)
    #     for cell in col:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2) * 1.2
    #     worksheet.column_dimensions[column_letter].width = adjusted_width
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        column_letter = chr(64 + column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Set a fixed width for the "Sr No" column
    worksheet.column_dimensions['A'].width = 10

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

    workbook.save(response)

    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customer_refill_machine_report_download(request):
    try:
        
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        assigned_machines = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')

        serialized_data = [{'org_name': item.assigned_customer.organization,'location':item.machine.installation_location} for item in assigned_machines if item.assigned_user and item.assigned_user.organization]
        
        machine_data = []

        for machine in assigned_machines:
            queryset = MStatus.objects.filter(
                m_id=machine.machine.machine_id,
                created_at__date__range=[date_from, date_to],
                stock=F('capacity')
            ).values('m_id', 'created_at__date', 'stock').annotate(
                count_capacity_equals_stock=Count('id')
            )

            for entry in queryset:
                stock_after_refill = entry['stock']
                stock_before_date = MStatus.objects.filter(
                    m_id=machine.machine.machine_id,
                    created_at__date__lt=entry['created_at__date']
                ).order_by('-created_at').values_list('stock', flat=True).first()

                refill_quantity = None
                if stock_after_refill is not None and stock_before_date is not None:
                    refill_quantity = stock_after_refill - stock_before_date

                machine_data.append({
                    'm_id': entry['m_id'],
                    'date': entry['created_at__date'],
                    'stock_before_refill': stock_before_date,
                    'refill_quantity': refill_quantity,
                    'stock_after_refill': entry['stock'],
                    'stock_capacity_equal_count': entry['count_capacity_equals_stock'],
                    
                    'location':machine.machine.installation_location
                })

        grand_total = {
            "stock_before_refill": sum(data.get("stock_before_refill", 0) or 0 for data in machine_data),
            "refill_quantity": sum(data.get("refill_quantity", 0) or 0 for data in machine_data),
            "stock_after_refill": sum(data.get("stock_after_refill", 0) or 0 for data in machine_data),
            "stock_capacity_equal_count": sum(data.get("stock_capacity_equal_count", 0) or 0 for data in machine_data),
        }

        workbook = Workbook()
        worksheet = workbook.active

        organization_name = serialized_data[0].get('org_name') if serialized_data else None
        # if organization_name:
        #     worksheet.merge_cells('C1:D2')
        #     cell = worksheet['C1']
        #     cell.value = organization_name
        #     cell.font = Font(bold=True)
        #     cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # worksheet.merge_cells('C3:D3')
        # cell = worksheet['C3']
        # cell.value = "Refill Machine Report"
        # cell.font = Font(bold=True)
        # cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # worksheet.merge_cells('E1:E3')
        # date_range_cell = worksheet['E1']
        # date_range_cell.value = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
        # date_range_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if organization_name:
            worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
            cell = worksheet['A4']  # Get the top-left cell of the merged range
            cell.value = f"Organization Name: {organization_name}"
            cell.font = Font(bold=True,size=12)
            cell.alignment = Alignment(wrap_text=True)  # Align text to the center

        # Add the "Customer Machine Report" label in the middle of the third row
        worksheet.merge_cells('A1:F2')  # Merge cells for the label
        cell = worksheet['A1']  # Get the top-left cell of the merged range
        cell.value = "Refill Machine Report"
        cell.font = Font(bold=True,size=16)  # Make the label bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        worksheet.merge_cells('A3:F3')  # Merge cells for the date range
        date_cell = worksheet['A3']  # Get the top-left cell of the merged range
        date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
       
        # Set up border styles
        border_style = Side(border_style="thin")
        border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

        # Add headers for machine data
        headers = ["Sr No", "Machine ID", "Before Refill Stock", "Refill Quantity", "After Refill Stock", "Refill Count", 'Location', "Refill Date"]
        for col, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=5, column=col, value=header)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Adding border to header cells
            header_cell.border = border

        # Write machine data to the worksheet if available
        for row_num, data in enumerate(machine_data, start=6):
            sr_no_cell = worksheet.cell(row=row_num, column=1, value=row_num - 5)
            sr_no_cell.alignment = Alignment(horizontal='center', vertical='center')
            sr_no_cell.border = border
            
            m_id_cell = worksheet.cell(row=row_num, column=2, value=data['m_id'])
            m_id_cell.alignment = Alignment(horizontal='center', vertical='center')
            m_id_cell.border = border

            stock_before_refill_cell = worksheet.cell(row=row_num, column=3, value=data['stock_before_refill'])
            stock_before_refill_cell.alignment = Alignment(horizontal='center', vertical='center')
            stock_before_refill_cell.border = border

            refill_quantity_cell = worksheet.cell(row=row_num, column=4, value=data['refill_quantity'])
            refill_quantity_cell.alignment = Alignment(horizontal='center', vertical='center')
            refill_quantity_cell.border = border

            stock_after_refill_cell = worksheet.cell(row=row_num, column=5, value=data['stock_after_refill'])
            stock_after_refill_cell.alignment = Alignment(horizontal='center', vertical='center')
            stock_after_refill_cell.border = border

            refill_count_cell = worksheet.cell(row=row_num, column=6, value=data['stock_capacity_equal_count'])
            refill_count_cell.alignment = Alignment(horizontal='center', vertical='center')
            refill_count_cell.border = border

            location_cell = worksheet.cell(row=row_num, column=7, value=data['location'])
            location_cell.alignment = Alignment(horizontal='center', vertical='center')
            location_cell.border = border

            refill_date_cell = worksheet.cell(row=row_num, column=8, value=data['date'])
            refill_date_cell.alignment = Alignment(horizontal='center', vertical='center')
            refill_date_cell.border = border

        # Add grand total row
        grand_total_row = len(machine_data) + 6
        for col_num, value in enumerate(grand_total.values(), start=3):
            grand_total_cell = worksheet.cell(row=grand_total_row, column=col_num, value=value)
            grand_total_cell.alignment = Alignment(horizontal='center', vertical='center')
            grand_total_cell.border = border

        # Add label for grand total row
        worksheet.merge_cells(start_row=grand_total_row, start_column=2, end_row=grand_total_row, end_column=2)
        label_cell = worksheet.cell(row=grand_total_row, column=2, value="Grand Total")
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')
        label_cell.border = border

        # Adjust column widths based on content length
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            column_letter = chr(64 + column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set a fixed width for the "Sr No" column
        worksheet.column_dimensions['A'].width = 10

        # Set the response headers for file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # Write the Excel workbook to the response
        workbook.save(response)

        return response

    except Exception as e:
        # Handle exceptions, log the error, and return an appropriate error response
        return Response({'success': 0, 'message': f'Error processing data: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


import pandas as pd
from rest_framework.decorators import action
import logging
from rest_framework import status, viewsets

import qrcode
from io import BytesIO
from django.conf import settings

from .models import MachineMaster, QrCode, Product, ModelCapacity
from .serializers import MachineMasterSerializer1
from django.core.files.base import ContentFile


# class BulkUploadMachineMasterViewSet(viewsets.ModelViewSet):
#     queryset = MachineMaster.objects.all()
#     serializer_class = MachineMasterSerializer1

#     @action(detail=False, methods=['post'], url_path='bulk-upload')
#     def bulk_upload(self, request):
#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             df = pd.read_excel(file)
#             records = df.to_dict(orient='records')
            
#             qr_code_map = {}
#             product_map = {}
#             model_map = {}

#             def get_or_create_qr_code(record):
#                 key = (record['qrcodeid'], record['qrstoreid'], record['qrstorename'], record['upilink'])
#                 if key not in qr_code_map:
#                     upi_link = record['upilink']
#                     qr_img = qrcode.make(upi_link)

#                     buffer = BytesIO()
#                     qr_img.save(buffer, format='PNG')
#                     buffer.seek(0)

#                     qr_code_obj, _ = QrCode.objects.get_or_create(
#                         qr_code_id=record['qrcodeid'],
#                         qr_store_id=record['qrstoreid'],
#                         qr_store_name=record['qrstorename']
#                     )
#                     qr_code_obj.qr.save(f'{record["machineid"]}.png', ContentFile(buffer.getvalue()))
#                     qr_code_obj.save()

#                     qr_code_map[key] = qr_code_obj.id
#                 return qr_code_map[key]


#             def get_or_create_product(record):
#                 key = (record['producttype'], record['amount'])
#                 if key not in product_map:
#                     products = Product.objects.filter(product_type=record['producttype'], amount=record['amount'])
#                     if products.exists():
#                         if products.count() > 1:
#                             # Handle the case where multiple records are found
#                             product_obj = products.first()  # You can decide how to handle this, e.g., take the first one or raise an error
#                         else:
#                             product_obj = products.first()
#                     else:
#                         product_obj = Product.objects.create(product_type=record['producttype'], amount=record['amount'])
#                     product_map[key] = product_obj.id
#                 return product_map[key]


#             def get_or_create_model(record):
#                 key = (record['modelno'], record['name'])
#                 if key not in model_map:
#                     models = ModelCapacity.objects.filter(model_no=record['modelno'], name=record['name'])
#                     if models.exists():
#                         if models.count() > 1:
#                             model_obj = models.first()  # Choose the first one or handle as needed
#                         else:
#                             model_obj = models.first()
#                     else:
#                         model_obj = ModelCapacity.objects.create(model_no=record['modelno'], name=record['name'])
#                     model_map[key] = model_obj.id
#                 return model_map[key]


#             processed_records = []
#             for record in records:
#                 record['rqcode'] = get_or_create_qr_code(record)
#                 record['product'] = get_or_create_product(record)
#                 record['modelnumber'] = get_or_create_model(record)
#                 processed_records.append({
#                     'rqcode': record['rqcode'],
#                     'product': record['product'],
#                     'model_number': record['modelnumber'],
#                     'machine_id': record['machineid'],
#                     'installation_location': record.get('installationlocation'),
#                     'payment_type': record.get('paymenttype'),
#                     'machinelease': record.get('machinelease')
#                 })

#             # Serialize and validate the data
#             serializer = MachineMasterSerializer1(data=processed_records, many=True)
#             if serializer.is_valid():
#                 serializer.save()
#                 return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data}, status=status.HTTP_201_CREATED)
#             else:
#                 return Response({'success': 0, 'message': 'Not Found','result':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
#         except Exception as e:
#             return Response({"error": f"Error processing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

class BulkUploadMachineMasterViewSet(viewsets.ModelViewSet):
    queryset = MachineMaster.objects.all()
    serializer_class = MachineMasterSerializer1
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        logger.info("Starting bulk upload process")
        file = request.FILES.get('file')
        if not file:
            logger.error("No file provided")
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Define expected header names
        expected_headers = [
            'machineid', 'qrcodeid', 'qrstoreid', 'qrstorename', 'paymenttype', 'installationlocation',
            'producttype', 'amount', 'modelno', 'modelcapacity', 'machinelease', 'upilink'
        ]

        try:
            df = pd.read_excel(file)
            logger.info("File successfully read into DataFrame")

            # Check if headers match expected headers
            if not all(header in df.columns for header in expected_headers):
                logger.error("Invalid file format. Headers do not match expected headers check sample excel file.")
                return Response({'success': 0, "message": "Invalid file format. Headers do not match expected headers please check sample excel file."}, status=status.HTTP_400_BAD_REQUEST)

            # Check for empty cells in the required columns
            if df[expected_headers].isnull().values.any():
                logger.error("Empty cells found in required columns")
                return Response({'success': 0, "message": "Empty cells found in required columns."}, status=status.HTTP_400_BAD_REQUEST)


            records = df.to_dict(orient='records')
            logger.info(f"Parsed {len(records)} records from file")

            qr_code_map = {}
            product_map = {}
            model_map = {}

            def get_or_create_qr_code(record):
                key = (record['qrcodeid'], record['qrstoreid'], record['qrstorename'], record['upilink'])
                if key not in qr_code_map:
                    upi_link = record['upilink']
                    qr_img = qrcode.make(upi_link)
                    buffer = BytesIO()
                    qr_img.save(buffer, format='PNG')
                    buffer.seek(0)

                    qr_code_obj, _ = QrCode.objects.get_or_create(
                        qr_code_id=record['qrcodeid'],
                        qr_store_id=record['qrstoreid'],
                        qr_store_name=record['qrstorename']
                    )
                    qr_code_obj.qr.save(f'{record["machineid"]}.png', ContentFile(buffer.getvalue()))
                    qr_code_obj.save()

                    qr_code_map[key] = qr_code_obj.id
                return qr_code_map[key]

            def get_or_create_product(record):
                key = (record['producttype'], record['amount'])
                if key not in product_map:
                    products = Product.objects.filter(product_type=record['producttype'], amount=record['amount'])
                    if products.exists():
                        product_obj = products.first()
                    else:
                        product_obj = Product.objects.create(product_type=record['producttype'], amount=record['amount'])
                    product_map[key] = product_obj.id
                return product_map[key]

            def get_or_create_model(record):
                key = (record['modelno'], record['modelcapacity'])
                if key not in model_map:
                    models = ModelCapacity.objects.filter(model_no=record['modelno'], name=record['modelcapacity'])
                    if models.exists():
                        model_obj = models.first()
                    else:
                        model_obj = ModelCapacity.objects.create(model_no=record['modelno'], name=record['modelcapacity'])
                    model_map[key] = model_obj.id
                return model_map[key]

            processed_records = []
            for record in records:
                record['rqcode'] = get_or_create_qr_code(record)
                record['product'] = get_or_create_product(record)
                record['modelnumber'] = get_or_create_model(record)
                processed_records.append({
                    'rqcode': record['rqcode'],
                    'product': record['product'],
                    'model_number': record['modelnumber'],
                    'machine_id': record['machineid'],
                    'installation_location': record.get('installationlocation'),
                    'payment_type': record.get('paymenttype'),
                    'machinelease': record.get('machinelease')
                })

            logger.info("Records processed and mapped successfully")

            # Serialize and validate the data
            serializer = MachineMasterSerializer1(data=processed_records, many=True)
            if serializer.is_valid():
                serializer.save()
                logger.info("Data saved successfully")
                return Response({'success': 1, 'message': 'Data Found', 'result': serializer.data}, status=status.HTTP_201_CREATED)
            else:
                logger.error(f"Validation errors: {serializer.errors}")
                return Response({'success': 0, 'message': 'Not Found', 'result': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        except pd.errors.EmptyDataError:
            logger.error("Empty file or no data found in file")
            return Response({"error": "Empty file or no data found in file"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            return Response({"error": f"Error processing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

from django.core.exceptions import ObjectDoesNotExist

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_ticket(request):
    if request.method == 'POST':
        machine_name = request.data.get('machine_map')
        title = request.data.get('title')
        description = request.data.get('description')
        notes = request.data.get('notes')
        priority = request.data.get('priority', 'MEDIUM')
        problem_type = request.data.get('problem_type')

        errors = {}

        # Validate and get related objects
        machine_mapping = None
        machine_master = None
        if machine_name:
            try:
                machine_master = MachineMaster.objects.get(machine_id=machine_name)  
                machine_mapping = MachineUserMapping.objects.get(machine=machine_master)
            except MachineMaster.DoesNotExist:
                errors['machine_name'] = 'Machine name does not exist'
            except MachineUserMapping.DoesNotExist:
                errors['machine_mapping'] = 'No mapping found for the given machine name'

        # Check for errors before creating the ticket
        if errors:
            return Response({'success': 0, 'message': 'Not Created Data', 'result': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Create the Ticket object
        ticket = Ticket(
            user=request.user,  # Assign the authenticated user
            machine_mapping=machine_mapping,
            machine_map=machine_master,
            title=title,
            description=description,
            notes=notes,
            priority=priority,
            problem_type=problem_type
        )

        # Generate and assign ticket number
        ticket.ticsrno = generate_ticsrno()

        # Save the ticket
        ticket.save()

        # Create the response data
        response_data = {
            'id': ticket.id,
            'user': ticket.user.id if ticket.user else None,
            'machine_mapping': ticket.machine_mapping.id if ticket.machine_mapping else None,
            'machine_master': machine_master.id if machine_master else None,
            'title': ticket.title,
            'description': ticket.description,
            'notes': ticket.notes,
            'priority': ticket.priority,
            'problem_type': ticket.problem_type,
        }

        return Response({'success': 1, 'message': 'Ticket Created Successfully', 'result': response_data}, status=status.HTTP_201_CREATED)
    
    return Response({'success': 0, 'message': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)

def generate_ticsrno():
    return f"TICK-{now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:2].upper()}"



# class TicketViewSet(viewsets.ModelViewSet):
#     queryset = Ticket.objects.all()
#     serializer_class = TicketSerializer
#     permission_classes=[IsAuthenticated]

#     def create(self, request, *args, **kwargs):
#         try:
#             # machine_name = request.data.get('machine_map')  # Get machine name from request data
#             # machine = MachineMaster.objects.get(machine_id=machine_name)  # Assuming machine_id is the field to filter by
#             # print(machine,'mmm')
#             # print(machine.pk,'ppppp')
#             serializer = TicketSerializer(data=request.data)
#             serializer.is_valid(raise_exception=True)
#             # print(machine.pk,'llll')
#             # # Assign machine_map field with the primary key value
#             # serializer.validated_data['machine_map'] = machine.pk

#             self.perform_create(serializer)

#             return Response({'success': 1, 'message': 'Data Created Successfully', 'result': serializer.data}, status=status.HTTP_201_CREATED)

#         except MachineMaster.DoesNotExist:
#             return Response({'success': 0, 'message': 'Machine not found'}, status=status.HTTP_404_NOT_FOUND)

#         except Exception as e:
#             logging.error(f'Error creating Ticket: {str(e)}')
#             return Response({'success': 0, 'message': 'Not Created Data', 'result': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

#     def perform_create(self, serializer):
#         ticket = serializer.save(user_id=self.request.user.id)
#         ticket.ticsrno = self.generate_ticsrno()
#         ticket.save()
#         # Print the generated ticsrno to the terminal
#         print(f"Generated ticsrno: {ticket.ticsrno}")

#     def generate_ticsrno(self):
#         return f"TICK-{now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
    

    
import uuid
from django.utils.timezone import now



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ticket_list(request):
    try:
        user_profile = User.objects.get(id=request.user.id)
        print(user_profile,'user pro')
        if request.method == 'GET':
            if user_profile.role == '1':  # Superuser
                tickets = Ticket.objects.all()
            elif user_profile.role == '2':  # Admin
                # Admin sees tickets created by their own users
                own_users = User.objects.filter(created_by=user_profile.id)
                print(own_users,'own1')
                tickets = Ticket.objects.filter(Q(user__in=own_users) | Q(created_by__in=own_users))
            elif user_profile.role == '3':  # Customer
                # Customer sees tickets created by their own users and themselves
                own_users = User.objects.filter(created_by=user_profile.id)
                print(own_users,'own')
                tickets = Ticket.objects.filter(Q(user__in=own_users) | Q(created_by__in=own_users) | Q(created_by=user_profile))
            else:  # Other roles
                tickets = Ticket.objects.none()
            data=[
                {
                    'id': tic.id,
                    'machine_map':tic.machine_map.machine_id,
                    'user': tic.user.id if tic.user else None,
                    'email': tic.user.email if tic.user else None,
                    'name': tic.user.name if tic.user else None,
                    'title': tic.title,
                    'description': tic.description,
                    'notes': tic.notes,
                    'priority': tic.priority,
                    'problem_type': tic.problem_type,
                    'ticsrno':tic.ticsrno,
                    'created_at':tic.created_at,
                    'status':tic.status
                }for tic in tickets
            ]

            # serializer = TicketSerializer(tickets, many=True)
            return Response({'success':1,'message':'Data Found','result':data})
        
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)
    
# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def update_ticket(request, pk):
#     try:
#         user_profile = User.objects.get(id=request.user.id)
#         try:
#             ticket = Ticket.objects.get(pk=pk)
#         except Ticket.DoesNotExist:
#             return Response({'error': 'Ticket not found'}, status=404)

#         if request.method == 'PUT':
#             # Allow update if the user is Superuser or Admin, or if the user created the ticket
#             if user_profile.role in ['1', '2'] or ticket.user == request.user:
#                 serializer = TicketSerializer(ticket, data=request.data, partial=True)
#                 if serializer.is_valid():
#                     serializer.save()
#                     return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})
#                 return Response({'error': 'Invalid data', 'result': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
#             else:
#                 return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

#     except User.DoesNotExist:
#         return Response({'error': 'User profile not found'}, status=404)
#     except Exception as e:
#         return Response({'error': str(e)}, status=500)
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_ticket(request, pk):
    try:
        user_profile = request.user
        ticket = get_object_or_404(Ticket, pk=pk)
        
        if user_profile.role == '1':  # Superuser can update any ticket
            pass
        elif user_profile.role == '2':  # Admin can update tickets created by their own users
            own_users = User.objects.filter(created_by=user_profile.id)
            if ticket.user not in own_users and ticket.created_by not in own_users:
                return Response({'error': 'You do not have permission to update this ticket.'}, status=status.HTTP_403_FORBIDDEN)
        elif user_profile.role == '3':  # Customer can update their own tickets and those created by their own users
            own_users = User.objects.filter(created_by=user_profile.id)
            if ticket.user not in own_users and ticket.created_by not in own_users and ticket.created_by != user_profile:
                return Response({'error': 'You do not have permission to update this ticket.'}, status=status.HTTP_403_FORBIDDEN)
        else:  # Other roles cannot update tickets
            return Response({'error': 'You do not have permission to update this ticket.'}, status=status.HTTP_403_FORBIDDEN)
        
        # Update fields manually
        ticket_fields = ['title', 'description', 'notes', 'priority', 'status', 'assigned_to']
        for field in ticket_fields:
            if field in request.data:
                setattr(ticket, field, request.data[field])
        
        ticket.save()
        
        data = {
            'id': ticket.id,
            'machine_map': ticket.machine_map.machine_id if ticket.machine_map else None,
            'user': ticket.user.id if ticket.user else None,
            'title': ticket.title,
            'description': ticket.description,
            'notes': ticket.notes,
            'priority': ticket.priority,
            'problem_type': ticket.problem_type,
            'ticsrno': ticket.ticsrno,
            'created_at': ticket.created_at,
            'status': ticket.status,
            'created_by': ticket.created_by.id if ticket.created_by else None
        }
        
        return Response({'success': 1, 'message': 'Ticket updated successfully', 'result': data}, status=status.HTTP_200_OK)
    
    except Ticket.DoesNotExist:
        return Response({'error': 'Ticket not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_machine_mapped(request):
    if request.user.is_authenticated:
        try:
            # Check if the authenticated user is an assigned customer
            machine_user_mappings = MachineUserMapping.objects.filter(assigned_customer=request.user)
            
            # If the user is not an assigned customer, check if they are an assigned user
            if not machine_user_mappings.exists():
                machine_user_mappings = MachineUserMapping.objects.filter(assigned_user=request.user)
            
            if machine_user_mappings.exists():
                mapped_machines_data = []
                for mapping in machine_user_mappings:
                    # Access machine details through the mapping object
                    machine_data = {
                        'id':mapping.machine.id if mapping.machine else None,
                        'machine_id': mapping.machine.machine_id if mapping.machine else None,  
                        # 'mapping_details': MachineUserMappingSerializer(mapping).data
                    }
                    mapped_machines_data.append(machine_data)
                
                return Response({'success': 1, 'message': 'Mapped machines retrieved successfully', 'result': mapped_machines_data}, status=status.HTTP_200_OK)
            else:
                return Response({"message": "User does not have any mapped machines"}, status=status.HTTP_404_NOT_FOUND)
        except MachineUserMapping.DoesNotExist:
            return Response({"message": "User does not have any mapped machines"}, status=status.HTTP_404_NOT_FOUND)
    else:
        return Response({"detail": "Authentication credentials were not provided"}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_ticket_list_customer(request):
    try:
        user_profile = User.objects.get(id=request.user.id)
        if request.method == 'GET':
            tickets = Ticket.objects.filter(user=user_profile)
            
            data = [
                {
                    'machine_id': tic.machine_map.machine_id if tic.machine_map else None,
                    "title": tic.title,
                    "description": tic.description,
                    "notes": tic.notes,
                    "created_at": tic.created_at,
                    "resolved_at": tic.resolved_at,
                    "priority": tic.priority,
                    "status": tic.status,
                    "problem_type": tic.problem_type,
                    "ticsrno": tic.ticsrno,
                    "user": tic.user.name,
                }
                for tic in tickets
            ]
            return Response({'success': 1, 'message': 'Data Found', 'result': data})

    except User.DoesNotExist:
        return Response({'success': 0, 'message': 'User not found'}, status=404)
    except Exception as e:
        logging.error(f'Error retrieving tickets: {str(e)}')
        return Response({'success': 0, 'message': 'Internal Server Error', 'error': str(e)}, status=500)

import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from django.shortcuts import get_object_or_404
from .models import Role, Module, SubModule, RoleModuleAssignment,  ExcelFile
from .serializers import RoleSerializer, ModuleSerializer, SubModuleSerializer, RoleModuleAssignmentSerializer,  ExcelFileUploadSerializer

# class ExcelFileUploadView(APIView):
#     def post(self, request, *args, **kwargs):
#         serializer = ExcelFileUploadSerializer(data=request.data)
#         if serializer.is_valid():
#             excel_file_instance = serializer.save()

#             # Parse the Excel file
#             parse_excel_file(excel_file_instance.file.path)

#             return Response({"message": "File uploaded and data stored successfully"}, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# def parse_excel_file(file_path):
#     # Read the Excel file
#     excel_data = pd.ExcelFile(file_path)

#     # Assuming the first sheet contains Module and SubModule data
#     sheet1_data = pd.read_excel(excel_data, sheet_name='Sheet1')

#     for _, row in sheet1_data.iterrows():
#         module_name = row.get('Module')
#         sub_module_name = row.get('Sub Module')
#         if module_name:
#             module, _ = Module.objects.get_or_create(name=module_name)
#             if sub_module_name:
#                 SubModule.objects.get_or_create(module=module, name=sub_module_name)

#     # Assuming the second sheet contains Role assignments
#     sheet2_data = pd.read_excel(excel_data, sheet_name='Sheet2')

#     for _, row in sheet2_data.iterrows():
#         role_name = row.get('role')
#         module_name = row.get('module')
#         sub_module_name = row.get('sub module')

#         if role_name:
#             role, _ = Role.objects.get_or_create(name=role_name)
#             if module_name:
#                 module, created = Module.objects.get_or_create(name=module_name)
#                 RoleModuleAssignment.objects.get_or_create(role=role, module=module)
#             if sub_module_name:
#                 sub_module, created = SubModule.objects.get_or_create(name=sub_module_name, module=module)
#                 RoleSubModuleAssignment.objects.get_or_create(role=role, sub_module=sub_module)

class RoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer

    def list(self, request, *args, **kwargs):
        try:
            queryset = Role.objects.all().order_by('-id')
            serializer = RoleSerializer(queryset, many=True)
            return Response({'success': 1, 'message': 'Manage Membership List', 'result': serializer.data})
        except Exception as e:
            return Response({'success': 0, 'message': 'Not Found', 'result': str(e)})

    def retrieve(self, request, pk, *args, **kwargs):
        try:
            chp = Role.objects.get(pk=pk)
            serializer = RoleSerializer(chp)
            return Response({'success': 1, 'message': 'Manage Membership', 'result': serializer.data})
        except Role.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def create(self, request, *args, **kwargs):
        try:
            serializer = RoleSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            
            return Response({'success': 1, 'message': 'Data Created', 'result': serializer.data})
        except ValidationError as ve:
            return Response({'success': 0, 'message': 'Not Created', 'result': ve.detail})

    def update(self, request, pk, *args, **kwargs):
        try:
            chp = Role.objects.get(pk=pk)
            if 'rqcode' not in request.data:
                return Response({'success': 0, 'message': 'rqcode is required'}, status=status.HTTP_400_BAD_REQUEST)
            serializer = RoleSerializer(chp, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})
        except Role.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

    def destroy(self, request, pk, *args, **kwargs):
        try:
            machine = Role.objects.get(pk=pk)

            machine.delete()
            # chp = Role.objects.get(pk=pk)
            # chp.delete()
            return Response({'success': 1, 'message': 'Data Deleted'})
        except Role.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found'})

class ModuleViewSet(viewsets.ModelViewSet):
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer

class SubModuleViewSet(viewsets.ModelViewSet):
    queryset = SubModule.objects.all()
    serializer_class = SubModuleSerializer

class RoleModuleAssignmentViewSet(viewsets.ModelViewSet):
    queryset = RoleModuleAssignment.objects.all()
    serializer_class = RoleModuleAssignmentSerializer

# class RoleSubModuleAssignmentViewSet(viewsets.ModelViewSet):
#     queryset = RoleSubModuleAssignment.objects.all()
#     serializer_class = RoleSubModuleAssignmentSerializer

@api_view(['GET'])
def get_modules_for_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    modules = Module.objects.filter(role_assignments__role=role)
    submodules = SubModule.objects.filter(role_assignments__role=role)
    
    module_serializer = ModuleSerializer(modules, many=True)
    submodule_serializer = SubModuleSerializer(submodules, many=True)
    
    return Response({
        'modules': module_serializer.data,
        'submodules': submodule_serializer.data
    })
from django.contrib.auth.models import Permission

from django.contrib.auth.models import Group
from userapp.models import User
from django.contrib.contenttypes.models import ContentType
# from django.contrib.auth.models import User, Group

# class RoleWithModulesOrSubModulesAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     def post(self, request):
#         user_profile = request.user
#         if user_profile.role != '2':  # Adjust as per your role checking logic
#             return Response({'error': 'You do not have permission to assign permissions.'}, status=status.HTTP_403_FORBIDDEN)

#         data = request.data
#         role_name = data.get('role_name')
#         modules = data.get('modules')

#         if not role_name:
#             return Response({'error': 'Role name is required'}, status=status.HTTP_400_BAD_REQUEST)

#         role, created = Role.objects.get_or_create(name=role_name)

#         for module_data in modules:
#             module_name = module_data.get('name')
#             submodules_data = module_data.get('submodules')

#             module, created = Module.objects.get_or_create(name=module_name, role=role)

#             for submodule_data in submodules_data:
#                 submodule_name = submodule_data.get('name')
#                 permissions_data = submodule_data.get('permissions')

#                 submodule, created = SubModule.objects.get_or_create(name=submodule_name, module=module)

#                 for permission_data in permissions_data:
#                     permission_name = permission_data.get('name')
#                     codename = permission_data.get('codename')

#                     custom_permission, created = CustomPermission.objects.get_or_create(
#                         name=permission_name,
#                         codename=codename,
#                         submodule=submodule
#                     )

#                     content_type = ContentType.objects.get_for_model(CustomPermission)
#                     permission, created = Permission.objects.get_or_create(
#                         codename=codename,
#                         name=permission_name,
#                         content_type=content_type
#                     )

#                     # Add the built-in permission to the role
#                     role.permissions.add(permission)

#         return Response({'message': 'Role, modules, and sub-modules created successfully'}, status=status.HTTP_200_OK)

class RoleWithModulesOrSubModulesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user_profile = request.user
        if user_profile.role != '2':  # Adjust as per your role checking logic
            return Response({'error': 'You do not have permission to assign permissions.'}, status=status.HTTP_403_FORBIDDEN)

        data = request.data
        role_name = data.get('role_name')
        modules = data.get('modules')

        if not role_name:
            return Response({'error': 'Role name is required'}, status=status.HTTP_400_BAD_REQUEST)

        role, created = Role.objects.get_or_create(name=role_name)

        for module_data in modules:
            module_name = module_data.get('name')
            module, created = Module.objects.get_or_create(name=module_name, role=role)

            # No need to handle submodules and permissions here

        return Response({'message': 'Role and modules created successfully','success':1}, status=status.HTTP_200_OK)

class AssignRoleToUserAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        user_id = data.get('user_id')
        role_name = data.get('role_name')

        user = User.objects.get(id=user_id)
        role = Role.objects.get(name=role_name)

        role.users.add(user)

        # Get the built-in Permission instances associated with the role's custom permissions
        permissions = role.permissions.all()
        user.user_permissions.set(permissions)

        return Response({'message': 'Role and permissions assigned to user successfully'}, status=status.HTTP_200_OK)


# class UserRolesModulesAndSubModulesAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         user = request.user
#         roles = user.roles.all()

#         if not roles:
#             return Response({'error': 'User has no roles assigned'}, status=status.HTTP_404_NOT_FOUND)

#         data = []

#         for role in roles:
#             role_data = {
#                 'role_name': role.name,
#                 'modules': []
#             }
#             modules = Module.objects.filter(role=role)
#             for module in modules:
#                 module_data = {
#                     'module_name': module.name,
#                     'submodules': []
#                 }
#                 submodules = SubModule.objects.filter(module=module)
#                 for submodule in submodules:
#                     permissions = CustomPermission.objects.filter(submodule=submodule)
#                     permission_data = [{'name': permission.name, 'codename': permission.codename} for permission in permissions]
#                     module_data['submodules'].append({
#                         'name': submodule.name,
#                         'permissions': permission_data
#                     })
#                 role_data['modules'].append(module_data)
#             data.append(role_data)

#         return Response(data, status=status.HTTP_200_OK)
class UserRolesModulesAndSubModulesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        user_id = user.id
        roles = Role.objects.filter(users=user)
        roles_data = []

        for role in roles:
            modules = Module.objects.filter(role=role)
            modules_data = []

            for module in modules:
                submodules = SubModule.objects.filter(module=module)
                submodules_data = []

                for submodule in submodules:
                    permissions = CustomPermission.objects.filter(submodule=submodule)
                    permissions_data = CustomPermissionSerializer(permissions, many=True).data
                    submodules_data.append({
                        'name': submodule.name,
                        'permissions': permissions_data
                    })

                modules_data.append({
                    'name': module.name,
                    'submodules': submodules_data
                })

            roles_data.append({
                'role_name': role.name,
                'modules': modules_data
            })

        response_data = {
            'user_id': user_id,
            'roles': roles_data
        }

        return Response(response_data)