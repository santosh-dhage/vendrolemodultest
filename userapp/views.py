from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
from rest_framework.permissions import AllowAny,IsAuthenticatedOrReadOnly,IsAuthenticated
from rest_framework import viewsets
from userapp.models import MQTTData, MOrder, MStatus, PaymentHistory
from machinedataapp.models import *
from .serializers import MQTTDataSerializer, MOrderSerializer, MStatusSerializer, PaymentHistorySerializer,UserMasterSerializer,ChangePasswordSerializer
from rest_framework.decorators import api_view,permission_classes,authentication_classes
import jwt
from rest_framework import exceptions
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework_simplejwt.tokens import RefreshToken
from .models import User
from .auth import generate_access_token, generate_refresh_token
from rest_framework import exceptions
import paho.mqtt.publish
from django.db.models import Q
from django.core.mail import send_mail

from django.views.generic import TemplateView
from django.views.decorators.cache import never_cache
import logging
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
import logging
from rest_framework import serializers
from django.http import JsonResponse

from django.core.mail import EmailMessage
from django.core.mail.backends.smtp import EmailBackend
from django.utils.crypto import get_random_string
from django.utils.encoding import force_bytes
from django.utils.encoding import force_str as force_text
from tenant.models import Domain
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.tokens import default_token_generator
from .serializers import *
import datetime

BROKER = '65.109.139.161'
PORT = 1883
MQTT_USERNAME = 'lldnoeHIULHU87678'
MQTT_PASSWORD = 'oSBVFIELH8734109283Vkjvbk'

# MQTT topic to which you want to send the response
TOPIC_M_ORDER = 'EMBI/M_Order'
# Logging configuration
LOG_FILE = 'log/app.log'
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
from django.db import transaction


index = never_cache(TemplateView.as_view(template_name='index.html'))
from .permissions import CustomPermissionRequired

class UserMasterViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserMasterSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        try:
            queryset = User.objects.all().order_by('-id')
            serializer = UserMasterSerializer(queryset, many=True)
            return Response({'success': 1, 'message': 'User master List', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error listing UserMaster: {str(e)}')
            return Response({'success': 0, 'message': 'Not Found', 'result': serializer.errors})
        
    def retrieve(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            serializer = UserMasterSerializer(user)
            return Response({'success': 1, 'message': 'User master', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error retrieving UserMaster: {str(e)}')
            return Response({'success': 0, 'message': 'Not Found'})

    def create(self, request, *args, **kwargs):
        domain = self.get_domain(request)
        tenant = domain.tenant
        start_time=time.time()
        try:
            # Validate request data
            serializer = UserMasterSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)

            email = request.data.get('email')
            mobile_no = request.data.get('mobile_no')
            password = request.data['password']
            print(password)

            # Check if email or mobile_no already exist
            if User.objects.filter(email=email).exists() or User.objects.filter(mobile_no=mobile_no).exists():
                elapsed_time = time.time() - start_time
                logging.info(f"Time taken for create_user: {elapsed_time} seconds")
                return Response({'success': 0, 'message': 'Email or mobile_no already exists.'})

            # Save user
            password_hased = make_password(request.data['password'])
            user = serializer.save(password=password_hased, created_by=request.user.id)

            # Welcome email
            subject = 'Welcome to iVendMSoft!'
            message = f"Dear {serializer.data['name']},\n\nThank you for registering in Vending Machine.\n\n Your Gmail Id is {request.data['email']}\n\nYour password is: {password}\n\nYour Domain name is: {tenant.schema_name}\n\nBest regards,\nThe Vending Machine Team"

            
            self.send_custom_email(subject, message, [email])

            # Log successful creation
            logging.info(f'Data Created: {serializer.data}')
            logging.info(f'Welcome email sent to {email}')
            elapsed_time = time.time() - start_time
            logging.info(f"Time taken for create_user: {elapsed_time} seconds")
            logging.info(f'Data Created: {serializer.data}')
            logging.info(f'Welcome email sent to {email}')

            return Response({'success': 1, 'message': 'Data Created Successfully', 'result': serializer.data}, status=status.HTTP_201_CREATED)

        
        except Exception as e:
            logging.error(f'Error creating UserMaster: {str(e)}')
            return Response({'success': 0, 'message': 'Not Created Data', 'result': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    

    def get_domain(self, request):
        start_time=time.time()
        domain_name = request.get_host().split(':')[0]  # Get the domain name from the request
        print(domain_name)

        elapsed_time = time.time() - start_time
        logging.info(f"Time taken for get_domain: {elapsed_time} seconds")
        return Domain.objects.get(domain=domain_name)
    
    def send_custom_email(self,subject, message, recipient_list):
        start_time=time.time()
        try:
            # Check if CustomEmail settings are configured for the tenant
            config = CustomEmail.objects.first()  # Assuming there's only one configuration
            print(config,'config')
            if config:
                # Use CustomEmail settings
                from_email = config.email_host_user
                backend = EmailBackend(
                    host=config.email_host,
                    port=config.email_port,
                    username=config.email_host_user,
                    password=config.email_host_password,
                    use_tls=config.email_use_tls,
                    fail_silently=config.email_fail_silently
                )
            else:
                # Use settings from settings.py
                from_email = settings.EMAIL_HOST_USER
                
                backend = None
            
            # Send email
            send_mail(subject, message, from_email, recipient_list, connection=backend)
            elapsed_time = time.time() - start_time
            logging.info(f"Time taken for send_custom_email: {elapsed_time} seconds")
            
            return True
        except Exception as e:
            logging.error(f'Error sending custom email: {e}')
            return False
    
    def update(self, request, pk, *args, **kwargs):
        try:
            chp = User.objects.get(pk=pk)
            serializer = UserMasterSerializer(chp, data=request.data, partial=True)  # Set partial=True to allow partial updates
            serializer.is_valid(raise_exception=True)
            
            # Remove 'password' field from validated data
            if 'password' in serializer.validated_data:
                del serializer.validated_data['password']
            
            updated_mobile = serializer.validated_data.get('mobile_no')
            if updated_mobile and User.objects.filter(~Q(pk=pk), mobile_no=updated_mobile).exists():
                logging.error(f'Mobile number already exists')
                return Response({'success': 0, 'message': 'Mobile number already exists'})

            serializer.save()
            logging.info(f'Data Updated Successfully: {serializer.data}')
            return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error listing UserMaster: {str(e)}')
            return Response({'success': 0, 'message': 'Not Found', 'result': serializer.errors})
        
    def destroy(self, request, pk, *args, **kwargs):
        try:
            user = User.objects.get(pk=pk)
            user.delete()
            return Response({'success': 1, 'message': 'Data Deleted'})
        except Exception as e:
            logging.error(f'Error deleting UserMaster: {str(e)}')
            return Response({'success': 0, 'message': 'Not Found'}, status=status.HTTP_400_BAD_REQUEST)


from django.shortcuts import get_object_or_404
#2
# class UserMasterViewSet(viewsets.ModelViewSet):
#     queryset = User.objects.all()
#     serializer_class = UserMasterSerializer
#     permission_classes = [IsAuthenticated]

#     def list(self, request, *args, **kwargs):
#         queryset = self.queryset.order_by('-id')
#         serializer = self.get_serializer(queryset, many=True)
#         return Response({'success': 1, 'message': 'User master List', 'result': serializer.data})

#     def retrieve(self, request, pk, *args, **kwargs):
#         user = get_object_or_404(User, pk=pk)
#         serializer = self.get_serializer(user)
#         return Response({'success': 1, 'message': 'User master', 'result': serializer.data})

#     def create(self, request, *args, **kwargs):
#         start_time = time.time()
#         domain = self.get_domain(request)
#         tenant = domain.tenant  # Assuming Domain has a ForeignKey to Tenant
#         serializer = self.get_serializer(data=request.data)
#         serializer.is_valid(raise_exception=True)

#         email = request.data.get('email')
#         mobile_no = request.data.get('mobile_no')

#         if User.objects.filter(Q(email=email) | Q(mobile_no=mobile_no)).exists():
#             return Response({'success': 0, 'message': 'Email or mobile_no already exists.'}, status=status.HTTP_400_BAD_REQUEST)

#         user = serializer.save(password=make_password(request.data['password']), created_by=request.user.id)

#         # Send welcome email
#         email_success = self.send_welcome_email(serializer.data, tenant, email, request.data['password'])

#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for create_user: {elapsed_time} seconds")

#         if email_success:
#             return Response({'success': 1, 'message': 'Data Created Successfully and email sent', 'result': serializer.data}, status=status.HTTP_201_CREATED)
#         else:
#             return Response({'success': 1, 'message': 'Data Created Successfully but email not sent', 'result': serializer.data}, status=status.HTTP_201_CREATED)

#     def update(self, request, pk, *args, **kwargs):
#         user = get_object_or_404(User, pk=pk)
#         serializer = self.get_serializer(user, data=request.data, partial=True)
#         serializer.is_valid(raise_exception=True)

#         if 'mobile_no' in serializer.validated_data and User.objects.filter(~Q(pk=pk), mobile_no=serializer.validated_data['mobile_no']).exists():
#             return Response({'success': 0, 'message': 'Mobile number already exists'}, status=status.HTTP_400_BAD_REQUEST)

#         serializer.save()
#         return Response({'success': 1, 'message': 'Data Updated', 'result': serializer.data})

#     def destroy(self, request, pk, *args, **kwargs):
#         user = get_object_or_404(User, pk=pk)
#         user.delete()
#         return Response({'success': 1, 'message': 'Data Deleted'})

#     def get_domain(self, request):
#         domain_name = request.get_host().split(':')[0]
#         return get_object_or_404(Domain, domain=domain_name)

#     def send_welcome_email(self, user_data, tenant, email, password):
#         subject = 'Welcome to iVendMSoft!'
#         # schema_name = tenant.schema_name()
#         message = (
#             f"Dear {user_data['name']},\n\n"
#             f"Thank you for registering in Vending Machine.\n\n"
#             f"Your Domain name is: {tenant.schema_name}\n\n"
#             f"Your User Id is: {email}\n\n"
#             f"Your password is: {password}\n\n"
#             "Best regards,\nThe Vending Machine Team"
#         )
#         return self.send_custom_email(subject, message, [email])

#     def send_custom_email(self, subject, message, recipient_list):
#         try:
#             config = CustomEmail.objects.first()
#             if config:
#                 backend = EmailBackend(
#                     host=config.email_host,
#                     port=config.email_port,
#                     username=config.email_host_user,
#                     password=config.email_host_password,
#                     use_tls=config.email_use_tls,
#                     fail_silently=config.email_fail_silently
#                 )
#                 from_email = config.email_host_user
#             else:
#                 backend = None
#                 from_email = settings.EMAIL_HOST_USER

#             send_mail(subject, message, from_email, recipient_list, connection=backend)
#             logging.info(f"Email sent to {recipient_list}")
#             return True
#         except Exception as e:
#             logging.error(f'Error sending custom email: {e}')
#             return False



import threading
from django.core.mail import send_mail
from django.conf import settings


   
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])  
def get_user_master(request, user_id):
    if request.method == 'GET':
        try:
            # Retrieve user data based on the provided user ID
            user_master = User.objects.get(id=user_id, role='4')
            # serializer = UserMasterSerializer(user_master)

            role_mapping = {'4': 'user'}

            # Retrieve email of the user who created the record
            created_by_user = User.objects.get(id=user_master.created_by)
            created_by_email = created_by_user.email if created_by_user else None

            # Customize the structure of user data as needed
            custom_user_data = {
                'id': user_master.id,  # Use user ID as the identifier
                'name': user_master.name,
                'email': user_master.email,
                'role': role_mapping.get(user_master.role, user_master.role),
                "mobile_no": user_master.mobile_no,
                "address1": user_master.address1,
                "address2": user_master.address2,
                "pincode": user_master.pincode,
                "state": user_master.state,
                "country": user_master.country,
                "landmark": user_master.landmark,
                "created_at": user_master.created_at,
                "status": user_master.status,
                "is_active": user_master.is_active,
                'created_by': created_by_email,
                
            }

            return Response({'success': 1, 'message': 'Data Found', 'result': custom_user_data})
        except User.DoesNotExist:
            return Response({'success': 0, 'message': 'Not Found', 'result': f'User with ID {user_id} not found'})
        except Exception as e:
            logging.error(f'Error processing request: {e}')
            return Response({'success': 0, 'message': 'Error', 'result': f'{e}'})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_master(request, customer_id):
    if request.method == 'GET':
        # try:
            user_master = User.objects.get(id=customer_id, role='3')

            role_mapping = {'3': 'customer'}

            # Retrieve email of the user who created the record
            created_by_user = User.objects.get(id=user_master.created_by)
            # created_by_email = created_by_user.email if created_by_user else None
            if created_by_user:
                created_by_email = created_by_user.email
            else:
                created_by_email = None
            # Customize the structure of user data as needed
            custom_user_data = {
                'id': user_master.id,  # Use user ID as the identifier
                'name': user_master.name,
                'email': user_master.email,
                'role': role_mapping.get(user_master.role, user_master.role) ,
                "mobile_no": user_master.mobile_no,
                "address1": user_master.address1,
                "address2": user_master.address2,
                "pincode": user_master.pincode,
                "state": user_master.state,
                "country": user_master.country,
                "landmark": user_master.landmark,
                "created_at": user_master.created_at,
                "status": user_master.status,
                "is_active": user_master.is_active,
                'created_by': created_by_email,
                
            }

            return Response({'success': 1, 'message': 'Data Found', 'result': custom_user_data})
        # except UserMaster.DoesNotExist:
        #     return Response({'success': 0, 'message': 'Not Found', 'result': f'User with ID {customer_id} not found'})
        # except Exception as e:
        #     logging.error(f'Error processing request: {e}')
        #     return Response({'success': 0, 'message': 'result', 'result': f'{e}'})



@api_view(['POST'])
@permission_classes([AllowAny])
@ensure_csrf_cookie
def login_view(request):
    
    try:
        email = request.data.get('email')
        password = request.data.get('password')
        response = Response()
        if (email is None) or (password is None):
            response.data = {'success': 0,
                             'message': 'email and password required',
                             'result': ""}
            return response
        
        user = User.objects.filter(email=email).first()
        
        if user is None:
            response.data = {'success': 0,
                             'message': 'User Not Found',
                             'result': ""}
            return response
        if not user.check_password(password):
            response.data = {'success': 0,
                             'message': 'wrong password',
                             'result': ""}
            return response
        
        if not user.is_active:
            response.data = {'success': 0,
                            'message': 'Your account is not active So please contact administrator',
                            'result': {}}
            return response
        

        serialized_user = UserMasterSerializer(user).data
        

        access_token = generate_access_token(user)
        refresh_token = generate_refresh_token(user)

        response.set_cookie(key='refreshtoken', value=refresh_token, httponly=True)
        response.data = {
            'success': 1,
            'message': 'Login successfully',
            'result': {
                'access': access_token,
                'refresh': refresh_token,
                'user': serialized_user,
            }
        }

        logging.info(f'User {email} logged in successfully.')

    except Exception as e:
        logging.error(f'Login failed. An error occurred: {str(e)}')
        response.data = {'success': 0, 'message': 'Login failed. An error occurred.', 'result': ""}
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard(request):
    if request.user.is_authenticated:
        serializer = UserMasterSerializer(request.user)
        
        data={
            'name':serializer.data['name'],
            'organization':serializer.data['organization']
        }
        return Response({'success':1,'message':'Data Found','result': data})
    else:
        return Response({'success':0,'message': 'Authentication required'})


@api_view(['POST'])
@permission_classes([AllowAny])
# @csrf_protect
def refresh_token_view(request):
    User = get_user_model()
    refresh_token = request.data.get('refresh')
    if refresh_token is None:
        raise exceptions.AuthenticationFailed(
            'Authentication credentials were not provided.')
    try:
        payload = jwt.decode(
            refresh_token, settings.SECRET_KEY, algorithms=['HS256'])
    except jwt.ExpiredSignatureError:
        raise exceptions.AuthenticationFailed(
            'expired refresh token, please login again.')

    user = User.objects.filter(id=payload.get('user_id')).first()
    if user is None:
        raise exceptions.AuthenticationFailed('User not found')

    if not user.is_active:
        raise exceptions.AuthenticationFailed('user is inactive')

    serialized_user = UserMasterSerializer(user).data

    access_token = generate_access_token(user)
    # refresh_token = generate_refresh_token(user)
    return Response({'access': access_token, 'refresh': refresh_token, 'user': serialized_user})



from rest_framework.generics import UpdateAPIView
from django.contrib.auth.hashers import make_password, check_password

UserModel = get_user_model()

class ChangePasswordView(UpdateAPIView):
    serializer_class = ChangePasswordSerializer
    permission_classes = [IsAuthenticated]

    def validate_request_data(self, data, user):
        # Validate old password if provided
        if 'old_password' not in data:
            raise serializers.ValidationError("Old password is required")

        # Check if the old password matches the user's current password
        if not check_password(data['old_password'], user.password):
            raise serializers.ValidationError("Incorrect old password")

        # Check if the new password and confirm new password match
        if data['new_password'] != data['confirm_new_password']:
            raise serializers.ValidationError("Passwords do not match")

    def update(self, request, *args, **kwargs):
        try:
            # If the user is authenticated, use request.user, otherwise, create a new user
            if request.user.is_authenticated:
                user = request.user
            else:
                # Create a new user with a known identifier like email
                email = request.data.get('email')  # Assuming email is provided in the request
                if email:
                    user, created = get_user_model().objects.get_or_create(email=email)
                else:
                    return Response({'success': 0, 'message': 'Email not provided'}, status=status.HTTP_400_BAD_REQUEST)

            # Validate the request data
            self.validate_request_data(request.data, user)

            # Update user's password only if the user is not an anonymous user
            if not user.is_anonymous:
                user.set_password(request.data['new_password'])
                user.save()

            return Response({'success': 1, 'message': 'Password reset successfully'})
        except serializers.ValidationError as validation_error:
            # Log the validation error
            logger.error(f"Validation error: {str(validation_error)}")
            return Response({'success': 0, 'message': str(validation_error)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            # Log the exception
            logger.error(f"Error updating password: {str(e)}")
            return Response({'success': 0, 'message': 'Invalid Data'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from django.core.mail import EmailMultiAlternatives

User=get_user_model()


@api_view(['POST'])
@permission_classes([AllowAny])
def forgot_password(request):
    serializer = ForgotPasswordSerializer(data=request.data)
    try:
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']

        # Attempt to get the user
        user = User.objects.filter(email=email)
        if len(user) == 0:
            return Response({'success':0,'message': 'User not Found '})
        # Generate and save OTP
        otp = get_random_string(length=4, allowed_chars='0123456789')
        print(otp)
        user[0].otp = otp
        user[0].save()

        # Send OTP via email using EmailMultiAlternatives with inline HTML
        subject = 'Reset Your Password'
        
        message = f"Dear ,\n\nThank you for registration in Vending Machine.\n\n Your Gmail Id is {serializer.data['email']}\n\n Your OTP is: {otp}\n\nBest regards,\nThe Vending Machine Team"
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [serializer.data['email'], ]
        send_mail(subject, message, email_from, recipient_list)

        return Response({'success':1,'message': 'OTP sent successfully'}, status=status.HTTP_200_OK)

    except UserMaster.DoesNotExist:
        return Response({'success':0,'message': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    except serializers.ValidationError as e:
        return Response({'success':0,'message': f'Invalid data: {e.detail}'}, status=status.HTTP_400_BAD_REQUEST)

    # except Exception as e:
    #     return Response({'success':0,'message': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
@permission_classes([AllowAny])
def verify_otp(request):
    serializer = VerifyOtpSerializer(data=request.data)

    try:
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']
        provided_otp = serializer.validated_data['otp']

        # Attempt to get the user
        user = User.objects.get(email=email)

        # Verify OTP
        if user.otp == provided_otp:
            # OTP is valid, you can perform additional actions here if needed
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            logger.info(f'OTP verification successful for {email}')
            return Response({'success':1,'uid': uid, 'token': token, 'message': 'OTP verification successful'}, status=status.HTTP_200_OK)
        else:
            return Response({'success':0,'Message': 'Invalid OTP'})

    except UserMaster.DoesNotExist:
        return Response({'success':0,'Message': 'User not found'})

    except serializers.ValidationError as e:
        return Response({'success':0,'Message': f'Invalid data: {e.detail}'}, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response({'success':0,'Message': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password(request):
    serializer = ResetPasswordSerializer(data=request.data)

    try:
        serializer.is_valid(raise_exception=True)
    except ValidationError as e:
        return Response({'success':0,'message': f'Invalid data: {e.detail}'}, status=status.HTTP_400_BAD_REQUEST)

    uidb64 = serializer.validated_data['uidb64']
    token = serializer.validated_data['token']
    new_password = serializer.validated_data['new_password']

    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
    except (TypeError, ValueError, OverflowError):
        return Response({'success':0,'message': 'Invalid UID'}, status=status.HTTP_400_BAD_REQUEST)

    user = None
    if uid is not None:
        try:
            user = User.objects.get(pk=uid)
        except User.DoesNotExist:
            return Response({'success':0,'message': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        if user and default_token_generator.check_token(user, token):
            if user.check_password(new_password):
                return Response({'success':0,'Message': 'New password cannot be the same as the current password'},status=status.HTTP_400_BAD_REQUEST)
            # Set the new password
            user.set_password(new_password)
            user.save()

            # Send an email notification
            subject = 'Password Reset Successful'
            message = 'Your password has been successfully reset.'
            email_message = EmailMessage(subject, message, to=[user.email])
            email_message.send()

            logger.info(f'Password reset successful for {user.email}')
            return Response({'success': 1,'message': 'Password reset successful'}, status=status.HTTP_200_OK)
        else:
            return Response({'success':0,'message': 'Invalid token or user'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Internal Server Error: {str(e)}')
        return Response({'success':0,'message': 'Internal Server Error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class MQTTDataViewSet(viewsets.ModelViewSet):
    queryset = MQTTData.objects.all().order_by('-id')
    serializer_class = MQTTDataSerializer
    permission_classes = [IsAuthenticated]


class MOrderViewSet(viewsets.ModelViewSet):
    queryset = MOrder.objects.all().order_by('-id')
    serializer_class = MOrderSerializer
    permission_classes = [IsAuthenticated]


class MStatusViewSet(viewsets.ModelViewSet):
    queryset = MStatus.objects.all().order_by('-id')
    serializer_class = MStatusSerializer
    permission_classes = [IsAuthenticated]

class PaymentHistoryViewSet(viewsets.ModelViewSet):
    queryset = PaymentHistory.objects.all().order_by('-id')
    serializer_class = PaymentHistorySerializer
    permission_classes = [IsAuthenticated]


@api_view(['GET'])
@permission_classes([IsAuthenticatedOrReadOnly])
@csrf_exempt
def mstatusbymachineid(request, M_Id=None):
    if request.method == 'GET':
        try:
            messages = MStatus.objects.filter(m_id=M_Id,mode__isnull=True).order_by('-id')
            serializer = MStatusSerializer(
                messages, many=True, context={'request': request})
            # return JsonResponse(serializer.data, safe=False)
            return JsonResponse({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            logger(str(e), 'e')
            return JsonResponse({'success': 0, 'message': 'Not Found', 'result': serializer.errors})
from django.utils import timezone
from datetime import datetime, timedelta
import traceback
import inspect

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getmstatusdetail(request):
    if request.method == 'GET':
        try:
            # Get all machine IDs from MachineMaster
            machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

            # Define time threshold (two minutes ago)
            threshold_time = timezone.now() - timedelta(minutes=2)

            # Fetch machine data along with their creation times using cursor
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT payload, created_at
                    FROM userapp_mqttdata
                    WHERE created_at >= %s
                    """,
                    [threshold_time]
                )
                rows = cursor.fetchall()

            # Prepare dictionaries to store machine statuses and durations
            machine_status = {}
            machine_off = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"
            offline_durations = {machine_id: timedelta(0) for machine_id in machine_ids}  # Initialize all machines' offline durations as 0
            online_durations = {machine_id: timedelta(0) for machine_id in machine_ids}  # Initialize all machines' online durations as 0

            # Iterate through the fetched data to determine machine statuses and calculate durations
            # for row in rows:
            #     payload = json.loads(row[0])
            #     machine_id = payload.get('M_Id')
            #     if machine_id in machine_ids:
            #         machine_status[machine_id] = "ON"
            #         machine_off.pop(machine_id, None)  # Remove machine from the OFF list if it's online
            #     else:
            #         created_at = row[1]
            #         offline_duration = timezone.now() - created_at
            #         offline_durations[machine_id] += offline_duration
            for row in rows:
                payload = json.loads(row[0])
                machine_id = payload.get('M_Id')
                if machine_id in machine_ids:
                    machine_status[machine_id] = "ON"
                    machine_off.pop(machine_id, None)  # Remove machine from the OFF list if it's online
                    # Calculate offline duration for machines that are online
                    if machine_id in offline_durations:
                        created_at = row[1]
                        offline_duration = timezone.now() - created_at
                        offline_durations[machine_id] += offline_duration
                else:
                    created_at = row[1]
                    offline_duration = timezone.now() - created_at
                    offline_durations[machine_id] = offline_duration  

            # Calculate online durations for machines that are currently online
            for machine_id, status in machine_status.items():
                if status == "ON":
                    online_duration = timezone.now() - threshold_time
                    online_durations[machine_id] += online_duration

            # Fetch latest stock data for each machine
            stocks = []
            for machine_id in machine_ids:
                stock = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()
                if stock:
                    # color = 'green' if 16 <= stock.stock <= 200 else 'yellow' if 6 <= stock.stock <= 15 else 'red' if 1 <= stock.stock <= 5 else 'Refill'
                    # Get the stock value for the current machine
                    stock_value = stock.stock

                    # Determine color dynamically based on stock value
                    if stock_value >= 16:
                        color = 'green'
                    elif 6 <= stock_value <= 15:
                        color = 'yellow'
                    elif 1 <= stock_value <= 5:
                        color = 'red'
                    else:
                        color = 'Refill'
                    machine_on_off = "ON" if machine_status.get(machine_id) else "OFF"
                    offline_duration_hours = offline_durations[machine_id].total_seconds() / 3600  # Calculate offline duration in hours
                    online_duration_hours = online_durations[machine_id].total_seconds() / 3600  # Calculate online duration in hours
                    stocks.append({
                        'M_Id': stock.m_id,
                        'Capacity': stock.capacity,
                        'Stock': stock.stock,
                        'Status': color,
                        'Mode': stock.mode,
                        'Color': color,
                        'created_at': stock.created_at,
                        'created_by': stock.created_by,
                        'machine_status': machine_on_off,
                        'offline_duration_hours': offline_duration_hours,
                        'online_duration_hours': online_duration_hours
                    })

            # Prepare the response
            response_data = {
                'success': 1,
                'message': 'Data Found',
                'results': stocks
            }
            logger.info("Data retrieval successful")
            return Response(response_data)

        except Exception as e:
            # Get the name of the current function
            current_function = inspect.currentframe().f_code.co_name
            # Log the function name along with the error message
            logger.error(f"Error in function {current_function}: {str(e)}")
            logger.error(traceback.format_exc())  # Log the traceback
            return Response({'success': 0, 'message': str(e)})

        
@api_view(['GET'])
@permission_classes([IsAuthenticatedOrReadOnly])
@csrf_exempt
def mstatusbymachineid_for_payment(request, M_Id=None):
    if request.method == 'GET':
        try:
            messages = MStatus.objects.filter(mode__isnull=False).filter(m_id=M_Id)
            serializer = MStatusSerializer(
                messages, many=True, context={'request': request})
            # return JsonResponse(serializer.data, safe=False)
            return JsonResponse({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            logger(str(e), 'e')
            return JsonResponse({'success': 0, 'message': 'Not Found', 'result': serializer.errors})

@api_view(['GET'])
@permission_classes([IsAuthenticatedOrReadOnly])
@csrf_exempt
def get_customer_list(request):
    if request.method == 'GET':
        try:
            messages = User.objects.filter(role = '3').order_by('-id')
            serializer = UserMasterSerializer(
                messages, many=True, context={'request': request})
            # return JsonResponse(serializer.data, safe=False)
            return JsonResponse({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            logger(str(e), 'e')
            return JsonResponse({'success': 0, 'message': 'Not Found', 'result': serializer.errors})
        
@api_view(['GET'])
@permission_classes([IsAuthenticatedOrReadOnly])
@csrf_exempt
def get_user_list(request):
    if request.method == 'GET':
        try:
            
            messages = User.objects.filter(role = '4',created_by =request.user.id).order_by('-id')
            serializer = UserMasterSerializer(
                messages, many=True, context={'request': request})
            # return JsonResponse(serializer.data, safe=False)
            return JsonResponse({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            logger(str(e), 'e')
            return JsonResponse({'success': 0, 'message': 'Not Found', 'result': serializer.errors})
        
        
@api_view(['GET'])
@permission_classes([IsAuthenticatedOrReadOnly])
@csrf_exempt
def get_user_list_for_customer(request ,pk):
    if request.method == 'GET':
        try:
            
            messages = User.objects.filter(role = '4',created_by =pk).order_by('-id')
            serializer = UserMasterSerializer(
                messages, many=True, context={'request': request})
            # return JsonResponse(serializer.data, safe=False)
            return JsonResponse({'success': 1, 'message': 'Data Found', 'result': serializer.data})
        except Exception as e:
            logger(str(e), 'e')
            return JsonResponse({'success': 0, 'message': 'Not Found', 'result': serializer.errors})
        

import base64
import jwt
import json
from paho.mqtt import client as mqtt
#old 
# @api_view(['POST'])
# @csrf_exempt
# def paymentcallback(request):
#     if request.method == 'POST':
#         try:
#             encoded_string = request.data['response']
#             logging.info(f'payment Data: {str(encoded_string)}')
#             jwt_token = base64.urlsafe_b64decode(encoded_string + '=' * (4 - len(encoded_string) % 4))
#             logging.info(f'payment Data: {str(jwt_token)}')
#             json_string = jwt_token.decode('utf-8')

#             # Parse the JSON string to a Python dictionary
#             json_data = json.loads(json_string)
#             # MQTT message payload (response)
#             if json_data['code'] ==  'PAYMENT_SUCCESS':
#                 print(json_data['data'])
#                 storeName= QrCode.objects.filter(qr_code_id = json_data['data']['storeId'])
#                 logging.info(f'payment Data: {str(storeName)}')
#                 if len(storeName) != 0:
#                     payload = json.dumps({
#                         "MACHINE_ID": storeName[0].qr_store_name,
#                         "AID": json_data['data']['merchantId'],
#                         "TID": json_data['data']['transactionId'],
#                         "RID": json_data['data']['terminalId'],
#                         "QTY":1
#                         })

#                     paho.mqtt.publish.single(TOPIC_M_ORDER, payload, hostname=BROKER, port=PORT , auth={'username': MQTT_USERNAME, 'password': MQTT_PASSWORD})
#             return JsonResponse({'success': 1, 'message': 'Data Found', 'result': json_data})
#         except Exception as e:
#             logging.error(f'Error payment: {str(e)}')
#             return JsonResponse({'success': 0, 'message': 'Not Found', 'result': ''})
payment_logger = logging.getLogger('payment_logger')

@api_view(['POST'])
@csrf_exempt
def paymentcallback(request):
    if request.method == 'POST':
        try:
            encoded_string = request.data.get('response')
            payment_logger.info(f'Payment Data (encoded): {encoded_string}')
            
            if not encoded_string:
                payment_logger.error('Encoded string is missing or empty')
                return JsonResponse({'success': 0, 'message': 'Missing encoded string in request data'})

            try:
                padded_encoded_string = encoded_string + '=' * (4 - len(encoded_string) % 4)
                jwt_token = base64.urlsafe_b64decode(padded_encoded_string)
                payment_logger.info(f'Payment Data (decoded JWT): {jwt_token}')
            except Exception as decode_error:
                payment_logger.error(f'Error decoding JWT token: {decode_error}', exc_info=True)
                return JsonResponse({'success': 0, 'message': 'Error decoding JWT token', 'result': str(decode_error)})

            try:
                json_string = jwt_token.decode('utf-8')
                json_data = json.loads(json_string)
                payment_logger.info(f'Payment Data (JSON): {json_data}')
            except Exception as json_decode_error:
                payment_logger.error(f'Error decoding JSON data: {json_decode_error}', exc_info=True)
                return JsonResponse({'success': 0, 'message': 'Error decoding JSON data', 'result': str(json_decode_error)})

            required_fields = ['code', 'data']
            if not all(field in json_data for field in required_fields):
                payment_logger.error(f'Missing required fields in JSON data: {json_data}')
                return JsonResponse({'success': 0, 'message': 'Missing required fields in JSON data'})

            if json_data.get('code') == 'PAYMENT_SUCCESS':
                data_fields = ['storeId', 'merchantId', 'transactionId', 'terminalId']
                if not all(field in json_data['data'] for field in data_fields):
                    payment_logger.error(f'Missing required data fields in JSON data: {json_data["data"]}')
                    return JsonResponse({'success': 0, 'message': 'Missing required data fields in JSON data'})

                store_name_record = QrCode.objects.filter(qr_store_name=json_data['data'].get('storeId'))
                if store_name_record.exists():
                    store_name = store_name_record.first().qr_store_name
                    payment_logger.info(f'Store Name: {store_name}')
                    
                    payload = json.dumps({
                        "MACHINE_ID": store_name,
                        "AID": json_data['data'].get('merchantId'),
                        "TID": json_data['data'].get('transactionId'),
                        "RID": json_data['data'].get('terminalId'),
                        "QTY": 1
                    })
                    payment_logger.info(f'MQTT Payload: {payload}')
                    
                    try:
                        paho.mqtt.publish.single(TOPIC_M_ORDER, payload, hostname=BROKER, port=PORT, auth={'username': MQTT_USERNAME, 'password': MQTT_PASSWORD})
                        payment_logger.info('Payment success and MQTT message published successfully')
                    except Exception as mqtt_publish_error:
                        payment_logger.error(f'Error publishing to MQTT: {mqtt_publish_error}', exc_info=True)
                        return JsonResponse({'success': 0, 'message': 'Error publishing to MQTT', 'result': str(mqtt_publish_error)})
                else:
                    payment_logger.warning(f'Store ID {json_data["data"].get("storeId")} not found in the database')
                    return JsonResponse({'success': 0, 'message': 'Store ID not found', 'result': 'Store ID not found in the database'})
            else:
                payment_logger.warning(f'Payment failed with code: {json_data.get("code")}')
                return JsonResponse({'success': 0, 'message': 'Payment failed', 'result': json_data})

            return JsonResponse({'success': 1, 'message': 'Data Found', 'result': json_data})
        
        except Exception as e:
            payment_logger.error(f'Error processing payment: {e}', exc_info=True)
            return JsonResponse({'success': 0, 'message': 'An error occurred', 'result': str(e)})
        
# @api_view(['GET'])
# @csrf_exempt
# def paymentsuccess(request):
#     if request.method == 'GET':
#         try:
#             json_data={"success": True,"code": "PAYMENT_SUCCESS","message": "Your payment is successful.","data": {"merchantId": "INDEFTTECHNOLOGYSOLUTIONS","transactionId": "TXSCAN2403011707204009674431","providerReferenceId": "T2403011707226646428279","amount": 500,"updateTimestamp": 1709293046573,"paymentState": "COMPLETED","payResponseCode": "SUCCESS","transactionContext": {"qrCodeId": "EQR2402271212310092552891","storeId": "MS2402231810481193910571","terminalId": "ET2402271212310102552082"},"storeId": "MS2402231810481193910571","terminalId": "ET2402271212310102552082"}}
#             if json_data['code'] ==  'PAYMENT_SUCCESS':
#                 print(json_data['data'])
#                 storeName= QrCode.objects.filter(qr_code_id = json_data['data']['storeId'])
#                 print('hello')
#                 payment_logger.info('Data Of payment Success')
#                 print(storeName)
#                 if len(storeName) != 0: 
#                     payload = json.dumps({
#                         "MACHINE_ID": storeName[0].qr_store_name,
#                         "AID": json_data['data']['merchantId'],
#                         "TID": json_data['data']['transactionId'],
#                         "RID": json_data['data']['terminalId'],
#                         "QTY":1
#                         })
#                     payment_logger.info('Data Of payment Success')
#                     paho.mqtt.publish.single(TOPIC_M_ORDER, payload, hostname=BROKER, port=PORT, auth={'username': MQTT_USERNAME, 'password': MQTT_PASSWORD})
#             return JsonResponse({'success': 1, 'message': 'Data Found'})
#         except Exception as e:
#             payment_logger.error(f'Error payment: {str(e)}')
#             return JsonResponse({'success': 0, 'message': 'Not Found', 'result': ''})
@api_view(['GET'])
@csrf_exempt
def paymentsuccess(request):
    if request.method == 'GET':
        try:
            json_data = {
                "success": True,
                "code": "PAYMENT_SUCCESS",
                "message": "Your payment is successful.",
                "data": {
                    "merchantId": "INDEFTTECHNOLOGYSOLUTIONS",
                    "transactionId": "TXSCAN2403011707204009674431",
                    "providerReferenceId": "T2403011707226646428279",
                    "amount": 500,
                    "updateTimestamp": 1709293046573,
                    "paymentState": "COMPLETED",
                    "payResponseCode": "SUCCESS",
                    "transactionContext": {
                        "qrCodeId": "EQR2402271212310092552891",
                        "storeId": "MS2402231810481193910571",
                        "terminalId": "ET2402271212310102552082"
                    },
                    "storeId": "MS2402231810481193910571",
                    "terminalId": "ET2402271212310102552082"
                }
            }

            if json_data['code'] == 'PAYMENT_SUCCESS':
                payment_logger.info(f"Payment success data: {json_data['data']}")

                store_name_record = QrCode.objects.filter(qr_code_id=json_data['data']['storeId'])
                payment_logger.info('Store record query executed')

                if store_name_record.exists():
                    store_name = store_name_record.first().qr_store_name
                    payment_logger.info(f'Store Name: {store_name}')

                    payload = json.dumps({
                        "MACHINE_ID": store_name,
                        "AID": json_data['data']['merchantId'],
                        "TID": json_data['data']['transactionId'],
                        "RID": json_data['data']['terminalId'],
                        "QTY": 1
                    })
                    payment_logger.info(f'MQTT Payload: {payload}')

                    try:
                        paho.mqtt.publish.single(TOPIC_M_ORDER, payload, hostname=BROKER, port=PORT, auth={'username': MQTT_USERNAME, 'password': MQTT_PASSWORD})
                        payment_logger.info('MQTT message published successfully')
                    except Exception as mqtt_publish_error:
                        payment_logger.error(f'Error publishing to MQTT: {mqtt_publish_error}', exc_info=True)
                        return JsonResponse({'success': 0, 'message': 'Error publishing to MQTT', 'result': str(mqtt_publish_error)})
                else:
                    payment_logger.warning(f'Store ID {json_data["data"]["storeId"]} not found in the database')
                    return JsonResponse({'success': 0, 'message': 'Store ID not found', 'result': 'Store ID not found in the database'})
            else:
                payment_logger.warning(f'Payment failed with code: {json_data["code"]}')
                return JsonResponse({'success': 0, 'message': 'Payment failed', 'result': json_data})

            return JsonResponse({'success': 1, 'message': 'Data Found'})
        
        except Exception as e:
            payment_logger.error(f'Error processing payment: {str(e)}', exc_info=True)
            return JsonResponse({'success': 0, 'message': 'An error occurred', 'result': str(e)})

from django.db.models import Count
from django.db import connection


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def all_schema_names(request):
    """
    Retrieve all schema names.
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT schema_name FROM information_schema.schemata WHERE schema_name NOT IN ('information_schema', 'public')")
            rows = cursor.fetchall()
            schema_names = [row[0] for row in rows]
            return Response({"schema_names": schema_names}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import MQTTData
from datetime import datetime, timedelta
from django.core.cache import cache
from django.utils import timezone
import json
from django.core.exceptions import ObjectDoesNotExist

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_machine_status(request):
    current_time = timezone.now()
    threshold_time = current_time - timedelta(minutes=2)

    # Get all machine IDs from MachineMaster
    machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

    # Fetch machine data along with their creation times
    machine_data = MQTTData.objects.filter(created_at__gte=threshold_time).values('payload', 'created_at')

    # Prepare dictionaries to store machine statuses
    machine_status = {}
    machine_off = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"

    # Iterate through the fetched data
    for data in machine_data:
        payload = data['payload']
        created_at = data['created_at']
        
        # Deserialize the payload string into a dictionary
        payload_dict = json.loads(payload)
        
        # Access the 'M_Id' key from the payload dictionary
        M_Id = payload_dict.get('M_Id')

        if M_Id in machine_ids:  # Check if the machine ID is in the MachineMaster list
            # Determine the status of the machine based on the last data received
            if created_at >= threshold_time:
                machine_status[M_Id] = "ON"
                machine_off.pop(M_Id, None)  # Remove the machine from the OFF list if it's online

    # Prepare the response
    response_data = {
        'success': 1,
        'message': 'Data Found',
        'machine_status': machine_status,
        'machine_off': machine_off,
    }

    return Response(response_data, status=status.HTTP_200_OK)

from django.db.models import Count




@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_payment_api_view(request):
    if request.method == 'GET':
        # Retrieve MStatus objects
        mstatus_objects = MStatus.objects.all()

        # Initialize nested dictionaries for each mode type per m_id
        machine_mode_counts = defaultdict(lambda: defaultdict(int))

        # Separate counts for each mode type per machine
        for obj in mstatus_objects:
            if obj.m_id and obj.mode:
                machine_mode_counts[obj.m_id][obj.mode] += 1

        # Construct mode_objects with separated counts for each machine
        mode_objects = []
        for m_id, modes in machine_mode_counts.items():
            # Extract counts for 'COIN' and 'QR', defaulting to 0 if not present
            coin_count = modes.get('COIN', 0)
            qr_count = modes.get('QR', 0)
            
            # You can calculate percentages if needed here, similar to previous examples
            # For simplicity, this example focuses on the counts

            mode_object = {
                'm_id': m_id,
                'coin_mode_counts': coin_count,
                'qr_mode_counts': qr_count
                # Add percentages or other calculations as needed
            }
            mode_objects.append(mode_object)

        # Constructing the response data
        response_data = {
            'success': 1,
            'message': 'Data Found',
            'result': mode_objects
        }

        return Response(response_data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_payment_api_view_user(request):
    try:
        if request.method == 'GET':
            # Fetch all distinct machine IDs associated with the logged-in user
            assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
            
            

            # Retrieve MStatus objects related to the logged-in user
            mstatus_objects = MStatus.objects.filter(m_id__in=[machine.machine.machine_id for machine in assigned_machines])
            
            # Initialize nested dictionaries for each mode type per m_id
            machine_mode_counts = defaultdict(lambda: defaultdict(int))

            # Separate counts for each mode type per machine
            for obj in mstatus_objects:
                if obj.m_id and obj.mode:
                    machine_mode_counts[obj.m_id][obj.mode] += 1

            # Construct mode_objects with separated counts for each machine
            mode_objects = []
            for m_id in [machine.machine.machine_id for machine in assigned_machines]:
                # Extract counts for 'COIN' and 'QR', defaulting to 0 if not present
                coin_count = machine_mode_counts[m_id].get('COIN', 0)
                qr_count = machine_mode_counts[m_id].get('QR', 0)
                
                mode_object = {
                    'm_id': m_id,
                    'coin_mode_counts': coin_count,
                    'qr_mode_counts': qr_count
                }
                mode_objects.append(mode_object)

            # Constructing the response data
            response_data = {
                'success': 1,
                'message': 'Data Found',
                'result': mode_objects
            }

            return Response(response_data, status=status.HTTP_200_OK)
        else:
            return Response({'success': 0, 'message': 'Invalid request method'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_qr_coin_view_for_dashboard(request):
    if request.method == 'GET':
        # Retrieve MStatus objects
        mstatus_objects = MStatus.objects.all()

        # Initialize dictionaries to store counts of 'COIN' and 'QR' modes
        coin_count = 0
        qr_count = 0

        # Count occurrences of each mode type
        for obj in mstatus_objects:
            if obj.mode == 'COIN':
                coin_count += 1
            elif obj.mode == 'QR':
                qr_count += 1

        # Calculate percentages
        total_count = coin_count + qr_count
        coin_percentage = (coin_count / total_count) * 100 if total_count != 0 else 0
        qr_percentage = (qr_count / total_count) * 100 if total_count != 0 else 0

        data={
            'coin_count': coin_count,
            'qr_count': qr_count,
            'coin_percentage': coin_percentage,
            'qr_percentage': qr_percentage
        }

        # Constructing the response data
        response_data = {
            'success': 1,
            'message': 'Data Found',
            'result':data
            
        }

        return Response(response_data, status=status.HTTP_200_OK)

from dateutil import parser
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from dateutil import parser
from .models import MStatus
from collections import defaultdict

import pandas as pd

from rest_framework import status
from .models import MStatus  # Ensure this matches your model's actual location and name
from collections import defaultdict

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def post_payment_report_excel(request):
    date_from_raw = request.data.get('from_date')
    date_to_raw = request.data.get('to_date')

    if not date_from_raw or not date_to_raw:
        return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()
        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        mstatus_records = MStatus.objects.filter(created_at__range=(date_from, date_to))

        mode_counts_per_machine = defaultdict(lambda: defaultdict(int))

        for record in mstatus_records:
            if record.m_id and record.mode:
                mode_counts_per_machine[record.m_id][record.mode] += 1

        data_for_response = []
        for m_id, modes in mode_counts_per_machine.items():
            row = {'m_id': m_id}
            total_modes = sum(modes.values())

            # Ensure both coin and qr mode counts are initialized to 0
            row['coin_mode_counts'] = modes.get('COIN', 0)
            row['qr_mode_counts'] = modes.get('QR', 0)

            # Calculate percentages, ensuring division by zero is handled
            row['coin_mode_percentage'] = (row['coin_mode_counts'] / total_modes) * 100 if total_modes else 0
            row['qr_mode_percentage'] = (row['qr_mode_counts'] / total_modes) * 100 if total_modes else 0

            data_for_response.append(row)

        return Response({
            'success': 1,
            'message': 'Data found',
            'result': data_for_response
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

      
from rest_framework.response import Response
from rest_framework import status
from .models import MQTTData, MStatus, RefillStockHistory
from django.db.models import Max
from django.utils import timezone        
from openpyxl import Workbook
from django.http import FileResponse, HttpResponse
from datetime import datetime
import os
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from django.http import FileResponse, HttpResponse
from datetime import datetime
import datetime
from django.conf import settings
import os
from openpyxl.styles import Alignment,Border, Side
from openpyxl.styles import Font
from openpyxl.utils.units import points_to_pixels    

from openpyxl.workbook import Workbook

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_and_download_excel_report_for_user(request):
    try:
       
        
        today_date = timezone.localdate()
        
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Check if both 'from_date' and 'to_date' are provided
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, "message": "Both 'from_date' and 'to_date' are required."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        # Get all machine IDs from MachineMaster
        machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

        # Define time threshold (two minutes ago)
        threshold_time = timezone.now() - timedelta(minutes=2)

        # Fetch machine data along with their creation times using cursor
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload, created_at
                FROM userapp_mqttdata
                WHERE created_at >= %s
                """,
                [threshold_time]
            )
            rows = cursor.fetchall()

        # Prepare dictionaries to store machine statuses and stocks
        machine_status = {m_id: "OFF" for m_id in machine_ids}
        stocks = {m_id: None for m_id in machine_ids}

        # Iterate through the fetched data to determine machine statuses
        for row in rows:
            payload = json.loads(row[0])
            m_id = payload.get('M_Id')
            if m_id in machine_ids:
                machine_status[m_id] = "ON"

        # Fetch latest stock data for each machine from MStatus model
        mstatus_data = MStatus.objects.filter(m_id__in=machine_ids)
        for mstatus in mstatus_data:
            stocks[mstatus.m_id] = mstatus.stock

        # Fetch latest refill information for each machine
        refill_data = {}
        for machine_id in machine_ids:
            all_machine_ids = MStatus.objects.filter(
                m_id=machine_id,
                stock=F('capacity')  # Filter where stock capacity equals stock
            ).values('m_id', 'created_at__date', 'stock').annotate(
                count_capacity_equals_stock=Count('id')
            )

            machine_data = []

            for entry in all_machine_ids:
                stock_after_refill = entry['stock']

                stock_before_date_query = MStatus.objects.filter(
                    m_id=machine_id,
                    created_at__date__lt=entry['created_at__date']
                ).order_by('-created_at')

                stock_before_date = stock_before_date_query.values_list('stock', flat=True).first()

                # To ensure that the stock before date exists and it's not a refill that just happened
                if stock_before_date is not None and stock_before_date_query.count() > 1:
                    # Calculate refill quantity if stock_before_date exists
                    refill_quantity = stock_after_refill - stock_before_date

                    machine_data.append({
                        'machine_id': entry['m_id'],
                        'refilled_date': entry['created_at__date'],
                        'refilled_quantity': refill_quantity
                    })

            refill_data[machine_id] = machine_data
                
        # Query data based on the date range and assigned_user being None
        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from,date_to),assigned_customer__isnull=True)
        
        current_datetime = datetime.now()
            
        # Format the current date and time as strings
        report_date = current_datetime.strftime("%Y-%m-%d")
       
        report_time = current_datetime.strftime("%H:%M:%S")
       
        serialized_data = []
        for item in report_data:
            machine_id = item.machine.machine_id
            serialized_item = {
                'report_date': report_date,  # Fill this with actual report date if available
                'report_time': report_time,  # Fill this with actual report time if available
                'tendryl': {
                    'user_id': item.assigned_user.id,
                    'name': item.assigned_user.name if item.assigned_user else None,
                    'user_moblie_no': item.assigned_user.mobile_no if item.assigned_user else None,
                    'email': item.assigned_user.email if item.assigned_user else None,
                },
                'machine': {
                    'machine_id': item.machine.machine_id,
                    'model_no': item.machine.model_number.model_no if item.machine.model_number else None,
                    'location': item.machine.installation_location if item.machine.installation_location else None,
                    'current_status_on_off': machine_status.get(item.machine.machine_id, "OFF"),
                    'current_inventory_count': stocks.get(item.machine.machine_id, None),
                },
                'product': {
                    'product_type': item.machine.product.product_type if item.machine.product else None,  
                    'refill_date': None,  
                    'refill_quantity': None,  
                    'recent_case_collection_date': '',  
                    'recent_amount_collected': '',  
                    
                },
            }

            serialized_data.append(serialized_item)

            refill_items = refill_data.get(machine_id, [])
            for refill_item in refill_items:
                refill_serialized_item = serialized_item.copy()  # Create a copy of the base serialized item
                refill_serialized_item['product']['refill_date'] = refill_item['refilled_date']
                refill_serialized_item['product']['refill_quantity'] = refill_item['refilled_quantity']
                serialized_data.append(refill_serialized_item)

            # Append the base serialized item if no refill records found
            if not refill_items:
                serialized_data.append(serialized_item)
            # Create Excel workbook and add data to worksheet
        wb = Workbook()
        ws = wb.active

        # Add header row
        ws.append([
            "No", "Report Date", "Report Time", 
            "TENDRYL", "", "", "", 
            "MACHINE", "", "", "", 
            "PRODUCT", "", "", "", "",
        ])

        # Set background colors for specific columns
        customer_fill = PatternFill(start_color="43E443", end_color="43E443", fill_type="solid")
        machine_fill = PatternFill(start_color="f8356f", end_color="f8356f", fill_type="solid")
        product_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Apply colors to the "CUSTOMER", "MACHINE", and "PRODUCT" sections
        for col in range(1, 17):  # Loop through relevant columns
            if 4 <= col <= 7:
                ws.cell(row=1, column=col).fill = customer_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make CUSTOMER header bold
            elif 8 <= col <= 12:
                ws.cell(row=1, column=col).fill = machine_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make MACHINE header bold
            elif 13 <= col <= 17:
                ws.cell(row=1, column=col).fill = product_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make PRODUCT header bold

        # Add borders to the cells
        border_bold = Border(left=Side(style='medium'), 
                            right=Side(style='medium'), 
                            top=Side(style='medium'), 
                            bottom=Side(style='medium'))
                
        # Add borders to the cells
        border_medium_bold = Border(left=Side(style='thin'), 
                                    # right=Side(style='medium'), 
                                    # top=Side(style='medium'), 
                                    bottom=Side(style='medium'))

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=17):
            for cell in row:
                cell.border = border_bold


        # Add values for "Report" and "Date" in the first set of merged cell
        # Merge cells A1:A2
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Merge cells B1:B2
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Merge cells C1:C2
        ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                

        # Merge cells for "CUSTOMER"
        ws.cell(row=1, column=4).value = "TENDRYL"
        ws.cell(row=1, column=4).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)  # CUSTOMER

        # Add details for "CUSTOMER" section in the second row
        ws.cell(row=2, column=4).value = "User ID"
        ws.cell(row=2, column=4).font = Font(bold=True)  # Make SPOC ID bold
        ws.cell(row=2, column=4).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=5).value = "Name"
        ws.cell(row=2, column=5).font = Font(bold=True)  # Make Name bold
        ws.cell(row=2, column=5).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=6).value = "Mobile"
        ws.cell(row=2, column=6).font = Font(bold=True)  # Make Mobile bold
        ws.cell(row=2, column=6).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=7).value = "Email"
        ws.cell(row=2, column=7).font = Font(bold=True)  # Make Email bold
        ws.cell(row=2, column=7).alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells for "MACHINE"
        ws.cell(row=1, column=8).value = "MACHINE"
        ws.cell(row=1, column=8).font = Font(bold=True,color=WHITE)
        ws.cell(row=1, column=8).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)  # MACHINE

        # Add details for "MACHINE" section in the second row
        ws.cell(row=2, column=8).value = "Machine ID"
        ws.cell(row=2, column=8).font = Font(bold=True)  # Make Machine ID bold
        ws.cell(row=2, column=8).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=9).value = "Model\nNo"
        ws.cell(row=2, column=9).font = Font(bold=True)  # Make Model No bold
        ws.cell(row=2, column=9).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        
        ws.cell(row=2, column=10).value = "Location"
        ws.cell(row=2, column=10).font = Font(bold=True)  # Make Location bold
        ws.cell(row=2, column=10).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=11).value = "Current\nStatus\n(On / off)"
        ws.cell(row=2, column=11).font = Font(bold=True)  # Make Current Status (Off / On) bold
        ws.cell(row=2, column=11).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        ws.cell(row=2, column=12).value = "Current\nInventory\nCount"
        ws.cell(row=2, column=12).font = Font(bold=True)  # Make Current Status (Off / On) bold
        ws.cell(row=2, column=12).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        # Merge cells for "PRODUCT"
        ws.cell(row=1, column=13).value = "PRODUCT"
        ws.cell(row=1, column=13).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=17)  # PRODUCT

        # Add details for "PRODUCT" section in the second row
        ws.cell(row=2, column=13).value = "Product\nType"
        ws.cell(row=2, column=13).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=13).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        ws.cell(row=2, column=14).value = "Recent\nRefill Date"
        ws.cell(row=2, column=14).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=14).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        ws.cell(row=2, column=15).value = "Quantity\nRefilled"
        ws.cell(row=2, column=15).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=15).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        ws.cell(row=2, column=16).value = "Recent Cash\nCollection Date"
        ws.cell(row=2, column=16).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=16).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        ws.cell(row=2, column=17).value = "Recent Amount\nCollected"
        ws.cell(row=2, column=17).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=17).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

       

        # Set height for row 2
        # ws.row_dimensions[2].height = points_to_pixels(28)
        ws.row_dimensions[2].height = 45

        # Set width for columns in row 2
        for col in range(1, 18):  # Assuming you have 17 columns
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Add report data
            for idx, data in enumerate(serialized_data, start=3):
               
                ws.cell(row=idx, column=1).value = idx - 2  # Serial number
                ws.cell(row=idx, column=1).border = border_bold
                ws.cell(row=idx, column=2).value = data["report_date"]
                ws.cell(row=idx, column=2).border = border_bold
                ws.cell(row=idx, column=3).value = data["report_time"]
                ws.cell(row=idx, column=3).border = border_bold

                # Customer details
                tendryl = data["tendryl"]
                ws.cell(row=idx, column=4).value = tendryl["user_id"]
                ws.cell(row=idx, column=4).border = border_bold
                ws.cell(row=idx, column=5).value = tendryl["name"]
                ws.cell(row=idx, column=5).border = border_bold
                ws.cell(row=idx, column=6).value = tendryl["user_moblie_no"]
                ws.cell(row=idx, column=6).border = border_bold
                ws.cell(row=idx, column=7).value = tendryl["email"]
                ws.cell(row=idx, column=7).border = border_bold

                # Machine details
                machine = data["machine"]
                ws.cell(row=idx, column=8).value = machine["machine_id"]
                ws.cell(row=idx, column=8).border = border_bold
                ws.cell(row=idx, column=9).value = machine["model_no"]
                ws.cell(row=idx, column=9).border = border_bold
                ws.cell(row=idx, column=10).value = machine["location"]
                ws.cell(row=idx, column=10).border = border_bold
                ws.cell(row=idx, column=11).value = machine["current_status_on_off"]
                ws.cell(row=idx, column=11).border = border_bold
                ws.cell(row=idx, column=12).value = machine["current_inventory_count"]
                ws.cell(row=idx, column=12).border = border_bold

                # Product details
                product = data["product"]
                ws.cell(row=idx, column=13).value = product["product_type"]
                ws.cell(row=idx, column=13).border = border_bold
                ws.cell(row=idx, column=14).value = product["refill_date"]
                ws.cell(row=idx, column=14).border = border_bold
                ws.cell(row=idx, column=15).value = product["refill_quantity"]
                ws.cell(row=idx, column=15).border = border_bold
                ws.cell(row=idx, column=16).value = product["recent_case_collection_date"]
                ws.cell(row=idx, column=16).border = border_bold
                ws.cell(row=idx, column=17).value = product["recent_amount_collected"]
                ws.cell(row=idx, column=17).border = border_bold
                

         # Auto-adjust column width based on content size
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        # Add thin and bold border to all cells
        thin_border_bold = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.border = thin_border_bold

        # Add medium bold border to the last row
        for cell in ws[ws.max_row]:
            cell.border = border_medium_bold
    

            # Generate file name with date and time
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f'user_excel_report__{now}.xlsx'
            
        # Save the workbook in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Set response content type
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Set content disposition header to force download
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        return response
        
    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mis_report_for_customer(request):
    try:
        # Get all machine IDs from MachineMaster
        machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

        # Define time threshold (two minutes ago)
        threshold_time = timezone.now() - timedelta(minutes=2)

        # Fetch machine data along with their creation times using cursor
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload, created_at
                FROM userapp_mqttdata
                WHERE created_at >= %s
                """,
                [threshold_time]
            )
            rows = cursor.fetchall()

        # Prepare dictionaries to store machine statuses and current inventory counts
        machine_status = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"
        stocks = defaultdict(int)

        # Iterate through the fetched data to determine machine statuses and calculate current inventory counts
        for row in rows:
            payload = json.loads(row[0])
            M_Id = payload.get('M_Id')
            if M_Id in machine_ids:
                machine_status[M_Id] = "ON"
                stocks[M_Id] = payload.get('Stock')
        
        # Fetch latest stock data for each machine from MStatus model
        mstatus_data = MStatus.objects.filter(m_id__in=machine_ids)
        for mstatus in mstatus_data:
            stocks[mstatus.m_id] = mstatus.stock

        # Fetch machine-user mapping data
        report_data = MachineUserMapping.objects.filter(assigned_customer__isnull=False).order_by('-id')

        # Serialize the data
        serialized_data = []
        for item in report_data:
            serialized_data.append({
                'spoc_id': item.assigned_customer.spoc_id if item.assigned_customer else None,
                'name': item.assigned_customer.name if item.assigned_customer else None,
                'user_mobile_no': item.assigned_customer.mobile_no if item.assigned_customer else None,
                'email': item.assigned_customer.email if item.assigned_customer else None,
                'machine_id': item.machine.machine_id,
                'model_no': item.machine.model_number.model_no if item.machine.model_number else None,
                'location': item.machine.installation_location if item.machine else None,
                'product': item.machine.model_number.name if item.machine.model_number else None,
                'date': item.created_at.date().isoformat(),
                'time': item.created_at.time().isoformat(),
                'type': item.machine.product.product_type if item.machine.product else None,
                'amount': item.machine.product.amount if item.machine.product else None,
                'success': '',
                'refund_status': '',
                'machine_status': machine_status.get(item.machine.machine_id, "OFF"),  # Fetch machine status
                'current_inventory_count': stocks.get(item.machine.machine_id, None),  # Fetch current inventory count
            })

        return JsonResponse({'success': 1, 'message': 'Data Found', 'result': serialized_data})
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mis_report_for_user(request):
    try:
        
        # Get all machine IDs from MachineMaster
        machine_ids = list(MachineMaster.objects.values_list('machine_id', flat=True))

        # Define time threshold (two minutes ago)
        threshold_time = timezone.now() - timedelta(minutes=2)

        # Fetch machine data along with their creation times using a cursor
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload
                FROM userapp_mqttdata
                WHERE created_at >= %s
                """,
                [threshold_time]
            )
            rows = cursor.fetchall()

        # Prepare dictionaries to store machine statuses and stocks
        machine_status = {m_id: "OFF" for m_id in machine_ids}
        stocks = {m_id: None for m_id in machine_ids}

        # Iterate through the fetched data to determine machine statuses
        for row in rows:
            payload = json.loads(row[0])
            m_id = payload.get('M_Id')
            if m_id in machine_ids:
                machine_status[m_id] = "ON"

        # Fetch latest stock data for each machine from MStatus model
        mstatus_data = MStatus.objects.filter(m_id__in=machine_ids)
        for mstatus in mstatus_data:
            stocks[mstatus.m_id] = mstatus.stock

        # Fetch latest refill information for each machine
        refill_data = {}
        for machine_id in machine_ids:
            all_machine_ids = MStatus.objects.filter(
                m_id=machine_id,
                stock=F('capacity')  # Filter where stock capacity equals stock
            ).values('m_id', 'created_at__date', 'stock').annotate(
                count_capacity_equals_stock=Count('id')
            )

            machine_data = []

            for entry in all_machine_ids:
                stock_after_refill = entry['stock']

                stock_before_date_query = MStatus.objects.filter(
                    m_id=machine_id,
                    created_at__date__lt=entry['created_at__date']
                ).order_by('-created_at')

                stock_before_date = stock_before_date_query.values_list('stock', flat=True).first()

                # To ensure that the stock before date exists and it's not a refill that just happened
                if stock_before_date is not None and stock_before_date_query.count() > 1:
                    # Calculate refill quantity if stock_before_date exists
                    refill_quantity = stock_after_refill - stock_before_date

                    machine_data.append({
                        'machine_id': entry['m_id'],
                        'refilled_date': entry['created_at__date'],
                        'refilled_quantity': refill_quantity
                    })

            refill_data[machine_id] = machine_data

        # Fetch latest user-machine mappings
        report_data = MachineUserMapping.objects.filter(assigned_customer__isnull=True).order_by('-id')

        # Serialize data
        serialized_data = []
        for item in report_data:
            machine_id = item.machine.machine_id

            # Include basic machine data in serialized item
            serialized_item = {
                'user_id': item.assigned_user.id if item.assigned_user else None,
                'name': item.assigned_user.name if item.assigned_user else None,
                'user_mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
                'email': item.assigned_user.email if item.assigned_user else None,
                'machine_id': machine_id,
                'model_no': item.machine.model_number.model_no if item.machine and item.machine.model_number else None,
                'location': item.machine.installation_location if item.machine else None,
                'machine_status': machine_status.get(machine_id, "OFF"),
                'current_inventory_count': stocks.get(machine_id, None),
                'product': item.machine.model_number.name if item.machine and item.machine.model_number else None,
                'type': item.machine.product.product_type if item.machine and item.machine.product else None,
                'amount': '',  # This needs to be clarified where it comes from
                'success': '',  # This needs to be clarified where it comes from
                'refund_status': '',  # This needs to be clarified where it comes from
            }

            # Append separate serialized items for each refill record
            refill_items = refill_data.get(machine_id, [])
            for refill_item in refill_items:
                refill_serialized_item = serialized_item.copy()  # Create a copy of the base serialized item
                refill_serialized_item['refill_date'] = refill_item['refilled_date']
                refill_serialized_item['refill_quantity'] = refill_item['refilled_quantity']
                serialized_data.append(refill_serialized_item)

            # Append the base serialized item if no refill records found
            if not refill_items:
                serialized_data.append(serialized_item)

        return Response({'success': 1, 'message': 'Data Found', 'result': serialized_data})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from datetime import datetime, timedelta
from django.utils import timezone
import calendar
from rest_framework.request import Request as RestRequest



import tempfile
from openpyxl import Workbook
import io
from io import BytesIO
from django.http import HttpResponse
import pandas as pd
from openpyxl.styles.colors import WHITE
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_and_download_excel_report_for_customer(request):
    try:
        # Get today's date in the current timezone
        today_date = timezone.localdate()
        
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        # Check if both 'from_date' and 'to_date' are provided
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, "message": "Both 'from_date' and 'to_date' are required."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)
        # Get all machine IDs from MachineMaster
        machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

        

        # Define time threshold (two minutes ago)
        threshold_time = timezone.now() - timedelta(minutes=2)

        # Fetch machine data along with their creation times using cursor
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload, created_at
                FROM userapp_mqttdata
                WHERE created_at >= %s
                """,
                [threshold_time]
            )
            rows = cursor.fetchall()

        # Prepare dictionaries to store machine statuses
        machine_status = {}
        machine_off = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"

        # Iterate through the fetched data to determine machine statuses
        for row in rows:
            payload = json.loads(row[0])
            M_Id = payload.get('M_Id')
            if M_Id in machine_ids:
                machine_status[M_Id] = "ON"
                machine_off.pop(M_Id, None)  # Remove machine from the OFF list if it's online

        # Fetch latest stock data for each machine
        stocks = {}
        for machine_id in machine_ids:
            last_stock_entry = MStatus.objects.filter(m_id=machine_id).aggregate(last_stock=Max('stock'))
            stocks[machine_id] = last_stock_entry['last_stock'] if last_stock_entry['last_stock'] is not None else None

    
        # Query data based on the date range and assigned_user being None
        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to),assigned_customer__isnull=False).order_by('-id')
        

        # if not report_data:
        #     return Response({'success': 0, "message": "No data found for the given date range."}, status=status.HTTP_404_NOT_FOUND)
        
        current_datetime = datetime.now()
            
        # Format the current date and time as strings
        report_date = current_datetime.strftime("%Y-%m-%d")
        
        report_time = current_datetime.strftime("%H:%M:%S")
        data_list = []
        for item in report_data:
            data = {
                'report_date': report_date,  
                'report_time': report_time,
                'customer': {
                    'spoc_id': item.assigned_customer.spoc_id if item.assigned_customer else None,
                    'name': item.assigned_customer.name if item.assigned_customer else None,
                    'user_moblie_no': item.assigned_customer.mobile_no if item.assigned_customer else None,
                    'email': item.assigned_customer.email if item.assigned_customer else None,
                },
                'machine': {
                    'machine_id': item.machine.machine_id,
                    'model_no': item.machine.model_number.model_no if item.machine.model_number else None,
                    'location': item.machine.installation_location if item.machine.installation_location else None,
                    'current_status_on_off': machine_status.get(item.machine.machine_id, "OFF"),
                    'current_inventory_count': stocks.get(item.machine.machine_id, None),
                },
                'product': {
                    'product': item.machine.model_number.name if item.machine.model_number else None,
                    'date': item.created_at.date().isoformat(),  
                    'time': item.created_at.time().isoformat(),  
                    'type': item.machine.product.product_type if item.machine.product else None, 
                    'amount': item.machine.product.amount if item.machine.product else None,  
                    'success': '',  
                    'refund_status': '',  
                }
            }
            data_list.append(data)
            
        
        # Create Excel workbook and add data to worksheet
        wb = Workbook()
        ws = wb.active

        # Add header row
        ws.append([
            "No", "Report Date", "Report Time", 
            "CUSTOMER", "", "", "", 
            "MACHINE", "", "", "", 
            "PRODUCT", "", "", "", "", "", "", ""
        ])


        # Set background colors for specific columns
        customer_fill = PatternFill(start_color="43E443", end_color="43E443", fill_type="solid")
        machine_fill = PatternFill(start_color="f8356f", end_color="f8356f", fill_type="solid")
        product_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Apply colors to the "CUSTOMER", "MACHINE", and "PRODUCT" sections
        for col in range(1, 20):  # Loop through relevant columns
            if 4 <= col <= 7:
                ws.cell(row=1, column=col).fill = customer_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make CUSTOMER header bold
            elif 8 <= col <= 12:
                ws.cell(row=1, column=col).fill = machine_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make MACHINE header bold
            elif 13 <= col <= 20:
                ws.cell(row=1, column=col).fill = product_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make PRODUCT header bold

        # Add borders to the cells
        border_bold = Border(left=Side(style='medium'), 
                            right=Side(style='medium'), 
                            top=Side(style='medium'), 
                            bottom=Side(style='medium'))


            # Add borders to the cells
        border_medium_bold = Border(left=Side(style='thin'), 
                                    # right=Side(style='medium'), 
                                    # top=Side(style='medium'), 
                                    bottom=Side(style='medium'))


        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=19):
            for cell in row:
                cell.border = border_bold

        # Add values for "Report" and "Date" in the first set of merged cell
        # Merge cells A1:A2
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Merge cells B1:B2
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Merge cells C1:C2
        ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Merge cells for "CUSTOMER"
        ws.cell(row=1, column=4).value = "CUSTOMER"
        ws.cell(row=1, column=4).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)  # CUSTOMER

        # Add details for "CUSTOMER" section in the second row
        ws.cell(row=2, column=4).value = "SPOC ID"
        ws.cell(row=2, column=4).font = Font(bold=True)  # Make SPOC ID bold
        ws.cell(row=2, column=4).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=5).value = "Name"
        ws.cell(row=2, column=5).font = Font(bold=True)  # Make Name bold
        ws.cell(row=2, column=5).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=6).value = "Mobile"
        ws.cell(row=2, column=6).font = Font(bold=True)  # Make Mobile bold
        ws.cell(row=2, column=6).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=7).value = "Email"
        ws.cell(row=2, column=7).font = Font(bold=True)  # Make Email bold
        ws.cell(row=2, column=7).alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells for "MACHINE"
        ws.cell(row=1, column=8).value = "MACHINE"
        ws.cell(row=1, column=8).font = Font(bold=True,color=WHITE)
        ws.cell(row=1, column=8).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)  # MACHINE

        # Add details for "MACHINE" section in the second row
        ws.cell(row=2, column=8).value = "Machine\nID"
        ws.cell(row=2, column=8).font = Font(bold=True)  # Make Machine ID bold
        ws.cell(row=2, column=8).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        
        ws.cell(row=2, column=9).value = "Model\nNo"
        ws.cell(row=2, column=9).font = Font(bold=True)  # Make Model No bold
        ws.cell(row=2, column=9).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        
        ws.cell(row=2, column=10).value = "Location"
        ws.cell(row=2, column=10).font = Font(bold=True)  # Make Location bold
        ws.cell(row=2, column=10).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=11).value = "Current\nStatus\n(On / Off)"
        ws.cell(row=2, column=11).font = Font(bold=True)  # Make Current Status (Off / On) bold
        ws.cell(row=2, column=11).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        ws.cell(row=2, column=12).value = "Current\nInventory\nCount"
        ws.cell(row=2, column=12).font = Font(bold=True)  # Make Current Status (Off / On) bold
        ws.cell(row=2, column=12).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        # Merge cells for "PRODUCT"
        ws.cell(row=1, column=13).value = "PRODUCT"
        ws.cell(row=1, column=13).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=19)  # PRODUCT

        # Add details for "PRODUCT" section in the second row
        ws.cell(row=2, column=13).value = "Product"
        ws.cell(row=2, column=13).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=13).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=14).value = "Date"
        ws.cell(row=2, column=14).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=14).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=15).value = "Time"
        ws.cell(row=2, column=15).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=15).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=16).value = "Type"
        ws.cell(row=2, column=16).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=16).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=17).value = "Amount"
        ws.cell(row=2, column=17).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=17).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=18).value = "Success\n(Y/N)"
        ws.cell(row=2, column=18).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=18).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        ws.cell(row=2, column=19).value = "Refund\nStatus"
        ws.cell(row=2, column=19).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=19).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        # Set height for row 2
        ws.row_dimensions[2].height = points_to_pixels(28)

        # Set height for row 2
        # ws.row_dimensions[2].height = points_to_pixels(28)
        ws.row_dimensions[2].height = 45

        # Set width for columns in row 2
        for col in range(1, 20):  # Assuming you have 17 columns
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Add report data
            for idx, data in enumerate(data_list, start=3):
                
                ws.cell(row=idx, column=1).value = idx - 2  # Serial number
                ws.cell(row=idx, column=1).border = border_bold
                ws.cell(row=idx, column=2).value = data["report_date"]
                ws.cell(row=idx, column=2).border = border_bold
                ws.cell(row=idx, column=3).value = data["report_time"]
                ws.cell(row=idx, column=3).border = border_bold

                # Customer details
                customer = data["customer"]
                ws.cell(row=idx, column=4).value = customer["spoc_id"]
                ws.cell(row=idx, column=4).border = border_bold
                ws.cell(row=idx, column=5).value = customer["name"]
                ws.cell(row=idx, column=5).border = border_bold
                ws.cell(row=idx, column=6).value = customer["user_moblie_no"]
                ws.cell(row=idx, column=6).border = border_bold
                ws.cell(row=idx, column=7).value = customer["email"]
                ws.cell(row=idx, column=7).border = border_bold

                # Machine details
                machine = data["machine"]
                ws.cell(row=idx, column=8).value = machine["machine_id"]
                # ws.cell(row=idx, column=1).font = Font(bold=True)  # Make mid number bold
                ws.cell(row=idx, column=8).border = border_bold
                ws.cell(row=idx, column=9).value = machine["model_no"]
                ws.cell(row=idx, column=9).border = border_bold
                ws.cell(row=idx, column=10).value = machine["location"]
                ws.cell(row=idx, column=10).border = border_bold
                ws.cell(row=idx, column=11).value = machine["current_status_on_off"]
                ws.cell(row=idx, column=11).border = border_bold
                ws.cell(row=idx, column=12).value = machine["current_inventory_count"]
                ws.cell(row=idx, column=12).border = border_bold

                # Product details
                product = data["product"]
                ws.cell(row=idx, column=13).value = product["product"]
                ws.cell(row=idx, column=13).border = border_bold
                ws.cell(row=idx, column=14).value = product["date"]
                ws.cell(row=idx, column=14).border = border_bold
                ws.cell(row=idx, column=15).value = product["time"]
                ws.cell(row=idx, column=15).border = border_bold
                ws.cell(row=idx, column=16).value = product["type"]
                ws.cell(row=idx, column=16).border = border_bold
                ws.cell(row=idx, column=17).value = product["amount"]
                ws.cell(row=idx, column=17).border = border_bold
                ws.cell(row=idx, column=18).value = product["success"]
                ws.cell(row=idx, column=18).border = border_bold
                ws.cell(row=idx, column=19).value = product["refund_status"]
                ws.cell(row=idx, column=19).border = border_bold

        # Auto-adjust column width based on content size
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        # Add thin and bold border to all cells
        thin_border_bold = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.border = thin_border_bold

        # Add medium bold border to the last row
        for cell in ws[ws.max_row]:
            cell.border = border_medium_bold


        # Generate file name with date and time
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f'customer_excel_report__{now}.xlsx'
            
        # Save the workbook in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Set response content type
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Set content disposition header to force download
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        
        return response
        # return Response({'success':1,'message':'Data Found','result':''})
    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def post_mis_report_for_customer_pdf(request):
    try:
        
        # Get today's date in the current timezone
        today_date = timezone.localdate()
        date_from_raw = request.data.get('from_date')
        date_to_raw = request.data.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success':0,'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Get all machine IDs from MachineMaster
        machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

        # Define time threshold (two minutes ago)
        threshold_time = timezone.now() - timedelta(minutes=2)

        # Fetch machine data along with their creation times using cursor
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload, created_at
                FROM userapp_mqttdata
                WHERE created_at >= %s
                """,
                [threshold_time]
            )
            rows = cursor.fetchall()

        # Prepare dictionaries to store machine statuses
        machine_status = {}
        machine_off = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"

        stocks = defaultdict(int)
        # Iterate through the fetched data to determine machine statuses
        for row in rows:
            payload = json.loads(row[0])
            M_Id = payload.get('M_Id')
            if M_Id in machine_ids:
                machine_status[M_Id] = "ON"
                machine_off.pop(M_Id, None)  # Remove machine from the OFF list if it's online
                stocks[M_Id] = payload.get('Stock')

        # Fetch latest stock data for each machine from MStatus model
        mstatus_data = MStatus.objects.filter(m_id__in=machine_ids)
        for mstatus in mstatus_data:
            stocks[mstatus.m_id] = mstatus.stock

        # Fetch machine-user mapping data within the specified date range
        report_data = MachineUserMapping.objects.filter(
            created_at__range=(date_from, date_to),
            assigned_customer__isnull=False
        ).order_by('-id')

        # Serialize the data
        serialized_data = []
        for item in report_data:
            serialized_data.append({
                'spoc_id': item.assigned_customer.spoc_id if item.assigned_customer else None,
                'name': item.assigned_customer.name if item.assigned_customer else None,
                'user_mobile_no': item.assigned_customer.mobile_no if item.assigned_customer else None,
                'email': item.assigned_customer.email if item.assigned_customer else None,
                'machine_id': item.machine.machine_id,
                'model_no': item.machine.model_number.model_no if item.machine.model_number else None,
                'location': item.machine.installation_location if item.machine else None,
                'product': item.machine.model_number.name if item.machine.model_number else None,
                'date': item.created_at.date().isoformat(),
                'time': item.created_at.time().isoformat(),
                'type': item.machine.product.product_type if item.machine.product else None,
                'amount': item.machine.product.amount if item.machine.product else None,
                'success': '',
                'refund_status': '',
                'machine_status': machine_status.get(item.machine.machine_id, "OFF"),  # Fetch machine status
                'current_inventory_count': stocks.get(item.machine.machine_id, None),  # Fetch last stock
                'address1': item.assigned_customer.address1 if item.assigned_customer else None,
                'address2': item.assigned_customer.address2 if item.assigned_customer else None,
                'country': item.assigned_customer.country if item.assigned_customer else None,
                'pincode': item.assigned_customer.pincode if item.assigned_customer else None,
                'is_active': item.is_active,
            })

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success':0,"message": "No data found for the provided date range."})

        return Response({'success':1,'message':'Data Found','result': serialized_data})
    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def post_mis_report_for_user_pdf(request):
    try:
        today_date = timezone.localdate()
        
        date_from_raw = request.data.get('from_date')
        date_to_raw = request.data.get('to_date')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()

        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        # Get all machine IDs from MachineMaster
        machine_ids = MachineMaster.objects.values_list('machine_id', flat=True)

        # Define time threshold (two minutes ago)
        threshold_time = timezone.now() - timedelta(minutes=2)

        # Fetch machine data along with their creation times using cursor
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload, created_at
                FROM userapp_mqttdata
                WHERE created_at >= %s
                """,
                [threshold_time]
            )
            rows = cursor.fetchall()

        # Prepare dictionaries to store machine statuses and stocks
        machine_status = {m_id: "OFF" for m_id in machine_ids}
        stocks = {m_id: None for m_id in machine_ids}

        # Iterate through the fetched data to determine machine statuses
        for row in rows:
            payload = json.loads(row[0])
            m_id = payload.get('M_Id')
            if m_id in machine_ids:
                machine_status[m_id] = "ON"

        # Fetch latest stock data for each machine from MStatus model
        mstatus_data = MStatus.objects.filter(m_id__in=machine_ids)
        for mstatus in mstatus_data:
            stocks[mstatus.m_id] = mstatus.stock

        # Fetch latest refill information for each machine
        refill_data = {}
        for machine_id in machine_ids:
            all_machine_ids = MStatus.objects.filter(
                m_id=machine_id,
                stock=F('capacity')  # Filter where stock capacity equals stock
            ).values('m_id', 'created_at__date', 'stock').annotate(
                count_capacity_equals_stock=Count('id')
            )

            machine_data = []

            for entry in all_machine_ids:
                stock_after_refill = entry['stock']

                stock_before_date_query = MStatus.objects.filter(
                    m_id=machine_id,
                    created_at__date__lt=entry['created_at__date']
                ).order_by('-created_at')

                stock_before_date = stock_before_date_query.values_list('stock', flat=True).first()

                # To ensure that the stock before date exists and it's not a refill that just happened
                if stock_before_date is not None and stock_before_date_query.count() > 1:
                    # Calculate refill quantity if stock_before_date exists
                    refill_quantity = stock_after_refill - stock_before_date

                    machine_data.append({
                        'machine_id': entry['m_id'],
                        'refilled_date': entry['created_at__date'],
                        'refilled_quantity': refill_quantity
                    })

            refill_data[machine_id] = machine_data
        
        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to), assigned_customer__isnull=True)
        serialized_data = []
        for item in report_data:
            machine_id = item.machine.machine_id
            serialized_item = {
                'user_id': item.assigned_user.id if item.assigned_user else None,
                'name': item.assigned_user.name if item.assigned_user else None,
                'user_mobile_no': item.assigned_user.mobile_no if item.assigned_user else None,
                'email': item.assigned_user.email if item.assigned_user else None,

                'machine_id': item.machine.machine_id,
                'model_no': item.machine.model_number.model_no if item.machine.model_number else None,
                'location': item.machine.installation_location if item.machine else None,

                'product': item.machine.model_number.name if item.machine.model_number else None,
                'type': item.machine.product.product_type if item.machine.product else None,
                'amount': item.machine.product.amount if item.machine.product else None,
                'machine_status': machine_status.get(item.machine.machine_id, "OFF"),
                'refilled_date': None,
                'refilled_quantity': None,
                'current_inventory_count': stocks.get(item.machine.machine_id, None),
                'success': '',
                'refund_status': '',
            }
            # Append separate serialized items for each refill record
            refill_items = refill_data.get(machine_id, [])
            for refill_item in refill_items:
                refill_serialized_item = serialized_item.copy()  # Create a copy of the base serialized item
                refill_serialized_item['refill_date'] = refill_item['refilled_date']
                refill_serialized_item['refill_quantity'] = refill_item['refilled_quantity']
                serialized_data.append(refill_serialized_item)

            # Append the base serialized item if no refill records found
            if not refill_items:
                serialized_data.append(serialized_item)

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success': 0, "message": "No data found for the provided date range."})

        return Response({'success': 1, 'message': 'Data Found', 'result': serialized_data})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mstatus_detail_for_tenant_customer(request):
    if request.method == 'GET':
        try:
            map_obj = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')
            machine_ids = [obj.machine.machine_id for obj in map_obj]  # Get machine IDs associated with the user
            

            # Define time threshold (two minutes ago)
            threshold_time = timezone.now() - timedelta(minutes=2)

            # Fetch machine data along with their creation times using cursor
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT payload, created_at
                    FROM userapp_mqttdata
                    WHERE created_at >= %s
                    """,
                    [threshold_time]
                )
                rows = cursor.fetchall()

            # Prepare dictionaries to store machine statuses
            machine_status = {}
            machine_off = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"

            # Iterate through the fetched data to determine machine statuses
            for row in rows:
                payload = json.loads(row[0])
                M_Id = payload.get('M_Id')
                if M_Id in machine_ids:
                    machine_status[M_Id] = "ON"
                    machine_off.pop(M_Id, None)  # Remove machine from the OFF list if it's online

            # Fetch latest stock data for each machine
            stocks = []
            for machine_id in machine_ids:
                stock = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()
                if stock:
                    color = 'green' if 16 <= stock.stock <= 25 else 'yellow' if 6 <= stock.stock <= 15 else 'red' if 1 <= stock.stock <= 5 else 'Refill'
                    machine_on_off = "ON" if machine_status.get(machine_id) else "OFF"
                    stocks.append({
                        'M_Id': stock.m_id,
                        'Capacity': stock.capacity,
                        'Stock': stock.stock,
                        'Status': color,
                        'Mode': stock.mode,
                        'Color': color,
                        'created_at': stock.created_at,
                        'created_by': stock.created_by,
                        'machine_status': machine_on_off,
                        'user_id': request.user.id  # Including the user ID in the response
                    })

            # Prepare the response
            response_data = {
                'success': 1,
                'message': 'Data Found',
                'results': stocks
            }

            return Response(response_data)

        except Exception as e:
            return Response({'success': 0, 'message': str(e)})
        

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mstatus_detail_for_tenant_customer_user(request):
    if request.method == 'GET':
        try:
            map_obj = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
            machine_ids = [obj.machine.machine_id for obj in map_obj]  # Get machine IDs associated with the user
            

            # Define time threshold (two minutes ago)
            threshold_time = timezone.now() - timedelta(minutes=2)

            # Fetch machine data along with their creation times using cursor
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT payload, created_at
                    FROM userapp_mqttdata
                    WHERE created_at >= %s
                    """,
                    [threshold_time]
                )
                rows = cursor.fetchall()

            # Prepare dictionaries to store machine statuses
            machine_status = {}
            machine_off = {machine_id: "OFF" for machine_id in machine_ids}  # Initialize all machines as "OFF"

            # Iterate through the fetched data to determine machine statuses
            for row in rows:
                payload = json.loads(row[0])
                M_Id = payload.get('M_Id')
                if M_Id in machine_ids:
                    machine_status[M_Id] = "ON"
                    machine_off.pop(M_Id, None)  # Remove machine from the OFF list if it's online

            # Fetch latest stock data for each machine
            stocks = []
            for machine_id in machine_ids:
                stock = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()
                if stock:
                    color = 'green' if 16 <= stock.stock <= 25 else 'yellow' if 6 <= stock.stock <= 15 else 'red' if 1 <= stock.stock <= 5 else 'Refill'
                    machine_on_off = "ON" if machine_status.get(machine_id) else "OFF"
                    stocks.append({
                        'M_Id': stock.m_id,
                        'Capacity': stock.capacity,
                        'Stock': stock.stock,
                        'Status': color,
                        'Mode': stock.mode,
                        'Color': color,
                        'created_at': stock.created_at,
                        'created_by': stock.created_by,
                        'machine_status': machine_on_off,
                        'user_id': request.user.id  # Including the user ID in the response
                    })

            # Prepare the response
            response_data = {
                'success': 1,
                'message': 'Data Found',
                'results': stocks
            }

            return Response(response_data)

        except Exception as e:
            return Response({'success': 0, 'message': str(e)})
        

from collections import defaultdict

    
from django.db.models.functions import TruncDate
# @api_view(['GET'])
# def get_stock_for_all_machine_refills(request):
#     try:
#         if request.method == 'GET':
#             # Fetch all distinct machine IDs
#             all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

#             machine_data = []

#             for machine_id in all_machine_ids:
#                 # Fetch the machine refill data grouped by day
#                 refill_data = MStatus.objects.filter(m_id=machine_id).annotate(
#                     refill_date=TruncDate('created_at')
#                 ).values('refill_date').annotate(
#                     stock_before_refill=F('stock'),  # Since we need the stock before refill
#                     stock_after_refill=Count('stock'),  # Count of stock after refill, we'll correct it later
#                     stock_capacity_equal_count=Count('stock', filter=F('stock') == F('capacity'))  # Count of stock equal to capacity
#                 )

#                 for refill_entry in refill_data:
#                     # Fetch the last entry for the current machine on the refill date
#                     last_entry_on_day = MStatus.objects.filter(
#                         m_id=machine_id,
#                         created_at__date=refill_entry['refill_date']
#                     ).order_by('-id').first()

#                     if last_entry_on_day:
#                         refill_entry['stock_after_refill'] = last_entry_on_day.stock

#                 # Append data for the current machine to the list
#                 machine_data.extend([
#                     {
#                         "machine_id": machine_id,
#                         "date": refill_entry['refill_date'],
#                         "stock_before_refill": refill_entry['stock_before_refill'],
#                         "stock_after_refill": refill_entry['stock_after_refill'],
#                         "stock_capacity_equal_count": refill_entry['stock_capacity_equal_count'],
#                     }
#                     for refill_entry in refill_data
#                 ])

#             return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})
#     except Exception as e:
#         return Response({'success': 0, 'message': str(e)})
# @api_view(['GET'])
# def get_stock_for_all_machine_refills(request):
#     try:
#         if request.method == 'GET':
#             machine_data = []

#             # Fetch all distinct machine IDs
#             all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

            # for machine_id in all_machine_ids:
            #     # Fetch the last entry for the current machine sorted by id
            #     last_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()

            #     if last_entry:
            #         # Fetch the second-last entry for the current machine sorted by id
            #         second_last_entry = MStatus.objects.filter(m_id=machine_id, id__lt=last_entry.id).order_by('-id').first()

            #         # If both last and second-last entries exist
            #         if second_last_entry:
            #             stock_before_refill = second_last_entry.stock
            #         else:
            #             stock_before_refill = None

            #         stock_after_refill = last_entry.stock

#                     # Count instances where stock is equal to capacity for the current machine
#                     stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

#                     # Calculate refill stock if stock and capacity are equal
#                     if stock_after_refill == last_entry.capacity:
#                         refill_stock = stock_after_refill - stock_before_refill
#                     else:
#                         refill_stock = None

#                     # Extract month and day from the date
#                     # month = last_entry.created_at.month
#                     # day = last_entry.created_at.day

#                     # Calculate per day and per month counts
#                     # per_day_count = MStatus.objects.filter(m_id=machine_id, created_at__day=day).count()
#                     # per_month_count = MStatus.objects.filter(m_id=machine_id, created_at__month=month).count()

#                     machine_data.append({
#                         "machine_id": machine_id,
#                         "date": last_entry.created_at.date().isoformat(),
#                         # "month": month,
#                         # "day": day,
#                         "stock_before_refill": stock_before_refill,
#                         "stock_after_refill": stock_after_refill,
#                         "stock_capacity_equal_count": stock_capacity_equal_count,
#                         "refill_stock": refill_stock,
#                         # "per_day_count": per_day_count,
#                         # "per_month_count": per_month_count,
#                     })
#                 else:
#                     # If no entry exists for the machine, set default values
#                     machine_data.append({
#                         "machine_id": machine_id,
#                         "date": None,
#                         # "month": None,
#                         # "day": None,
#                         "stock_before_refill": None,
#                         "stock_after_refill": None,
#                         "stock_capacity_equal_count": 0,  # No entries, so count is 0
#                         "refill_stock": None,
#                         # "per_day_count": 0,
#                         # "per_month_count": 0,
#                     })

#             return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})
#     except Exception as e:
#         return Response({'success': 0, 'message': str(e)})

# @api_view(['GET'])
# def get_stock_for_all_machine_refills(request):
#     try:
#         if request.method == 'GET':
#             machine_data = []

#             # Fetch all distinct machine IDs
#             all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

#             for machine_id in all_machine_ids:
#                 # Fetch all entries for the current machine sorted by id in descending order
#                 entries = MStatus.objects.filter(m_id=machine_id).order_by('-id')

#                 for idx, entry in enumerate(entries):
#                     # If it's not the last entry, calculate refill counts
#                     if idx < len(entries) - 1:
#                         next_entry = entries[idx + 1]
#                         stock_before_refill = entry.stock
#                         stock_after_refill = next_entry.stock
#                         if stock_after_refill == next_entry.capacity:
#                             per_day_refill_count = 1
#                         else:
#                             per_day_refill_count = 0
#                     else:
#                         stock_before_refill = None
#                         stock_after_refill = entry.stock
#                         per_day_refill_count = 0

#                     # Extract month and day from the date
#                     month = entry.created_at.month
#                     day = entry.created_at.day

#                     # Count instances where stock is equal to capacity for the current machine
#                     stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

#                     # Calculate per day and per month counts
#                     per_day_count = MStatus.objects.filter(m_id=machine_id, created_at__day=day).count()
#                     per_month_count = MStatus.objects.filter(m_id=machine_id, created_at__month=month).count()

#                     machine_data.append({
#                         "machine_id": machine_id,
#                         "date": entry.created_at.date().isoformat(),
#                         "month": month,
#                         "day": day,
#                         "stock_before_refill": stock_before_refill,
#                         "stock_after_refill": stock_after_refill,
#                         "stock_capacity_equal_count": stock_capacity_equal_count,
#                         "refill_stock": stock_after_refill - stock_before_refill if stock_before_refill is not None else None,
#                         "per_day_count": per_day_count,
#                         "per_month_count": per_month_count,
#                         "per_day_refill_count": per_day_refill_count,
#                     })

#             return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})
#     except Exception as e:
#         return Response({'success': 0, 'message': str(e)})
# @api_view(['GET'])
# def get_stock_for_all_machine_refills(request):
#     try:
#         if request.method == 'GET':
#             # Fetch all distinct machine IDs
#             all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

#             # Get current date
#             current_date = datetime.now().date()

#             machine_data = []

#             for machine_id in all_machine_ids:
#                 # Fetch stock data for the current machine, grouped by date of creation
#                 stock_data = MStatus.objects.filter(m_id=machine_id).annotate(
#                     creation_date=TruncDate('created_at')
#                 ).values('creation_date').annotate(
#                     stock_before_refill=F('stock'),
#                     stock_after_refill=F('stock'),
#                     stock_capacity_equal_count=Count('id')
#                 ).order_by('-creation_date')

#                 # Append data for the current machine to the list
#                 for stock_entry in stock_data:
#                     machine_data.append({
#                         "machine_id": machine_id,
#                         "date": stock_entry['creation_date'],
#                         "stock_before_refill": stock_entry['stock_before_refill'],
#                         "stock_after_refill": stock_entry['stock_after_refill'],
#                         "stock_capacity_equal_count": stock_entry['stock_capacity_equal_count'],
#                     })

#             return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})
#     except Exception as e:
#         return Response({'success': 0, 'message': str(e)})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_refill_counts_per_machine(request):
    try:
        if request.method == 'GET':
            # Get all distinct machine IDs
            all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

            # Initialize a dictionary to store counts per machine
            counts_per_machine = {}

            # Get today's date
            today = timezone.now().date()

            # Get counts per day and per month for each machine
            for machine_id in all_machine_ids:
                # Per day count
                per_day_count = RefillStockHistory.objects.filter(machine_id_store=machine_id, refill_date__date=today).count()

                # Per month count
                per_month_count = RefillStockHistory.objects.filter(machine_id_store=machine_id, refill_date__year=today.year, refill_date__month=today.month).count()

                # Store counts for the current machine
                counts_per_machine[machine_id] = {
                    'per_day_count': per_day_count,
                    'per_month_count': per_month_count
                }

            # Construct the response data
            response_data = {
                'success': 1,
                'message': 'Counts Retrieved Successfully',
                'counts_per_machine': counts_per_machine
            }

            return Response(response_data)

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_refill_counts(request):
    try:
        # Get today's date
        today = timezone.now().date()

        # Calculate per day count
        per_day_count = RefillStockHistory.objects.filter(refill_date__date=today).count()

        # Calculate total count
        total_count = RefillStockHistory.objects.count()

        # Response data
        response_data = {
            'per_day_count': per_day_count,
            'total_count': total_count
        }

        # Return response
        return Response({'success': 1, 'message': 'Data Found', 'result': response_data})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})
    
from django.db.models import F, Count

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stock_capacity_equal_count(request):
    try:
        if request.method == 'GET':
            # Aggregate count of instances where stock is equal to capacity, grouped by machine_id
            stock_capacity_equal_counts = MStatus.objects.filter(stock=F('capacity')).values('m_id').annotate(count=Count('m_id'))

            # Prepare the result dictionary
            result = {}
            for entry in stock_capacity_equal_counts:
                machine_id = entry['m_id']
                count = entry['count']
                result[machine_id] = count

            return Response({'success': 1, 'message': 'Data Found', 'result': result})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})


from django.db.models import Q
@api_view(['GET'])
@permission_classes([IsAuthenticated]) 
def get_map_customer_to_user_report(request):
    if request.method=='GET':
        try:
        
            # Filter report data based on the logged-in customer user's ID
            report_data = MachineUserMapping.objects.filter(assigned_customer=request.user.id)

            serialized_data = []
            for item in report_data:
                # Only include mapped machines where assigned_user is not None
                if item.assigned_user:
                    serialized_item = {
                        'machine_id': item.machine.machine_id if item.machine else None,
                        'user_id': item.assigned_user.id,
                        'name': item.assigned_user.name,
                        'user_mobile_no': item.assigned_user.mobile_no,
                        'email': item.assigned_user.email,
                        'assigned_machine_date':item.assigned_user_date
                    }
                    serialized_data.append(serialized_item)


            return Response({'success': 1, 'message': 'Data Found', 'result': serialized_data})

        except Exception as e:
            return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_user_payment_api_view(request):
    try:
        
        # Filter MachineUserMapping based on the logged-in customer user's ID
        assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)

        # Retrieve machine IDs assigned to the user
        machine_ids = assigned_customers.values_list('machine__machine_id', flat=True)
        
        # Retrieve MachineMaster objects with matching machine IDs
        machine_master_objects = MachineMaster.objects.filter(machine_id__in=machine_ids)

        response_data = {
            'success': 1,
            'message': 'Data Found',
            'result': []
        }

        for machine_master in machine_master_objects:
            # Retrieve MStatus objects for the current machine
            mstatus_objects = MStatus.objects.filter(m_id=machine_master.machine_id)
            
            # Count occurrences of 'COIN' and 'QR' modes for the current machine
            coin_mode_counts = mstatus_objects.filter(mode='COIN').count()
            qr_mode_counts = mstatus_objects.filter(mode='QR').count()

            total_mode_counts = coin_mode_counts + qr_mode_counts

            # Calculate percentages
            coin_mode_percentage = (coin_mode_counts / total_mode_counts) * 100 if total_mode_counts else 0
            qr_mode_percentage = (qr_mode_counts / total_mode_counts) * 100 if total_mode_counts else 0

            response_data['result'].append({
                'machine_id': machine_master.machine_id,
                'coin_mode_counts': coin_mode_counts,
                'qr_mode_counts': qr_mode_counts,
                'coin_mode_percentage': coin_mode_percentage,
                'qr_mode_percentage': qr_mode_percentage,
                'total_mode_counts': total_mode_counts
            })

        return Response(response_data, status=status.HTTP_200_OK)
    except MachineUserMapping.DoesNotExist:
        return Response({'success': 0, 'message': 'Machine User Mapping does not exist for the current user.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_customer_role_for_user_download_excel_report(request,user_id):
    try:
        # Get the logged-in user
        logged_in_user = user_id
        
        # Get the from_date and to_date from the request query parameters
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        # Parse the string dates into datetime objects
        from_date = timezone.make_aware(datetime.strptime(from_date_str, '%Y-%m-%d')) if from_date_str else None
        to_date = timezone.make_aware(datetime.strptime(to_date_str, '%Y-%m-%d')) if to_date_str else None

        # Validate that both from_date and to_date are provided
        if not from_date or not to_date:
            return Response({'success': False, "message": "Both 'from_date' and 'to_date' are required."}, status=400)

        # Validate that from_date is before to_date
        if from_date >= to_date:
            return Response({'success': False, "message": "'from_date' must be before 'to_date'."}, status=400)

        # Query machine-customer mappings for the logged-in user within the specified date range
        mappings = MachineUserMapping.objects.filter(assigned_customer=logged_in_user, created_at__range=[from_date, to_date])
        
        # Extract machine IDs from mappings
        machine_ids = mappings.values_list('machine_id', flat=True)

        # Query machines based on extracted machine IDs
        machines = MachineMaster.objects.filter(id__in=machine_ids)

        serialized_data = []

        for machine in machines:

            assigned_user_data = MachineUserMapping.objects.filter(machine=machine, assigned_customer=logged_in_user).first().assigned_user
        

            report_date = datetime.now().strftime("%Y-%m-%d")
            report_time = datetime.now().strftime("%H:%M:%S")

            serialized_item = {
                'report_date': report_date,
                'report_time': report_time,
                'tendryl': {
                    'user_id': assigned_user_data.id if assigned_user_data else None,
                    'name': assigned_user_data.name if assigned_user_data else None,
                    'user_moblie_no': assigned_user_data.mobile_no if assigned_user_data else None,
                    'email': assigned_user_data.email if assigned_user_data else None,
                },
                'machine': {
                    'machine_id': machine.machine_id,
                    'model_no': machine.model_number.model_no if machine.model_number else None,
                    'location': machine.installation_location,
                    'current_status_on_off': '',  # Fill this with actual data
                    'current_inventory_count': '',  # Fill this with actual data
                },
                'product': {
                    'product_type': machine.product.product_type if machine.product else None,
                    'recent_refill_date': '',  # Fill this with actual data
                    'quantity_refilled': '',  # Fill this with actual data
                    'recent_case_collection_date': '',  # Fill this with actual data
                    'recent_amount_collected': '',  # Fill this with actual data
                },
            }

            serialized_data.append(serialized_item)


        wb = Workbook()
        ws = wb.active

        # Add header row
        ws.append([
            "No", "Report Date", "Report Time", 
            "TENDRYL", "", "", "", 
            "MACHINE", "", "", "", 
            "PRODUCT", "", "", "", "",
        ])

        # Set background colors for specific columns
        customer_fill = PatternFill(start_color="43E443", end_color="43E443", fill_type="solid")
        machine_fill = PatternFill(start_color="f8356f", end_color="f8356f", fill_type="solid")
        product_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Apply colors to the "CUSTOMER", "MACHINE", and "PRODUCT" sections
        for col in range(1, 17):  # Loop through relevant columns
            if 4 <= col <= 7:
                ws.cell(row=1, column=col).fill = customer_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make CUSTOMER header bold
            elif 8 <= col <= 12:
                ws.cell(row=1, column=col).fill = machine_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make MACHINE header bold
            elif 13 <= col <= 17:
                ws.cell(row=1, column=col).fill = product_fill
                ws.cell(row=1, column=col).font = Font(bold=True)  # Make PRODUCT header bold

        # Add borders to the cells
        border_bold = Border(left=Side(style='medium'), 
                            right=Side(style='medium'), 
                            top=Side(style='medium'), 
                            bottom=Side(style='medium'))

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=17):
            for cell in row:
                cell.border = border_bold


        # Add values for "Report" and "Date" in the first set of merged cell
        # Merge cells A1:A2
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Merge cells B1:B2
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Merge cells C1:C2
        ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                

        # Merge cells for "CUSTOMER"
        ws.cell(row=1, column=4).value = "TENDRYL"
        ws.cell(row=1, column=4).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)  # CUSTOMER

        # Add details for "CUSTOMER" section in the second row
        ws.cell(row=2, column=4).value = "User ID"
        ws.cell(row=2, column=4).font = Font(bold=True)  # Make SPOC ID bold
        ws.cell(row=2, column=4).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=5).value = "Name"
        ws.cell(row=2, column=5).font = Font(bold=True)  # Make Name bold
        ws.cell(row=2, column=5).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=6).value = "Mobile"
        ws.cell(row=2, column=6).font = Font(bold=True)  # Make Mobile bold
        ws.cell(row=2, column=6).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=7).value = "Email"
        ws.cell(row=2, column=7).font = Font(bold=True)  # Make Email bold
        ws.cell(row=2, column=7).alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells for "MACHINE"
        ws.cell(row=1, column=8).value = "MACHINE"
        ws.cell(row=1, column=8).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)  # MACHINE

        # Add details for "MACHINE" section in the second row
        ws.cell(row=2, column=8).value = "Machine ID"
        ws.cell(row=2, column=8).font = Font(bold=True)  # Make Machine ID bold
        ws.cell(row=2, column=8).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=9).value = "Model No"
        ws.cell(row=2, column=9).font = Font(bold=True)  # Make Model No bold
        ws.cell(row=2, column=9).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=10).value = "Location"
        ws.cell(row=2, column=10).font = Font(bold=True)  # Make Location bold
        ws.cell(row=2, column=10).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=11).value = "Current Status (Off / On)"
        ws.cell(row=2, column=11).font = Font(bold=True)  # Make Current Status (Off / On) bold
        ws.cell(row=2, column=11).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=12).value = "Current Inventory Count"
        ws.cell(row=2, column=12).font = Font(bold=True)  # Make Current Status (Off / On) bold
        ws.cell(row=2, column=12).alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells for "PRODUCT"
        ws.cell(row=1, column=13).value = "PRODUCT"
        ws.cell(row=1, column=13).alignment = Alignment(horizontal='center', vertical='center')  # Align CUSTOMER header
        ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=17)  # PRODUCT

        # Add details for "PRODUCT" section in the second row
        ws.cell(row=2, column=13).value = "Product Type"
        ws.cell(row=2, column=13).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=13).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=14).value = "Recent Refill Date"
        ws.cell(row=2, column=14).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=14).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=15).value = "Quantity Refilled"
        ws.cell(row=2, column=15).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=15).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=16).value = "Recent Cash Collection Date"
        ws.cell(row=2, column=16).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=16).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=17).value = "Recent Amount Collected"
        ws.cell(row=2, column=17).font = Font(bold=True)  # Make Current Inventory Count bold
        ws.cell(row=2, column=17).alignment = Alignment(horizontal='center', vertical='center')

        # ws.cell(row=2, column=18).value = "Success (Y/N)"
        # ws.cell(row=2, column=18).font = Font(bold=True)  # Make Current Inventory Count bold
        # ws.cell(row=2, column=18).alignment = Alignment(horizontal='center', vertical='center')

        # ws.cell(row=2, column=19).value = "Refund Status"
        # ws.cell(row=2, column=19).font = Font(bold=True)  # Make Current Inventory Count bold
        # ws.cell(row=2, column=19).alignment = Alignment(horizontal='center', vertical='center')

        # Set height for row 2
        ws.row_dimensions[2].height = points_to_pixels(28)
        
        # Add report data
        for idx, data in enumerate(serialized_data, start=3):
            
            ws.cell(row=idx, column=1).value = idx - 2  # Serial number
            ws.cell(row=idx, column=1).border = border_bold
            ws.cell(row=idx, column=2).value = data["report_date"]
            ws.cell(row=idx, column=2).border = border_bold
            ws.cell(row=idx, column=3).value = data["report_time"]
            ws.cell(row=idx, column=3).border = border_bold

            # Customer details
            tendryl = data["tendryl"]
            ws.cell(row=idx, column=4).value = tendryl["user_id"]
            ws.cell(row=idx, column=4).border = border_bold
            ws.cell(row=idx, column=5).value = tendryl["name"]
            ws.cell(row=idx, column=5).border = border_bold
            ws.cell(row=idx, column=6).value = tendryl["user_moblie_no"]
            ws.cell(row=idx, column=6).border = border_bold
            ws.cell(row=idx, column=7).value = tendryl["email"]
            ws.cell(row=idx, column=7).border = border_bold

            # Machine details
            machine = data["machine"]
            ws.cell(row=idx, column=8).value = machine["machine_id"]
            ws.cell(row=idx, column=8).border = border_bold
            ws.cell(row=idx, column=9).value = machine["model_no"]
            ws.cell(row=idx, column=9).border = border_bold
            ws.cell(row=idx, column=10).value = machine["location"]
            ws.cell(row=idx, column=10).border = border_bold
            ws.cell(row=idx, column=11).value = machine["current_status_on_off"]
            ws.cell(row=idx, column=11).border = border_bold
            ws.cell(row=idx, column=12).value = machine["current_inventory_count"]
            ws.cell(row=idx, column=12).border = border_bold

            # Product details
            product = data["product"]
            ws.cell(row=idx, column=13).value = product["product_type"]
            ws.cell(row=idx, column=13).border = border_bold
            ws.cell(row=idx, column=14).value = product["recent_refill_date"]
            ws.cell(row=idx, column=14).border = border_bold
            ws.cell(row=idx, column=15).value = product["quantity_refilled"]
            ws.cell(row=idx, column=15).border = border_bold
            ws.cell(row=idx, column=16).value = product["recent_case_collection_date"]
            ws.cell(row=idx, column=16).border = border_bold
            ws.cell(row=idx, column=17).value = product["recent_amount_collected"]
            ws.cell(row=idx, column=17).border = border_bold
            # ws.cell(row=idx, column=18).value = product["success"]
            # ws.cell(row=idx, column=19).value = product["refund_status"]

        # Generate file name with date and time
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f'user_excel_report__{now}.xlsx'

        # Save workbook to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        return response

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})
    
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch

@api_view(['GET'])
def get_customer_role_for_user_download_pdf_report(request):
    # try:
        # Get the logged-in user
        logged_in_user = request.user

        # Get the from_date and to_date from the request query parameters
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        # Parse the string dates into datetime objects
        from_date = timezone.make_aware(datetime.strptime(from_date_str, '%Y-%m-%d')) if from_date_str else None
        to_date = timezone.make_aware(datetime.strptime(to_date_str, '%Y-%m-%d')) if to_date_str else None

        # Validate that both from_date and to_date are provided
        if not from_date or not to_date:
            return Response({'success': False, "message": "Both 'from_date' and 'to_date' are required."}, status=400)

        # Validate that from_date is before to_date
        if from_date >= to_date:
            return Response({'success': False, "message": "'from_date' must be before 'to_date'."}, status=400)

        # Query machine-customer mappings for the logged-in user within the specified date range
        mappings = MachineUserMapping.objects.filter(assigned_customer=logged_in_user, created_at__range=[from_date, to_date])

        # Extract unique dates from the mappings
        unique_dates = mappings.values_list('created_at__date', flat=True).distinct()

        # Create a list to hold the data for the PDF report
        pdf_data_pages = []

        # Iterate over unique dates
        for date in unique_dates:
            # Filter mappings for the current date
            mappings_for_date = mappings.filter(created_at__date=date)

            # Extract machine IDs from mappings
            machine_ids = mappings_for_date.values_list('machine_id', flat=True)

            # Query machines based on extracted machine IDs
            machines = MachineMaster.objects.filter(id__in=machine_ids)

            # Create PDF data for the current date
            pdf_data = []

            # Add headers to the PDF data
            headers = ["Field", "Value"]
            pdf_data.append(headers)

            # Add data rows to the PDF data
            for machine in machines:
                assigned_user_data = MachineUserMapping.objects.filter(machine=machine, assigned_customer=logged_in_user).first().assigned_user
                
                pdf_data.extend([
                    ["Date", date.strftime('%Y-%m-%d')],
                    ["TENDRYL User ID", str(assigned_user_data.id) if assigned_user_data else None],
                    ["Name", assigned_user_data.name if assigned_user_data else None],
                    ["Mobile", assigned_user_data.mobile_no if assigned_user_data else None],
                    ["Email", assigned_user_data.email if assigned_user_data else None],
                    ["Machine ID", str(machine.machine_id)],
                    ["Model No", machine.model_number.model_no if machine else None],
                    ["Location", machine.installation_location],
                    ["Current Status (Off / On)"],# machine.current_status_on_off],
                    ["Current Inventory Count"], #, machine.current_inventory_count],
                    ["Product Type", machine.product.product_type if machine.product else None],
                    ["Recent Refill Date"], #machine.product.recent_refill_date if machine.product else None],
                    ["Quantity Refilled"], #machine.product.quantity_refilled if machine.product else None],
                    ["Recent Case Collection Date"], #machine.product.recent_case_collection_date if machine.product else None],
                    ["Recent Amount Collected"], #machine.product.recent_amount_collected if machine.product else None],
                    ["", ""]  # Add an empty row for spacing
                ])

            # Append PDF data for the current date to the list of pages
            pdf_data_pages.append(pdf_data)

        # Generate file name with date and time
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f'user_pdf_report__{now}.pdf'

        # Create PDF document
        pdf = SimpleDocTemplate(file_name, pagesize=letter)

        # Create tables for each page
        for page_data in pdf_data_pages:
            # Create table from PDF data
            table = Table(page_data)

            # Add style to the table
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            table.setStyle(style)

            # Add table to PDF document
            pdf.build([table, PageBreak()])

        # Create response
        with open(file_name, 'rb') as pdf_file:
            response = HttpResponse(pdf_file.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        return response

    # except Exception as e:
    #     return Response({'success': 0, 'message': str(e)})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def post_mis_customer_role_for_user_excel(request):
    try:
        date_from_raw = request.data.get('from')
        date_to_raw = request.data.get('to')

        # Validate input data
        if not date_from_raw or not date_to_raw:
            return Response({'success':0,'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date strings into datetime objects using dateutil.parser
        date_from = parser.parse(date_from_raw)
        date_to = parser.parse(date_to_raw)

        report_data = MachineUserMapping.objects.filter(created_at__range=(date_from, date_to),assigned_customer=request.user)
        serialized_data = [{
            'name': item.assigned_customer.name if item.assigned_customer else None,
            'user_moblie_no': item.assigned_customer.mobile_no if item.assigned_customer else None,
            'email': item.assigned_customer.email if item.assigned_customer else None,

            'machine_id': item.machine.machine_id,
            'model_no': item.machine.model_number  if item.assigned_customer else None,
            'location': item.machine.installation_location  if item.assigned_customer else None,
            # 'current_status': item.machine   if item.assigned_customer else None,
            # 'current_inventory_count': item.machine.current_inventory_count,

            'product': item.machine.model_number.name if item.machine.model_number else None,  
            'date': item.created_at.date().isoformat(),  
            'time': item.created_at.time().isoformat(),  
            'type': item.machine.product.product_type if item.machine.product else None, 
            'amount': item.machine.product.amount if item.machine.product else None,  
            'success': '',  
            'refund_status': '',  

            'address1': item.assigned_customer.address1 if item.assigned_customer else None,
            'address2': item.assigned_customer.address2 if item.assigned_customer else None,
            'country': item.assigned_customer.country if item.assigned_customer else None,
            'pincode': item.assigned_customer.pincode if item.assigned_customer else None,
            'is_active': item.is_active,
        } for item in report_data]

        # Check if serialized_data is empty
        if not serialized_data:
            return Response({'success':0,"message": "No data found for the provided date range."})
        
        return JsonResponse({'success':1,'message':'Data Found','result': serialized_data})
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def post_stock_for_all_machine_refills_excel(request):
    try:
        if request.method == 'POST':
            # Get the from_date and to_date from the request data
            from_date = request.data.get('from_date')
            to_date = request.data.get('to_date')

            # Parse the dates into datetime objects
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

            machine_data = []

            # Fetch all distinct machine IDs within the date range
            all_machine_ids = MStatus.objects.filter(created_at__date__range=[from_date, to_date]).values_list('m_id', flat=True).distinct()

            for machine_id in all_machine_ids:
                # Fetch the last entry for the current machine sorted by id
                last_entry = MStatus.objects.filter(m_id=machine_id, created_at__date__range=[from_date, to_date]).order_by('-id').first()

                if last_entry:
                    
                    stock_before_refill = last_entry.stock
                    # Count instances where stock is equal to capacity for the current machine per day
                    # stock_capacity_equal_count_per_day = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).annotate(
                    #     day=Trunc('created_at', 'day', output_field=DateField())
                    # ).values('day').annotate(count=Count('id')).order_by('day')
                    stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()
                    

                    stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
                    if stock_after_refill_qs.exists():
                        stock_after_refill = stock_after_refill_qs.first().stock
                    else:
                        stock_after_refill = None
                        

                    # Fetch the last second entry for the current machine sorted by id
                    last_second_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id')[1:2].first()
                    if last_second_entry:
                        last_second_stock = last_second_entry.stock
                    else:
                        last_second_stock = None


                    if stock_after_refill is not None:
                        refill_quantity = stock_after_refill - last_second_stock
                    else:
                        refill_quantity = None

                    # Fetch refill date from MStatus model
                    refill_date = last_entry.created_at.date() if last_entry.created_at else None

                    # Fetch location from MachineMaster model
                    machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
                    location = machine_master_data.installation_location if machine_master_data else None


                    # Add data for the current machine to the list
                    machine_data.append({
                        
                        "machine_id": machine_id,
                        "stock_before_refill": last_second_stock,
                        "refill_quantity": refill_quantity,
                        "stock_after_refill": stock_after_refill,
                        "stock_capacity_equal_count": stock_capacity_equal_count,
                        'location':location,
                        "refill_date": refill_date,
                    })
                    if stock_after_refill == last_entry.capacity:
                        machine_data.append({
                            "machine_id": machine_id,
                            "stock_before_refill": last_entry.capacity,
                            "refill_quantity": 0,
                            "stock_after_refill": 0,
                            "stock_capacity_equal_count": stock_capacity_equal_count,
                            'location':location,
                            "refill_date": refill_date,
                        })
                else:
                    # If no entry exists for the machine, set both stock values to None
                    machine_data.append({
                        
                        "machine_id": machine_id,
                        "stock_before_refill": None,
                        "refill_quantity": None,
                        "stock_after_refill": None,
                        "stock_capacity_equal_count": None , # No entries, so count is 0
                        'location':location,
                        'refill_date':None
                    })

            return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stock(request):
    try:
        if request.method == 'GET':
            # Fetch all distinct machine IDs
            all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

            machine_data = []

            for machine_id in all_machine_ids:
                # Fetch the last entry for the current machine sorted by id
                last_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()

                # Get the stock from the last entry if available, otherwise set it to None
                stock = last_entry.stock if last_entry else None

        

                machine_data.append({
                    'machine_id': machine_id,
                    'stock': stock,
                })

            return JsonResponse({'success': 1, 'message': 'Data Found', 'result': machine_data})
    except Exception as e:
        return JsonResponse({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from django.contrib.auth import authenticate

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password_logged_in(request):
    try:
        # Get the current user
        user = request.user

        # Retrieve the current password and new password from the request data
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')

        # Check if the new password is the same as the current password
        if current_password == new_password:
            return Response({'success': 0, 'message': 'Please enter a different new password.'}, status=status.HTTP_400_BAD_REQUEST)


        # Check if the current password is correct
        if not authenticate(username=user, password=current_password):
            return Response({'success': 0, 'message': 'Current password is incorrect.'}, status=status.HTTP_400_BAD_REQUEST)

        # Change the password
        user.set_password(new_password)
        user.save()

        return Response({'success': 1, 'message': 'Password changed successfully.'}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stock_by_machine_customer_role_refills(request):
    try:
        if request.method == 'GET':
            # Fetch all distinct machine IDs associated with the logged-in user
            assigned_machines = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')
            
            
            all_machine_ids = [mapping.machine.machine_id for mapping in assigned_machines]

            machine_data = []

            for machine_id in all_machine_ids:
                
                # Fetch the last entry for the current machine sorted by id
                last_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()

                if last_entry:
                    stock_before_refill = last_entry.stock
                    # Count instances where stock is equal to capacity for the current machine per day
                    # stock_capacity_equal_count_per_day = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).annotate(
                    #     day=Trunc('created_at', 'day', output_field=DateField())
                    # ).values('day').annotate(count=Count('id')).order_by('day')
                    stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

                    stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
                    if stock_after_refill_qs.exists():
                        stock_after_refill = stock_after_refill_qs.first().stock
                    else:
                        stock_after_refill = None
                    

                    if stock_after_refill is not None:
                        refill_quantity = stock_after_refill - stock_before_refill
                    else:
                        refill_quantity = None
                    
                    # Fetch refill date from MStatus model
                    refill_date = last_entry.created_at.date() if last_entry.created_at else None

                    # Fetch location from MachineMaster model
                    machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
                    location = machine_master_data.installation_location if machine_master_data else None


                    # Add data for the current machine to the list
                    machine_data.append({
                        "machine_id": machine_id,
                        "stock_before_refill": stock_before_refill,
                        "stock_after_refill": stock_after_refill,
                        "refill_quantity": refill_quantity,
                        "stock_capacity_equal_count": stock_capacity_equal_count,
                        'location':location,
                        "refill_date": refill_date,
                    })
                else:
                    # If no entry exists for the machine, set both stock values to None
                    machine_data.append({
                        "machine_id": machine_id,
                        "stock_before_refill": 0,
                        "stock_after_refill": 0,
                        "refill_quantity": 0,
                        "stock_capacity_equal_count":0,  # No entries, so count is 0
                        'location':None,
                        "refill_date": '',
                    })

            return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stock_by_machine_user_role_refills(request):
    try:
        if request.method == 'GET':
            # Fetch all distinct machine IDs associated with the logged-in user
            assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
            
            
            all_machine_ids = [mapping.machine.machine_id for mapping in assigned_machines]

            machine_data = []

            for machine_id in all_machine_ids:
                # Fetch the last entry for the current machine sorted by id
                last_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()

                if last_entry:
                    stock_before_refill = last_entry.stock

                    stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

                    stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
                    if stock_after_refill_qs.exists():
                        stock_after_refill = stock_after_refill_qs.first().stock
                    else:
                        stock_after_refill = None
                    

                    if stock_after_refill is not None:
                        refill_quantity = stock_after_refill - stock_before_refill
                    else:
                        refill_quantity = None

                    # Fetch refill date from MStatus model
                    refill_date = last_entry.created_at.date() if last_entry.created_at else None

                    # Fetch location from MachineMaster model
                    machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
                    location = machine_master_data.installation_location if machine_master_data else None

                    # Add data for the current machine to the list
                    machine_data.append({
                        "machine_id": machine_id,
                        "stock_before_refill": stock_before_refill,
                        "refill_quantity": refill_quantity,
                        "stock_after_refill": stock_after_refill,
                        "stock_capacity_equal_count": stock_capacity_equal_count,
                        'location':location,
                        "refill_date": refill_date,
                    })
                else:
                    # If no entry exists for the machine, set both stock values to None
                    machine_data.append({
                        "machine_id": machine_id,
                        "stock_before_refill": 0,
                        "refill_quantity": 0,
                        "stock_after_refill": 0,
                        "stock_capacity_equal_count": 0,  # No entries, so count is 0
                        'location':None,
                        "refill_date": '',
                    })

            return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_payment_download(request):
    try:
        
        # Filter MachineUserMapping based on the logged-in customer user's ID
        assigned_customers = MachineUserMapping.objects.filter(assigned_customer=request.user.id)

        # Retrieve machine IDs assigned to the user
        machine_ids = assigned_customers.values_list('machine__machine_id', flat=True)
        
        # Retrieve MachineMaster objects with matching machine IDs
        machine_master_objects = MachineMaster.objects.filter(machine_id__in=machine_ids)

        response_data = {
            'success': 1,
            'message': 'Data Found',
            'result': []
        }

        for machine_master in machine_master_objects:
            # Retrieve MStatus objects for the current machine
            mstatus_objects = MStatus.objects.filter(m_id=machine_master.machine_id)
            
            # Count occurrences of 'COIN' and 'QR' modes for the current machine
            coin_mode_counts = mstatus_objects.filter(mode='COIN').count()
            qr_mode_counts = mstatus_objects.filter(mode='QR').count()

            total_mode_counts = coin_mode_counts + qr_mode_counts

            # Calculate percentages
            coin_mode_percentage = (coin_mode_counts / total_mode_counts) * 100 if total_mode_counts else 0
            qr_mode_percentage = (qr_mode_counts / total_mode_counts) * 100 if total_mode_counts else 0

            response_data['result'].append({
                'machine_id': machine_master.machine_id,
                'coin_mode_counts': coin_mode_counts,
                'qr_mode_counts': qr_mode_counts,
                'coin_mode_percentage': coin_mode_percentage,
                'qr_mode_percentage': qr_mode_percentage,
                'total_mode_counts': total_mode_counts
            })

        return Response(response_data, status=status.HTTP_200_OK)
    except MachineUserMapping.DoesNotExist:
        return Response({'success': 0, 'message': 'Machine User Mapping does not exist for the current user.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_payment_report_excel_download(request):
    try:
        tenant=request.tenant.user_id
        date_from_raw = request.GET.get('from_date')
        date_to_raw = request.GET.get('to_date')

        if not date_from_raw or not date_to_raw:
            return Response({'success': 0, 'message': 'Both from and to dates are required.'}, status=status.HTTP_400_BAD_REQUEST)

    
        date_from = parser.parse(date_from_raw).date()
        date_to = parser.parse(date_to_raw).date()
        date_to = date_to + timezone.timedelta(days=1) - timezone.timedelta(seconds=1)

        mstatus_records = MStatus.objects.filter(created_at__range=(date_from, date_to))

        mode_counts_per_machine = defaultdict(lambda: defaultdict(int))

        for record in mstatus_records:
            if record.m_id and record.mode:
                mode_counts_per_machine[record.m_id][record.mode] += 1

        data_for_response = []
        for m_id, modes in mode_counts_per_machine.items():
            row = {'m_id': m_id}
            total_modes = sum(modes.values())

            row['coin_mode_counts'] = modes.get('COIN', 0)
            row['qr_mode_counts'] = modes.get('QR', 0)

            row['coin_mode_percentage'] = (row['coin_mode_counts'] / total_modes) * 100 if total_modes else 0
            row['qr_mode_percentage'] = (row['qr_mode_counts'] / total_modes) * 100 if total_modes else 0

            data_for_response.append(row)

        # Fetch organization name from the database
        organization_name = None
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT organization_name
                FROM saasapp_companyregistration
                WHERE user_id = %s
            """, [tenant])
            row = cursor.fetchone()
            organization_name = row[0] if row else None

        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active

        if organization_name:
            worksheet.merge_cells('A4:F4')  # Merge cells for the organization name
            cell = worksheet['A4']  # Get the top-left cell of the merged range
            cell.value = f"Organization Name: {organization_name}"
            cell.font = Font(bold=True,size=12)
            cell.alignment = Alignment(wrap_text=True)  # Align text to the center

        # Add the "Customer Machine Report" label in the middle of the third row
        worksheet.merge_cells('A1:F2')  # Merge cells for the label
        cell = worksheet['A1']  # Get the top-left cell of the merged range
        cell.value = "Payment Machine Report"
        cell.font = Font(bold=True,size=16)  # Make the label bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        worksheet.merge_cells('A3:F3')  # Merge cells for the date range
        date_cell = worksheet['A3']  # Get the top-left cell of the merged range
        date_cell.value = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center
       

        # Add organization name
        # if organization_name:
        #     worksheet.merge_cells('C1:D2')  # Merge cells for the organization name
        #     cell = worksheet['C1']  # Get the top-left cell of the merged range
        #     cell.value = organization_name
        #     cell.font = Font(bold=True)
        #     cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        # # Add the "Refill Machine Report" label
        # worksheet.merge_cells('C3:D3')  # Merge cells for the label
        # cell = worksheet['C3']  # Get the top-left cell of the merged range
        # cell.value = "Payment Machine Report"
        # cell.font = Font(bold=True)  # Make the label bold
        # cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        # Merge cells for "FROM" and "TO" dates
        # worksheet.merge_cells('E1:E3')  # Merge cells for the date range
        # date_range_cell = worksheet['E1']  # Get the top-left cell of the merged range
        # date_range_cell.value = f"FROM: {date_from.strftime('%Y-%m-%d')} TO: {date_to.strftime('%Y-%m-%d')}"
        # date_range_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Align text to the center

        # Add headers for machine data
        headers = ["Sr No","Machine ID", "Coin", "QR", "Total", "Total Amount"]
        for col, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=5, column=col, value=header)
            header_cell.font = Font(bold=True)  # Make headers bold
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            border_style = Side(border_style="thin")
            border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            header_cell.border = border

        # Write machine data to the worksheet
        for row_num, data in enumerate(data_for_response, start=6):
            total_coin = data['coin_mode_counts'] * 10
            total_qr = data['qr_mode_counts'] * 10
            total_amount = total_coin + total_qr
            
            worksheet.cell(row=row_num, column=1, value=row_num - 5).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=2, value=data['m_id']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=3, value=data['coin_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=4, value=data['qr_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=5, value=data['coin_mode_counts'] + data['qr_mode_counts']).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=row_num, column=6, value=total_amount).alignment = Alignment(horizontal='center', vertical='center')
            for col in range(1, 7):
                    cell = worksheet.cell(row=row_num, column=col)
                    border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
                    cell.border = border
        # Adjust column widths based on content length
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            column_letter = chr(64 + column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set a fixed width for the "Sr No" column
        worksheet.column_dimensions['A'].width = 10

        # Set the response headers for file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # Write the Excel workbook to the response
        workbook.save(response)

        return response

    except Exception as e:
        return Response({'success': 0, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)


from django.db.models import Count, DateField
from django.db.models.functions import Trunc


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stock_for_all_machine_refills(request):
    try:
        
        if request.method == 'GET':
            # Fetch all distinct machine IDs
            all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

            machine_data = []

            for machine_id in all_machine_ids:
                # Fetch the last entry for the current machine sorted by id
                last_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()

                if last_entry:
                    last_data = last_entry.stock

                    stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

                    stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
                    if stock_after_refill_qs.exists():
                        stock_after_refill = stock_after_refill_qs.first().stock
                    else:
                        stock_after_refill = None
                    

                    # Fetch the last second entry for the current machine sorted by id
                    last_second_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id')[1:2].first()
                    if last_second_entry:
                        last_second_stock = last_second_entry.stock
                    else:
                        last_second_stock = None

                    
                    if stock_after_refill is not None:
                        refill_quantity = stock_after_refill - last_second_stock
                        
                    else:
                        refill_quantity = None

                    # Fetch refill date from MStatus model
                    refill_date = last_entry.created_at.date() if last_entry.created_at else None

                    # Fetch location from MachineMaster model
                    machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
                    location = machine_master_data.installation_location if machine_master_data else None

                    # Add data for the current machine to the list
                    machine_data.append({
                        "machine_id": machine_id,
                        "stock_before_refill": last_second_stock,
                        "stock_after_refill": stock_after_refill,
                        "refill_quantity": refill_quantity,
                        "stock_capacity_equal_count": stock_capacity_equal_count,
                        'location':location,
                        "refill_date": refill_date,
                        
                    })
                    if last_data ==last_entry.capacity:
                        
                        machine_data.append({
                            "machine_id": machine_id,
                            "stock_before_refill": last_data,
                            "stock_after_refill": 0,
                            "refill_quantity": 0,
                            "stock_capacity_equal_count": stock_capacity_equal_count,
                            'location':location,
                            "refill_date": refill_date,
                            
                        })  
                else:
                    # If no entry exists for the machine, set both stock values to None
                    machine_data.append({
                        "machine_id": None,
                        "stock_before_refill": None,
                        "stock_after_refill": None,
                        "refill_quantity": None,
                        "stock_capacity_equal_count": None,  # No entries, so count is 0
                        'location':None,
                        "refill_date": None,
                        
                    })
                

            return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})

    except Exception as e:
        return Response({'success': 0, 'message': str(e)})
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_stock_for_all_machine_refills(request):
#     try:
#         if request.method == 'GET':
#             # Fetch all distinct machine IDs
#             all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

#             machine_data = []

#             for machine_id in all_machine_ids:
#                 # Fetch all entries for the current machine sorted by id in descending order
#                 entries = MStatus.objects.filter(m_id=machine_id).order_by('-id')

#                 for index, entry in enumerate(entries):
#                     # Fetch the next entry in the list
#                     next_entry = entries[index + 1] if index + 1 < len(entries) else None

#                     # Calculate refill quantity and adjust stock if the next entry exists
#                     if next_entry:
#                         stock_before_refill = entry.stock
#                         stock_after_refill = next_entry.stock
#                         refill_quantity = stock_after_refill - stock_before_refill
#                         after_refill_stock = stock_after_refill + refill_quantity
#                         refill_date = next_entry.created_at.strftime('%Y-%m-%d') if next_entry else None

#                         # Fetch machine details
#                         machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
#                         location = machine_master_data.installation_location if machine_master_data else None
#                         stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

#                         # Add data for the current entry to the list
#                         machine_data.append({
#                             "machine_id": machine_id,
#                             "stock_before_refill": stock_before_refill,
#                             "refill_quantity": refill_quantity,
#                             "stock_after_refill": after_refill_stock,
#                             "refill_date": refill_date,
#                             "location": location,
#                             "stock_capacity_equal_count": stock_capacity_equal_count,
#                         })
#                     else:
#                         # If there's no next entry, consider the current entry as the stock after refill
#                         stock_after_refill = entry.stock

#                         # Fetch machine details
#                         machine_master_data = MachineMaster.objects.filter(machine_id=machine_id).first()
#                         location = machine_master_data.installation_location if machine_master_data else None
#                         stock_capacity_equal_count = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).count()

#                         # Add data for the current entry to the list
#                         machine_data.append({
#                             "machine_id": machine_id,
#                             "stock_before_refill": None,
#                             "refill_quantity": None,
#                             "stock_after_refill": stock_after_refill,
#                             "refill_date": None,
#                             "location": location,
#                             "stock_capacity_equal_count": stock_capacity_equal_count,
#                         })

#             # Sort machine_data list in reverse order based on refill_date, handling None values
#             # machine_data = sorted(machine_data, key=lambda x: x.get('refill_date', ''), reverse=True)

#             return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})

#     except Exception as e:
#         return Response({'success': 0, 'message': str(e)})

# @api_view(['GET'])
# def get_stock_for_all_machine_refills(request):
#     try:
#         if request.method == 'GET':
#             # Fetch all distinct machine IDs
#             all_machine_ids = MStatus.objects.values_list('m_id', flat=True).distinct()

#             machine_data = []

#             for machine_id in all_machine_ids:
#                 # Fetch the last entry for the current machine sorted by id
#                 last_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id').first()

#                 if last_entry:
#                     stock_before_refill = last_entry.stock

#                     # Count instances where stock is equal to capacity for the current machine per day
#                     stock_capacity_equal_count_per_day = MStatus.objects.filter(m_id=machine_id, stock=F('capacity')).annotate(
#                         day=TruncDate('created_at')
#                     ).values('day').annotate(count=Count('id')).order_by('day')

#                     stock_after_refill_qs = MStatus.objects.filter(m_id=machine_id, stock=F('capacity'))
#                     if stock_after_refill_qs.exists():
#                         stock_after_refill = stock_after_refill_qs.first().stock
#                     else:
#                         stock_after_refill = None

#                     if stock_after_refill is not None:
#                         refill_quantity = stock_after_refill - stock_before_refill
#                     else:
#                         refill_quantity = None

#                     # Add data for the current machine to the list
#                     machine_data.append({
#                         "machine_id": machine_id,
#                         "stock_before_refill": stock_before_refill,
#                         "stock_after_refill": stock_after_refill,
#                         "refill_quantity": refill_quantity,
#                         "stock_capacity_equal_count": stock_capacity_equal_count_per_day,
#                     })
#                 else:
#                     # If no entry exists for the machine, set both stock values to None
#                     machine_data.append({
#                         "machine_id": None,
#                         "stock_before_refill": None,
#                         "stock_after_refill": None,
#                         "refill_quantity": None,
#                         "stock_capacity_equal_count": None,  # No entries, so count is 0
#                     })

#             return Response({'success': 1, 'message': 'Data Found', 'result': machine_data})

#     except Exception as e:
#         return Response({'success': 0, 'message': str(e)})
class CustomEmailViewSet(viewsets.ModelViewSet):
    queryset = CustomEmail.objects.all()
    serializer_class = CustomEmailSerializer
    permission_classes = [IsAuthenticated]
    def list(self, request, *args, **kwargs):
        try:
            queryset = CustomEmail.objects.all().order_by('-id')
            serializer = CustomEmailSerializer(queryset, many=True)
            logging.info('List request processed successfully.')
            return Response({'success': 1, 'message': 'Custom Email list', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing list request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, pk,*args, **kwargs):
        try:
            instance = CustomEmail.objects.get(pk=pk)
            serializer = CustomEmailSerializer(instance)
            logging.info('Retrieve request processed successfully.')
            return Response({'success': 1, 'message': 'Custom Email details', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing retrieve request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def create(self, request, *args, **kwargs):
        try:
            serializer = CustomEmailSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            logging.info('Create request processed successfully. Custom Email created.')
            return Response({'success': 1, 'message': 'Custom Email created successfully', 'result': serializer.data}, status=status.HTTP_201_CREATED)
        except Exception as e:
            logging.error(f'Error processing create request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, pk,**kwargs):
        try:
            instance = CustomEmail.objects.get(pk=pk)
            serializer = CustomEmailSerializer(instance, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            logging.info('Update request processed successfully. Custom Email updated.')
            return Response({'success': 1, 'message': 'Custom Email updated successfully', 'result': serializer.data})
        except Exception as e:
            logging.error(f'Error processing update request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, pk,*args, **kwargs):
        try:
            instance = CustomEmail.objects.get(pk=pk)
            instance.delete()
            logging.info('Delete request processed successfully. Custom Email deleted.')
            return Response({'success': 1, 'message': 'Custom Email deleted successfully'})
        except Exception as e:
            logging.error(f'Error processing delete request: {e}')
            return Response({'success': 0, 'message': 'Error processing request'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


# @api_view(['GET'])
# def stock_capacity_equal_count_per_day(request, machine_id):
#     date_from_raw = request.GET.get('from_date')
#     date_to_raw = request.GET.get('to_date')

#     if not date_from_raw or not date_to_raw:
#         return Response({'success': False, 'message': 'Both from_date and to_date are required.'}, status=400)

#     try:
#         date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
#         date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
#     except ValueError:
#         return Response({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

#     queryset = MStatus.objects.filter(
#             m_id=machine_id,
#             created_at__date__range=[date_from, date_to],
#             stock=F('capacity')
#         ).values('m_id','created_at__date','stock').annotate(
#             count_capacity_equals_stock=Count('id')
#         )

#     # Combine the count_capacity_equals_stock and refill dates
#     data = []
#     for entry in queryset:
#         stock_after_refill = entry['stock']
#         last_second_entry = MStatus.objects.filter(m_id=machine_id).order_by('-id')[1:2].first()
#         last_second_stock = last_second_entry.stock if last_second_entry else None
#         refill_quantity = None
#         if stock_after_refill is not None and last_second_stock is not None:
#             # Calculate refill quantity if stock_after_refill and last_second_stock are not None
#             refill_quantity = stock_after_refill - last_second_stock
#         # Append data to the result
#         data.append({
#             'm_id': entry['m_id'],
#             'date': entry['created_at__date'],
#             'stock': entry['stock'],
#             'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
#             'refill_quantity': refill_quantity
#         })

#     return Response({'success': 1,'message':'Data Found' ,'result': data})
from django.db.models import Max

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stock_capacity_equal_count_per_day(request, machine_id):
    date_from_raw = request.GET.get('from_date')
    date_to_raw = request.GET.get('to_date')

    if not date_from_raw or not date_to_raw:
        return Response({'success': False, 'message': 'Both from_date and to_date are required.'}, status=400)

    try:
        date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
    except ValueError:
        return Response({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # Annotate to count instances where capacity equals stock, and retrieve stock value, grouped by date
    queryset = MStatus.objects.filter(
        m_id=machine_id,
        created_at__date__range=[date_from, date_to],
        stock=F('capacity')
    ).values('m_id', 'created_at__date', 'stock').annotate(
        count_capacity_equals_stock=Count('id')
    )

    # Combine the count_capacity_equals_stock, stock value, and refill dates
    data = []
    for entry in queryset:
        stock_after_refill = entry['stock']
        
        # Fetch the stock value for the day before the current date
        stock_before_date = MStatus.objects.filter(
            m_id=machine_id,
            created_at__date__lt=entry['created_at__date']
        ).order_by('-created_at').values_list('stock', flat=True).first()

        refill_quantity = None
        if stock_after_refill is not None and stock_before_date is not None:
            # Calculate refill quantity if stock_after_refill and stock_before_date are not None
            refill_quantity = stock_after_refill - stock_before_date
        
        # Append data to the result
        data.append({
            'm_id': entry['m_id'],
            'date': entry['created_at__date'],
            'stock': entry['stock'],
            'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
            'refill_quantity': refill_quantity,
            'stock_before_date':stock_before_date

        })

    return Response({'success': 1, 'message':'Data Found','result': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_stock_capacity_equal_count_per_day_by_id(request, machine_id):
    # Fetch all entries for the specified machine_id
    queryset = MStatus.objects.filter(
        m_id=machine_id,
        stock=F('capacity')  # Filter where stock capacity equals stock
    ).values('m_id', 'created_at__date', 'stock').annotate(
        count_capacity_equals_stock=Count('id')
    )

    # Combine the count_capacity_equals_stock, stock value, and refill dates
    data = []
    for entry in queryset:
        stock_after_refill = entry['stock']

        stock_before_date = MStatus.objects.filter(
            m_id=machine_id,
            created_at__date__lt=entry['created_at__date']
        ).order_by('-created_at').values_list('stock', flat=True).first()

        refill_quantity = None
        if stock_after_refill is not None and stock_before_date is not None:
            # Calculate refill quantity if stock_after_refill and stock_before_date are not None
            refill_quantity = stock_after_refill - stock_before_date

        # Append data to the result
        data.append({
            'm_id': entry['m_id'],
            'date': entry['created_at__date'],
            'stock': entry['stock'],
            'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
            'refill_quantity': refill_quantity,
            'stock_before_date': stock_before_date
        })

    return Response({'success': 1,'message':'Data Found' ,'result': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customer_stock_capacity_equal_count_per_day_by_id(request, machine_id):

    # Fetch all machines assigned to the current user for the specified machine_id
        assigned_machines = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')
        machine_ids = assigned_machines.values_list('machine__machine_id', flat=True)
        
        # Prepare data to be returned
        data = []
        
        # Iterate over each machine
        for machine_id in machine_ids:
            
            # Fetch data for the specified machine_id
            queryset = MStatus.objects.filter(
                m_id=machine_id,
                stock=F('capacity')  # Filter where stock capacity equals stock
            ).values('m_id', 'created_at__date', 'stock').annotate(
                count_capacity_equals_stock=Count('id')
            )
            
            # Process queryset entries
            for entry in queryset:
                stock_after_refill = entry['stock']

                # Fetch stock before the current date
                stock_before_date = MStatus.objects.filter(
                    m_id=machine_id,
                    created_at__date__lt=entry['created_at__date']
                ).order_by('-created_at').values_list('stock', flat=True).first()

                # Calculate refill quantity
                refill_quantity = None
                if stock_after_refill is not None and stock_before_date is not None:
                    refill_quantity = stock_after_refill - stock_before_date

                # Append data to the result
                data.append({
                    'm_id': entry['m_id'],
                    'date': entry['created_at__date'],
                    'stock': entry['stock'],
                    'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
                    'refill_quantity': refill_quantity,
                    'stock_before_date': stock_before_date
                })
        return Response({'success': 1,'message':'Data Found' ,'result': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def post_customer_stock_capacity_equal_count(request, machine_id):
    date_from_raw = request.GET.get('from_date')
    date_to_raw = request.GET.get('to_date')

    if not date_from_raw or not date_to_raw:
        return Response({'success': False, 'message': 'Both from_date and to_date are required.'}, status=400)

    try:
        date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
    except ValueError:
        return Response({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    assigned_machines = MachineUserMapping.objects.filter(assigned_customer=request.user.id).order_by('-id')
    machine_ids = assigned_machines.values_list('machine__machine_id', flat=True)
    
    data = []
    
    for machine_id in machine_ids:
        queryset = MStatus.objects.filter(
            m_id=machine_id,
            created_at__date__range=[date_from, date_to],
            stock=F('capacity')  # Filter where stock capacity equals stock
        ).values('m_id', 'created_at__date', 'stock').annotate(
            count_capacity_equals_stock=Count('id')
        )
        
        for entry in queryset:
            stock_after_refill = entry['stock']

            # Fetch stock before the current date
            stock_before_date = MStatus.objects.filter(
                m_id=machine_id,
                created_at__date__lt=entry['created_at__date']
            ).order_by('-created_at').values_list('stock', flat=True).first()

            # Calculate refill quantity
            refill_quantity = None
            if stock_after_refill is not None and stock_before_date is not None:
                refill_quantity = stock_after_refill - stock_before_date

            data.append({
                'm_id': entry['m_id'],
                'date': entry['created_at__date'],
                'stock': entry['stock'],
                'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
                'refill_quantity': refill_quantity,
                'stock_before_date': stock_before_date
            })

    return Response({'success': 1, 'message':'Data Found','result': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_stock_capacity_equal_count_per_day_by_id(request, machine_id):

    assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
    machine_ids = assigned_machines.values_list('machine__machine_id', flat=True)
    
    data = []
    
    for machine_id in machine_ids:
        queryset = MStatus.objects.filter(
            m_id=machine_id,
            
            stock=F('capacity')  # Filter where stock capacity equals stock
        ).values('m_id', 'created_at__date', 'stock').annotate(
            count_capacity_equals_stock=Count('id')
        )
        
        for entry in queryset:
            stock_after_refill = entry['stock']

            # Fetch stock before the current date
            stock_before_date = MStatus.objects.filter(
                m_id=machine_id,
                created_at__date__lt=entry['created_at__date']
            ).order_by('-created_at').values_list('stock', flat=True).first()

            # Calculate refill quantity
            refill_quantity = None
            if stock_after_refill is not None and stock_before_date is not None:
                refill_quantity = stock_after_refill - stock_before_date

            data.append({
                'm_id': entry['m_id'],
                'date': entry['created_at__date'],
                'stock': entry['stock'],
                'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
                'refill_quantity': refill_quantity,
                'stock_before_date': stock_before_date
            })

    return Response({'success': 1, 'message':'Data Found','result': data})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def post_user_stock_capacity_equal_count(request, machine_id):
    date_from_raw = request.GET.get('from_date')
    date_to_raw = request.GET.get('to_date')

    if not date_from_raw or not date_to_raw:
        return Response({'success': False, 'message': 'Both from_date and to_date are required.'}, status=400)

    try:
        date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
    except ValueError:
        return Response({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    assigned_machines = MachineUserMapping.objects.filter(assigned_user=request.user.id).order_by('-id')
    machine_ids = assigned_machines.values_list('machine__machine_id', flat=True)
    
    data = []
    
    for machine_id in machine_ids:
        queryset = MStatus.objects.filter(
            m_id=machine_id,
            created_at__date__range=[date_from, date_to],
            stock=F('capacity')  # Filter where stock capacity equals stock
        ).values('m_id', 'created_at__date', 'stock').annotate(
            count_capacity_equals_stock=Count('id')
        )
        
        for entry in queryset:
            stock_after_refill = entry['stock']

            # Fetch stock before the current date
            stock_before_date = MStatus.objects.filter(
                m_id=machine_id,
                created_at__date__lt=entry['created_at__date']
            ).order_by('-created_at').values_list('stock', flat=True).first()

            # Calculate refill quantity
            refill_quantity = None
            if stock_after_refill is not None and stock_before_date is not None:
                refill_quantity = stock_after_refill - stock_before_date

            

            data.append({
                'm_id': entry['m_id'],
                'date': entry['created_at__date'],
                'stock': entry['stock'],
                'count_capacity_equals_stock': entry['count_capacity_equals_stock'],
                'refill_quantity': refill_quantity,
                'stock_before_date': stock_before_date,
                
            })

    return Response({'success': 1, 'message':'Data Found','result': data})

import time
# @api_view(['POST'])
# def create_user(request):
#     start_time = time.time()
#     try:
#         serializer = UserMasterSerializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
        
#         email = request.data.get('email')
#         mobile_no = request.data.get('mobile_no')
#         password = request.data['password']

#         if User.objects.filter(Q(email=email) | Q(mobile_no=mobile_no)).exists():
#             elapsed_time = time.time() - start_time
#             logging.info(f"Time taken for create_user: {elapsed_time} seconds")
#             return Response({'success': 0, 'message': 'Email or mobile_no already exists.'}, status=status.HTTP_400_BAD_REQUEST)

#         user = serializer.save(created_by=request.user.id,commit=False)
#         domain = get_domain(request)
#         send_welcome_email(user, password, domain)

#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for create_user: {elapsed_time} seconds")
#         logging.info(f'Data Created: {serializer.data}')
#         logging.info(f'Welcome email sent to {email}')

#         return Response({'success': 1, 'message': 'Data Created Successfully', 'result': serializer.data}, status=status.HTTP_201_CREATED)
    
#     except Exception as e:
#         elapsed_time = time.time() - start_time
#         logging.error(f'Error creating user: {str(e)}')
#         logging.info(f"Time taken for create_user: {elapsed_time} seconds")
#         return Response({'success': 0, 'message': 'Error creating user'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# def get_domain(request):
#     start_time = time.time()
#     try:
#         domain_name = request.get_host().split(':')[0]  # Get the domain name from the request
#         # Split the hostname by "."
#         parts = domain_name.split(".")
#         print(parts,'parts')
#         # Extract the subdomain (the part before the first dot)
#         if len(parts) > 1:
#             subdomain = parts[0]
#         else:
#             subdomain = None  # No subdomain found

#         print(subdomain,'subdomain')
        
#         domain = Tenant.objects.filter(schema_name=subdomain)
#         print(domain,'domain name')
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for get_domain: {elapsed_time} seconds")
#         return domain
#     except Domain.DoesNotExist:
#         elapsed_time = time.time() - start_time
#         logging.error(f'Error retrieving domain: Domain not found')
#         logging.info(f"Time taken for get_domain: {elapsed_time} seconds")
#         raise
#     except Exception as e:
#         elapsed_time = time.time() - start_time
#         logging.error(f'Error retrieving domain: {str(e)}')
#         logging.info(f"Time taken for get_domain: {elapsed_time} seconds")
#         raise
    
# def send_welcome_email(user, password, domain):
#     start_time = time.time()
#     subject = 'Welcome to iVendMSoft!'
#     try:
#         # Retrieve the Tenant object
#         tenant = domain.first()  # Assuming domain is a queryset, retrieve the first object
#         if tenant is None:
#             logging.error('No Tenant object found.')
#             return  # Exit the function if no Tenant object is found
        
#         # Extract the name of the domain from the Tenant object
#         domain_name = tenant.schema_name  # Assuming 'schema_name' is the attribute storing the domain name
        
#         # Compose the email message
#         message = f"Dear {user.name},\n\nThank you for registering in Vending Machine.\n\nYour Gmail Id is {user.email}\n\nYour password is: {password}\n\nYour Domain name is: {domain_name}\n\nBest regards,\nThe Vending Machine Team"

#         # Send the email
#         send_mail(subject, message, settings.EMAIL_HOST_USER, [user.email], fail_silently=True)
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for send_welcome_email: {elapsed_time} seconds")
#     except Exception as e:
#         logging.error(f'Error sending welcome email: {str(e)}')
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for send_welcome_email: {elapsed_time} seconds")
# from smtplib import SMTPAuthenticationError

# def get_domain(request):
#     start_time = time.time()
#     try:
#         domain_name = request.get_host().split(':')[0]  # Get the domain name from the request
#         # Split the hostname by "."
#         parts = domain_name.split(".")
#         # Extract the subdomain (the part before the first dot)
#         if len(parts) > 1:
#             subdomain = parts[0]
#         else:
#             subdomain = None  # No subdomain found
        
#         domain = Tenant.objects.filter(schema_name=subdomain)
        
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for get_domain: {elapsed_time} seconds")
        
#         return domain
#     except Tenant.DoesNotExist:
#         elapsed_time = time.time() - start_time
#         logging.error(f'Error retrieving domain: Domain not found')
#         logging.info(f"Time taken for get_domain: {elapsed_time} seconds")
#         raise
#     except Exception as e:
#         elapsed_time = time.time() - start_time
#         logging.error(f'Error retrieving domain: {str(e)}')
#         logging.info(f"Time taken for get_domain: {elapsed_time} seconds")
#         raise

# def send_custom_email(subject, message, recipient_list):
#     start_time = time.time()
#     try:
#         config = CustomEmail.objects.first()
#         if config:
#             from_email = config.email_host_user
#             backend = EmailBackend(
#                 host=config.email_host,
#                 port=config.email_port,
#                 username=config.email_host_user,
#                 password=config.email_host_password,
#                 use_tls=config.email_use_tls,
#                 fail_silently=config.email_fail_silently
#             )
#         else:
#             from_email = settings.EMAIL_HOST_USER
#             backend = None
        
#         send_mail(subject, message, from_email, recipient_list, connection=backend)
        
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for send_custom_email: {elapsed_time} seconds")
#         return True
#     except SMTPAuthenticationError as e:
#         logging.error(f'Error sending custom email: SMTP authentication failed - {e}')
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for send_custom_email: {elapsed_time} seconds")
#         return False
#     except Exception as e:
#         logging.error(f'Error sending custom email: {e}')
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for send_custom_email: {elapsed_time} seconds")
#         return False

# def send_welcome_email(user, password, domain):
#     start_time = time.time()
#     subject = 'Welcome to iVendMSoft!'
#     try:
#         tenant = domain.first()
#         if tenant is None:
#             logging.error('No Tenant object found.')
#             return
        
#         domain_name = tenant.schema_name
        
#         message = f"Dear {user.name},\n\nThank you for registering in Vending Machine.\n\nYour Gmail Id is {user.email}\n\nYour password is: {password}\n\nYour Domain name is: {domain_name}\n\nBest regards,\nThe Vending Machine Team"

#         success = send_custom_email(subject, message, [user.email])
#         if success:
#             elapsed_time = time.time() - start_time
#             logging.info(f"Time taken for send_welcome_email: {elapsed_time} seconds")
#         else:
#             elapsed_time = time.time() - start_time
#             logging.error(f"Failed to send welcome email.")
#             logging.info(f"Time taken for send_welcome_email: {elapsed_time} seconds")
        
#     except Exception as e:
#         logging.error(f'Error sending welcome email: {str(e)}')
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for send_welcome_email: {elapsed_time} seconds")

# @api_view(['POST'])
# def create_user(request):
#     start_time = time.time()
#     try:
#         serializer = UserMasterSerializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
        
#         email = request.data.get('email')
#         mobile_no = request.data.get('mobile_no')
#         password = request.data['password']

#         if User.objects.filter(Q(email=email) | Q(mobile_no=mobile_no)).exists():
#             elapsed_time = time.time() - start_time
#             logging.info(f"Time taken for create_user: {elapsed_time} seconds")
#             return Response({'success': 0, 'message': 'Email or mobile_no already exists.'}, status=status.HTTP_400_BAD_REQUEST)

#         user = serializer.save(created_by=request.user.id)
        
#         domain = get_domain(request)
#         send_welcome_email(user, password, domain)
        
#         user.save()
        
#         elapsed_time = time.time() - start_time
#         logging.info(f"Time taken for create_user: {elapsed_time} seconds")
#         logging.info(f'Data Created: {serializer.data}')
#         logging.info(f'Welcome email sent to {email}')

#         return Response({'success': 1, 'message': 'Data Created Successfully', 'result': serializer.data}, status=status.HTTP_201_CREATED)
    
#     except serializers.ValidationError as e:
#         elapsed_time = time.time() - start_time
#         logging.error(f'Validation error creating user: {str(e)}')
#         logging.info(f"Time taken for create_user: {elapsed_time} seconds")
#         return Response({'success': 0, 'message': 'Validation error', 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

#     except Exception as e:
#         elapsed_time = time.time() - start_time
#         logging.error(f'Error creating user: {str(e)}')
#         logging.info(f"Time taken for create_user: {elapsed_time} seconds")
#         return Response({'success': 0, 'message': 'Error creating user'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
from rest_framework.views import APIView

class CustomAPIRoot(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return Response({
            'message': 'Welcome to the API'
        })
class EmptyRootView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        return Response({"Go Back Don't Be Too Smart."})